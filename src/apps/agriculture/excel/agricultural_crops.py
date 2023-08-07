from datetime import datetime
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from ..models import TreeTypes
from .db_service import AgriculturalDBQuery
from .utils import write_header_title
from ...accounts.models import Region


class AgriculturalCropsSheet(AgriculturalDBQuery):
    def __init__(self, user=None, start=None, end=None):
        self.user = user
        self.start_date = start
        self.end_date = end
        self.end_col = None
        self.align = Alignment(wrapText=True, horizontal='center', vertical='center')

    def generate_agricultural_product_excel_report(self):
        now = datetime.now()
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f"attachment; filename=qishloq_ekinlarini_ekish-{now.strftime('%d-%m-%Y')}.xlsx"
        wb = Workbook()
        sheet = wb.active
        align = Alignment(wrapText=True, horizontal='center', vertical='center')
        end_merge_column = f'A1:Z1'
        sheet.merge_cells(end_merge_column)
        sheet.column_dimensions['A'].width = 5
        sheet.column_dimensions['B'].width = 55
        sheet.row_dimensions[1].height = 45
        sheet.column_dimensions['C'].width = 14
        if self.start_date[:4] == self.end_date[:4]:
            year = f'{self.start_date[:4]}'
        else:
            year = f'{self.start_date[:4]}-{self.end_date[:4]}'

        title = sheet['A1']
        title.value = f"Ўрмон хўжалиги давлат қўмитаси тизимидаги ўрмон хўжаликларида " \
                      f"{year} йилда экиладиган қишлоқ хўжалиги " \
                      f"экинларини экиш ва махсулотларини " \
                      f"етиштириш тўғрисида тезкор маълумот"
        title.alignment = align
        title.font = Font(name='Times New Roman', bold=True, size=15)
        sheet.merge_cells('A2:A4')
        sheet.merge_cells('B2:B4')
        sheet.merge_cells('C2:C4')
        sheet.row_dimensions[2].height = 40
        sheet.row_dimensions[3].height = 60
        sheet.row_dimensions[4].height = 35
        write_header_title(sheet=sheet, row=2, col=1, size=12, is_bold=True, is_italic=False, value='№')
        write_header_title(sheet=sheet, row=2, col=2, size=15, is_bold=True, is_italic=False, value='Хўжалик номи')
        write_header_title(sheet=sheet, row=2, col=3, size=12, is_bold=True, is_italic=False, value='Курсаткич-лар')
        sheet.merge_cells('D2:F3')
        write_header_title(sheet=sheet, row=2, col=4, size=12, is_bold=True, is_italic=True, value='Жами қишлоқ хўжалиги')
        write_header_title(sheet=sheet, row=4, col=4, size=11, is_bold=True, is_italic=True, value='га')
        write_header_title(sheet=sheet, row=4, col=5, size=11, is_bold=True, is_italic=True, value='ц/га')
        write_header_title(sheet=sheet, row=4, col=6, size=11, is_bold=True, is_italic=True, value='тн')
        sheet.merge_cells('G2:G4')
        government_title = sheet.cell(column=7, row=2, value='Давлатга топширилган буғдой тонна')
        government_title.alignment = Alignment(textRotation=90, wrapText=True, horizontal='center', vertical='center')
        government_title.font = Font(name='Times New Roman', italic=True, bold=True, size=13)
        sheet.merge_cells('H2:Z2')
        write_header_title(sheet=sheet, row=2, col=8, size=13, is_bold=True, is_italic=True, value='Шу жумладан')
        products = TreeTypes.objects.filter(status=1).order_by('sort')
        col_index = 8
        for type_value in products:
            sheet.merge_cells(f'{get_column_letter(col_index)}3:{get_column_letter(col_index + 2)}3')
            tree_type_name = sheet.cell(row=3, column=col_index, value=type_value.name)
            tree_type_name.alignment = align
            tree_type_name.font = Font(name='Times New Roman', bold=True, italic=True, size=13)
            write_header_title(sheet=sheet, row=4, col=col_index, size=12, is_bold=True, is_italic=True, value='га')
            write_header_title(sheet=sheet, row=4, col=col_index + 1, size=12, is_bold=True, is_italic=True, value='ц/га')
            write_header_title(sheet=sheet, row=4, col=col_index + 2, size=12, is_bold=True, is_italic=True, value='тн')
            col_index += 3

        agro_data = self.get_agricultural_db_data(user=self.user, qs=products, start=self.start_date, end=self.end_date)
        row_index = 5
        region_row = []
        article_id = 0
        region_id = None

        for index, value in enumerate(agro_data):
            # print("1......", value)
            # print("2......", value['act'])
            # print("***************************")
            if region_id is None or value['region_id'] != region_id:
                article_id = 0
                region_row.append(row_index)
                region_id = self.write_region_data(
                    sheet=sheet, row=row_index, region_id=value['region_id'],
                    count=value['department_count'], data=value['plan'])
                row_index += 3
            article_id += 1
            sheet.merge_cells(f'A{row_index}:A{row_index + 2}')
            sheet.merge_cells(f'B{row_index}:B{row_index + 2}')
            department_id = sheet.cell(column=1, row=row_index, value=article_id)
            department_id.font = Font(name='Times New Roman', size=12)
            department_id.alignment = align
            if len(value['department']) > 50:
                name = f" {value['department'][:45]}..."
            else:
                name = f" {value['department']}"
            department_name = sheet.cell(column=2, row=row_index, value=name)
            department_name.alignment = Alignment(wrapText=True, vertical='center')
            department_name.font = Font(name='Times New Roman', size=13)
            sheet.cell(row=row_index, column=3, value='Топшириқ')
            sheet.cell(row=row_index + 1, column=3, value='Амалда')
            sheet.cell(row=row_index + 2, column=3, value='Фоиз')
            col = 8
            length = len(value['plan']) * 3
            self.end_col = length
            formula = self.build_write_formula(length=length, row=row_index)
            # print("F......", formula)
            # plan_total_hectare = '='
            # plan_total_ton = '='
            # act_total_hectare = '='
            # act_total_ton = '='
            # for zero in range(len(value['plan']) * 3):
            #     coordinate = col + zero
            #     if coordinate % 3 == 0:
            #         plan_total_hectare += f'+{get_column_letter(coordinate - 1)}{row_index}'
            #         plan_total_ton += f'+{get_column_letter(coordinate + 1)}{row_index}'
            #         act_total_hectare += f'+{get_column_letter(coordinate - 1)}{row_index + 1}'
            #         act_total_ton += f'+{get_column_letter(coordinate + 1)}{row_index + 1}'

            self.write_plan_act_value(sheet=sheet, row=row_index, col=col, data=value['plan'])
            self.write_plan_act_value(sheet=sheet, row=row_index + 1, col=col, data=value['act'])
            self.calc_percent_value(sheet=sheet, row=row_index + 2, col=col, length=len(value['act']))
            self.write_d_e_f(
                sheet=sheet, row=row_index,
                hectare_formula=formula['plan_total_hectare'], ton_formula=formula['plan_total_ton'])
            self.write_d_e_f(
                sheet=sheet, row=row_index + 1,
                hectare_formula=formula['act_total_hectare'], ton_formula=formula['act_total_ton'])
            self.calc_total_percent(sheet=sheet, row=row_index + 2, size=10)

            row_index += 3
        sheet.merge_cells(f'A{row_index}:A{row_index + 2}')
        sheet.merge_cells(f'B{row_index}:B{row_index + 2}')

        republic_title = sheet.cell(row=row_index, column=2, value="Республика бўйича жами")
        republic_title.alignment = align
        republic_title.font = Font(name='Times New Roman', size=13, bold=True)

        end_plan_col = sheet.cell(row=row_index, column=3, value='Топшириқ')
        end_plan_col.alignment = align
        end_plan_col.font = Font(name='Times New Roman', bold=True, size=12)

        end_act_col = sheet.cell(row=row_index + 1, column=3, value='Амалда')
        end_act_col.alignment = align
        end_act_col.font = Font(name='Times New Roman', bold=True, size=12)

        end_percent_col = sheet.cell(row=row_index + 2, column=3, value='Фоиз')
        end_percent_col.alignment = align
        end_percent_col.font = Font(name='Times New Roman', bold=True, size=12)

        col = 8
        length = self.end_col
        for zero in range(length):
            coordinate = col + zero
            if not coordinate % 3 == 0:
                plan_all_region = '='
                act_all_region = '='
                for row in region_row:
                    plan_all_region += f'+{get_column_letter(coordinate)}{row}'
                    act_all_region += f'+{get_column_letter(coordinate)}{row + 1}'

                region_hectare = sheet.cell(row=row_index, column=coordinate, value=plan_all_region)
                region_hectare.alignment = align
                region_hectare.font = Font(name='Times New Roman', bold=True, size=12)

                region_act = sheet.cell(row=row_index + 1, column=coordinate, value=act_all_region)
                region_act.alignment = align
                region_act.font = Font(name='Times New Roman', bold=True, size=12)
            else:
                begin = get_column_letter(coordinate - 1)
                end = get_column_letter(coordinate + 1)
                region_plan_cent = sheet.cell(
                    row=row_index, column=coordinate,
                    value=f'=IF(OR({begin}{row_index}={0}, {end}{row_index}={0}),{0},{end}{row_index}*{10}/{begin}{row_index})')
                region_plan_cent.alignment = self.align
                region_plan_cent.font = Font(name='Times New Roman', bold=True, size=12)
                region_plan_cent.number_format = '0.0'

                region_act_cent = sheet.cell(
                    row=row_index + 1, column=coordinate,
                    value=f'=IF(OR({begin}{row_index + 1}={0}, {end}{row_index + 1}={0}),{0},{end}{row_index + 1}*{10}/{begin}{row_index + 1})')
                region_act_cent.alignment = self.align
                region_act_cent.font = Font(name='Times New Roman', bold=True, size=12)
                region_act_cent.number_format = '0.0'

        row = row_index + 2
        for x in range(length):
            col_1 = get_column_letter(col)
            hectare_percent = sheet.cell(
                row=row, column=col,
                value=f'=IF(OR({col_1}{row - 1}={0}, {col_1}{row - 2}={0}),{0},{col_1}{row - 1}*{100}/{col_1}{row - 2})')
            hectare_percent.alignment = self.align
            hectare_percent.font = Font(name='Times New Roman', bold=True, size=12)
            hectare_percent.number_format = '0.0'
            col += 1

        total_formula = self.build_write_formula(length=length, row=row_index)
        # *************** PLAN ************
        republic_total_plan_h = sheet.cell(row=row_index, column=4, value=total_formula['plan_total_hectare'])
        republic_total_plan_h.alignment = align
        republic_total_plan_h.font = Font(name='Times New Roman', bold=True, size=12)

        republic_total_plan_t = sheet.cell(row=row_index, column=6, value=total_formula['plan_total_ton'])
        republic_total_plan_t.alignment = align
        republic_total_plan_t.font = Font(name='Times New Roman', bold=True, size=12)

        republic_total_plan_c = sheet.cell(
            row=row_index, column=5, value=f'=IF(OR(D{row_index}={0}, F{row_index}={0}),{0},F{row_index}*{10}/D{row_index})')
        republic_total_plan_c.alignment = self.align
        republic_total_plan_c.font = Font(name='Times New Roman', bold=True, size=12)
        republic_total_plan_c.number_format = '0.0'
        # *************** ACTUAL ************
        republic_total_act_h = sheet.cell(row=row_index + 1, column=4, value=total_formula['act_total_hectare'])
        republic_total_act_h.alignment = align
        republic_total_act_h.font = Font(name='Times New Roman', bold=True, size=12)

        republic_total_act_t = sheet.cell(row=row_index + 1, column=6, value=total_formula['act_total_ton'])
        republic_total_act_t.alignment = align
        republic_total_act_t.font = Font(name='Times New Roman', bold=True, size=12)

        republic_total_act_c = sheet.cell(
            row=row_index + 1, column=5,
            value=f'=IF(OR(D{row_index + 1}={0}, F{row_index + 1}={0}),{0},F{row_index + 1}*{10}/D{row_index + 1})')
        republic_total_act_c.alignment = self.align
        republic_total_act_c.font = Font(name='Times New Roman', bold=True, size=12)
        republic_total_act_c.number_format = '0.0'
        # *************** PERCENT ************
        col = 4
        for x in range(3):
            letter = get_column_letter(col)
            _f = f'=IF(OR({letter}{row - 1}={0}, {letter}{row - 2}={0}),{0},{letter}{row - 1}*{100}/{letter}{row - 2})'
            hectare_percent = sheet.cell(row=row, column=col, value=_f)
            hectare_percent.alignment = self.align
            hectare_percent.font = Font(name='Times New Roman', bold=True, size=12)
            hectare_percent.number_format = '0.0'
            col += 1

        sheet.freeze_panes = 'A5'
        wb.save(response)
        return response

    def write_region_data(self, sheet, row, region_id, count, data):
        qs_region = Region.objects.filter(id=region_id, status=1)
        col = 8
        length = len(data) * 3
        formula = self.build_write_formula(length=length, row=row)
        plan_column_c = sheet.cell(row=row, column=4, value=formula['plan_total_hectare'])
        plan_column_c.alignment = self.align
        plan_column_c.font = Font(name='Times New Roman', bold=True)

        plan_column_f = sheet.cell(row=row, column=6, value=formula['plan_total_ton'])
        plan_column_f.alignment = self.align
        plan_column_f.font = Font(name='Times New Roman', bold=True)

        plan_column_e = sheet.cell(row=row, column=5, value=f'=IF(OR(D{row}={0}, F{row}={0}),{0},F{row}*{10}/D{row})')
        plan_column_e.alignment = self.align
        plan_column_e.font = Font(name='Times New Roman', bold=True)
        plan_column_e.number_format = '0.0'

        act_column_c = sheet.cell(row=row + 1, column=4, value=formula['act_total_hectare'])
        act_column_c.alignment = self.align
        act_column_c.font = Font(name='Times New Roman', bold=True)

        act_column_f = sheet.cell(row=row + 1, column=6, value=formula['act_total_ton'])
        act_column_f.alignment = self.align
        act_column_f.font = Font(name='Times New Roman', bold=True)

        act_column_e = sheet.cell(row=row + 1, column=5, value=f'=IF(OR(D{row+1}={0}, F{row+1}={0}),{0},F{row+1}*{10}/D{row+1})')
        act_column_e.alignment = self.align
        act_column_e.font = Font(name='Times New Roman', bold=True)
        act_column_e.number_format = '0.0'

        self.calc_total_percent(sheet=sheet, row=row + 2, size=11)

        self.calc_plan_act_by_region(sheet=sheet, row=row, col=col, count=count, data=data)
        self.calc_plan_act_by_region(sheet=sheet, row=row + 1, col=col, count=count, data=data)
        self.calc_percent_by_region(sheet=sheet, row=row + 2, col=col, length=len(data))

        if qs_region.exists():
            qs_region_id = qs_region.first().pk
            sheet.merge_cells(f'A{row}:A{row + 2}')
            sheet.merge_cells(f'B{row}:B{row + 2}')

            region_name = sheet.cell(column=2, row=row, value=qs_region.first().name)
            region_name.alignment = self.align
            region_name.font = Font(name='Times New Roman', bold=True, size=15)
            write_header_title(sheet=sheet, row=row, col=3, size=12, is_bold=True, is_italic=False, value='Топшириқ')
            write_header_title(sheet=sheet, row=row + 1, col=3, size=12, is_bold=True, is_italic=False, value='Амалда')
            write_header_title(sheet=sheet, row=row + 2, col=3, size=12, is_bold=True, is_italic=False, value='Фоиз')
            return qs_region_id

    def write_d_e_f(self, sheet, row, hectare_formula, ton_formula):
        column_d = sheet.cell(row=row, column=4, value=hectare_formula)
        column_d.alignment = self.align
        column_d.font = Font(name='Times New Roman', size=10, bold=True)

        column_f = sheet.cell(row=row, column=6, value=ton_formula)
        column_f.alignment = self.align
        column_f.font = Font(name='Times New Roman', size=10, bold=True)

        column_e = sheet.cell(row=row, column=5, value=f'=IF(OR(D{row}={0}, F{row}={0}),{0},F{row}*{10}/D{row})')
        column_e.alignment = self.align
        column_e.font = Font(name='Times New Roman', size=10, bold=True)
        column_e.number_format = '0.0'

    def calc_total_percent(self, sheet, row, size):
        # =IF(OR(A6=0, A5=0),0,A6*100/A5)
        column_d = sheet.cell(
            row=row, column=4,
            value=f'=IF(OR(D{row - 1}={0}, D{row - 2}={0}),{0},D{row - 1}*{100}/D{row - 2})')
        column_d.alignment = self.align
        column_d.font = Font(name='Times New Roman', size=size, bold=True)
        column_d.number_format = '0.0'

        column_e = sheet.cell(
            row=row, column=5,
            value=f'=IF(OR(E{row - 1}={0}, E{row - 2}={0}),{0},E{row - 1}*{100}/E{row - 2})')
        column_e.alignment = self.align
        column_e.font = Font(name='Times New Roman', size=size, bold=True)
        column_e.number_format = '0.0'

        column_f = sheet.cell(
            row=row, column=6,
            value=f'=IF(OR(F{row - 1}={0}, F{row - 2}={0}),{0},F{row - 1}*{100}/F{row - 2})')
        column_f.alignment = self.align
        column_f.font = Font(name='Times New Roman', size=size, bold=True)
        column_f.number_format = '0.0'

    def write_plan_act_value(self, sheet, row, col, data):
        for _, item in enumerate(data):
            sheet.cell(row=row, column=col, value=item[0]).alignment = self.align
            sheet.cell(row=row, column=col + 2, value=item[1]).alignment = self.align
            col_1 = get_column_letter(col)
            col_3 = get_column_letter(col + 2)
            # ex: J5*10/H5
            cent = sheet.cell(
                row=row, column=col + 1,
                value=f'=IF(OR({col_1}{row}={0}, {col_3}{row}={0}),{0},{col_3}{row}*{10}/{col_1}{row})')
            cent.alignment = self.align
            cent.number_format = '0.0'
            col += 3

    def build_write_formula(self, length, row):
        col = 8
        plan_total_hectare = '='
        plan_total_ton = '='
        act_total_hectare = '='
        act_total_ton = '='
        for zero in range(length):
            coordinate = col + zero
            if coordinate % 3 == 0:
                plan_total_hectare += f'+{get_column_letter(coordinate - 1)}{row}'
                plan_total_ton += f'+{get_column_letter(coordinate + 1)}{row}'
                act_total_hectare += f'+{get_column_letter(coordinate - 1)}{row + 1}'
                act_total_ton += f'+{get_column_letter(coordinate + 1)}{row + 1}'
        return dict(
            plan_total_hectare=plan_total_hectare,
            plan_total_ton=plan_total_ton,
            act_total_hectare=act_total_hectare,
            act_total_ton=act_total_ton
        )

    def calc_percent_value(self, sheet, row, col, length):
        for x in range(length * 3):
            col_1 = get_column_letter(col)
            hectare_percent = sheet.cell(
                row=row, column=col,
                value=f'=IF(OR({col_1}{row - 1}={0}, {col_1}{row - 2}={0}),{0},{col_1}{row - 1}*{100}/{col_1}{row - 2})')
            hectare_percent.alignment = self.align
            hectare_percent.number_format = '0.0'
            col += 1

    def calc_percent_by_region(self, sheet, row, col, length):
        for x in range(length):
            col_1 = get_column_letter(col)
            hectare_percent = sheet.cell(
                row=row, column=col,
                value=f'=IF(OR({col_1}{row - 1}={0}, {col_1}{row - 2}={0}),{0},{col_1}{row - 1}*{100}/{col_1}{row - 2})')
            hectare_percent.alignment = self.align
            hectare_percent.font = Font(name='Times New Roman', bold=True)
            hectare_percent.number_format = '0.0'
            col_2 = get_column_letter(col + 1)
            centner_percent = sheet.cell(
                row=row, column=col + 1,
                value=f'=IF(OR({col_2}{row - 1}={0}, {col_2}{row - 2}={0}),{0},{col_2}{row - 1}*{100}/{col_2}{row - 2})')
            centner_percent.alignment = self.align
            centner_percent.font = Font(name='Times New Roman', bold=True)
            centner_percent.number_format = '0.0'
            col_3 = get_column_letter(col + 2)
            ton_percent = sheet.cell(
                row=row, column=col + 2,
                value=f'=IF(OR({col_3}{row - 1}={0}, {col_3}{row - 2}={0}),{0},{col_3}{row - 1}*{100}/{col_3}{row - 2})')
            ton_percent.alignment = self.align
            ton_percent.font = Font(name='Times New Roman', bold=True)
            ton_percent.number_format = '0.0'
            col += 3

    def calc_plan_act_by_region(self, sheet, row, col, count, data):
        for _, item in enumerate(data):
            hectare_formula = '='
            for x in range(1, count * 3 + 1):
                if x % 3 == 0:
                    hectare_formula += f'+{get_column_letter(col)}{x + row}'
            hectare = sheet.cell(column=col, row=row, value=hectare_formula)
            hectare.alignment = self.align
            hectare.font = Font(name='Times New Roman', bold=True)

            ton_formula = '='
            for x in range(1, count * 3 + 1):
                if x % 3 == 0:
                    ton_formula += f'+{get_column_letter(col + 2)}{x + row}'
            ton = sheet.cell(column=col + 2, row=row, value=ton_formula)
            ton.alignment = self.align
            ton.font = Font(name='Times New Roman', bold=True)

            col_1 = get_column_letter(col)
            col_3 = get_column_letter(col + 2)
            centner = sheet.cell(
                row=row, column=col + 1,
                value=f'=IF(OR({col_1}{row}={0}, {col_3}{row}={0}),{0},{col_3}{row}*{10}/{col_1}{row})')
            centner.alignment = self.align
            centner.font = Font(name='Times New Roman', bold=True)
            centner.number_format = '0.0'
            col += 3
