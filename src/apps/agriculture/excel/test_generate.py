from datetime import datetime
from django.db import connection
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter, column_index_from_string
from ...accounts.models import Region, UserDepartment


class TestSheet:

    def generate_excel(self, user):
        now = datetime.now()
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f"attachment; filename=test-{now.strftime('%d-%m-%Y')}.xlsx"
        wb = Workbook()
        sheet = wb.active
        sheet.column_dimensions['D'].width = 80
        sheet.cell(column=3, row=2, value='№')

        dep_title = sheet.cell(column=4, row=2, value='Departments')
        dep_title.alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
        dep_title.font = Font(name='Times New Roman', bold=True, size=12)
        qs_user_department = UserDepartment.objects.filter(user=user)
        if qs_user_department.exists():
            departments = qs_user_department[0].departments.filter(status=1).order_by('sort')
        else:
            departments = []
        region_id = None
        row_index = 3
        article_id = 0
        for index, department_value in enumerate(departments):
            # if region_id is None:
            #     region_id = self.write_region(sheet=sheet, row=row_index, region_pk=department_value.region.pk)
            #     row_index += 1
            # else:
            #     if department_value.region.pk != region_id:
            #         article_id = 0
            #         region_id = self.write_region(sheet=sheet, row=row_index, region_pk=department_value.region.pk)
            #         row_index += 1
            article_id += 1
            sheet.cell(column=3, row=row_index, value=article_id)
            sheet.cell(column=4, row=row_index, value=department_value.name)
            row_index += 1

        def set_border(ws, cell_range):
            thin = Side(border_style="thin", color="000000")
            for row in ws[cell_range]:
                for cell in row:
                    cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
        set_border(sheet, f'A1:{get_column_letter(sheet.max_column)}{row_index}')
        wb.save(response)
        return response

    # def write_region_data(self, sheet, row, region_pk):
    #     qs_region = Region.objects.filter(pk=region_pk)
    #     if qs_region.exists():
    #         region_id = qs_region.first().pk
    #         cell_region = sheet.cell(column=4, row=row, value=qs_region.first().name)
    #         cell_region.alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
    #         cell_region.font = Font(name='Times New Roman', bold=True, size=12)
    #         return region_id
