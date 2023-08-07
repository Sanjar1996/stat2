from datetime import datetime
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from ..models import ReportKey, ReportValue
from django.db.models import Sum, Count
from ...accounts.models import Region, Department, UserDepartment
from ...trees.excel.utils import write_republic_data, set_border


class DynamicExcelReport:
    def __init__(self, user=None, obj=None, start=None, end=None):
        self.user = user
        self.start_date = start
        self.end_date = end
        self.report = obj
        self.align = Alignment(wrapText=True, horizontal='center', vertical='center')

    def generate_dynamic_excel_report(self):
        now = datetime.now()
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f"attachment; filename=dynamic-{now.strftime('%d-%m-%Y')}.xlsx"
        wb = Workbook()
        sheet = wb.active
        align = Alignment(wrapText=True, horizontal='center', vertical='center')
        sheet.column_dimensions['A'].width = 5
        sheet.column_dimensions['B'].width = 50
        title_value = self.report.name
        if len(title_value) > 150:
            sheet.merge_cells('A1:AN1')
            sheet.row_dimensions[1].height = 60
        else:
            sheet.merge_cells('A1:T1')
            sheet.row_dimensions[1].height = 50
        title = sheet.cell(column=1, row=1, value=f"{self.end_date} {title_value}")
        title.font = Font(name='Times New Roman', bold=True, size=13)
        title.alignment = align
        sheet.merge_cells("C2:E2")
        currency = sheet.cell(column=3, row=2, value=f"{self.report.description}")
        currency.alignment = align

        sort_tick = sheet.cell(row=3, column=1, value='№')
        sort_tick.alignment = self.align
        sort_tick.font = Font(name='Times New Roman', size=16)

        first_groups = ReportKey.objects.values(
            'report_group_one_id', 'report_group_one__name').filter(
            report_id=self.report.id, status=1).annotate(
            total=Count('id')).order_by('report_group_one_id', 'report_group_one__name')
        sheet.row_dimensions[5].height = 60
        current_column = 3
        max_length_first_title = 0
        report_keys = []
        for first_group in first_groups:
            length = first_group['report_group_one__name']
            if int(len(length)) > int(max_length_first_title):
                max_length_first_title = int(len(length))
            first_group_title = sheet.cell(row=3, column=current_column, value=first_group['report_group_one__name'])
            first_group_title.font = Font(name='Times New Roman', bold=True, size=12)
            first_group_title.alignment = self.align
            begin = current_column
            end = current_column + first_group['total']
            # print("TOTAL...", first_group['total'])
            coordinate = f'{get_column_letter(begin)}{3}:{get_column_letter(end - 1)}{3}'
            sheet.merge_cells(coordinate)
            # print("CC.....", first_group['report_group_one_id'])
            second_groups = ReportKey.objects.values(
                'report_group_one_id', 'report_group_two_id', 'report_group_two__name'
            ).filter(
                report_id=self.report.id, report_group_one_id=first_group['report_group_one_id'],
                status=1).annotate(total=Count('id'))
            # print("second_groups...", second_groups, type(second_groups))
            second_group_col = current_column
            for second_group in second_groups:
                # print("???......", second_group)
                if second_group['report_group_two_id'] is None and second_group['total']:
                    keys = ReportKey.objects.filter(
                        report_group_one_id=second_group['report_group_one_id'], report_group_two_id=None)
                    for key_id, key in enumerate(keys):
                        report_keys.append(key.id)
                        # print("???......0", key.name)
                        cell_key_title = sheet.cell(row=4, column=second_group_col + key_id, value=key.name)
                        cell_key_title.font = Font(name='Times New Roman', size=10)
                        cell_key_title.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
                    for subtotal in range(int(second_group['total'])):
                        sheet.merge_cells(
                            f'{get_column_letter(second_group_col)}{4}:{get_column_letter(second_group_col)}{5}')
                        second_group_col += 1
                else:
                    keys = ReportKey.objects.filter(
                        report_group_one_id=second_group['report_group_one_id'],
                        report_group_two_id=second_group['report_group_two_id'])
                    for key_id, key in enumerate(keys):
                        report_keys.append(key.id)
                        cell_key_title = sheet.cell(row=5, column=second_group_col + key_id, value=key.name)
                        cell_key_title.font = Font(name='Times New Roman', size=10)
                        cell_key_title.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
                    second_group_begin = second_group_col
                    second_group_end = second_group_col + second_group['total']
                    sheet.merge_cells(
                        f'{get_column_letter(second_group_begin)}{4}:{get_column_letter(second_group_end - 1)}{4}')
                    second_group_title = sheet.cell(
                        row=4, column=second_group_col, value=second_group['report_group_two__name'])
                    second_group_title.font = Font(name='Times New Roman', bold=True, size=11)
                    second_group_title.alignment = align
                second_group_col += second_group['total']
            current_column += first_group['total']
        sheet.merge_cells("B2:B5")
        department_title = sheet.cell(column=2, row=2, value='Хўжаликлар номи')
        department_title.alignment = align
        department_title.font = Font(name='Times New Roman', bold=True, size=20)
        if max_length_first_title > 100:
            sheet.row_dimensions[3].height = 90
        else:
            sheet.row_dimensions[3].height = 60
        # print("report_keys......[]", report_keys)
        row_index = 6
        region_id = None
        qs_user_department = UserDepartment.objects.filter(user=self.user)
        if qs_user_department.exists():
            departments = qs_user_department[0].departments.filter(status=1).order_by('sort')
        else:
            departments = []
        region_row = []
        article_id = 0
        start_col = 3
        for index, value in enumerate(departments):
            if region_id is None or value.region.id != region_id:
                article_id = 0
                region_row.append(row_index)
                region_id = self.write_region_data(
                    sheet=sheet, row=row_index, col=start_col, region_id=value.region.id, length=len(report_keys))
                row_index += 1
            article_id += 1
            department_id = sheet.cell(row=row_index, column=1, value=article_id)
            department_id.alignment = align
            if len(value.name) > 53:
                name = f'{value.name[:45]}...'
            else:
                name = value.name
            department_name = sheet.cell(row=row_index, column=2, value=name)
            department_name.font = Font(name='Times New Roman', size=11)
            for numerate, report_key_id in enumerate(report_keys):
                report_value = ReportValue.objects.filter(
                    department=value, report_key_id=report_key_id,
                    date__gte=self.start_date, date__lt=self.end_date, status=1).aggregate(Sum('double'))
                final_value = report_value['double__sum'] if report_value['double__sum'] else 0
                report_value_cell = sheet.cell(row=row_index, column=start_col + numerate, value=final_value)
                report_value_cell.alignment = align
            row_index += 1

        write_republic_data(
            sheet=sheet, row=row_index, rep_title='Республика бўйича жами',
            begin=start_col, end=start_col + len(report_keys), array_row=region_row)
        set_border(sheet=sheet, cell_range=f'A1:{get_column_letter(sheet.max_column)}{row_index}')
        sheet.freeze_panes = 'A6'
        wb.save(response)
        return response

    def write_region_data(self, sheet, row, col, region_id, length):
        qs_region = Region.objects.filter(id=region_id, status=1)
        count = Department.objects.filter(region_id=region_id, status=1).count()
        if qs_region.exists():
            begin_row = row + 1
            end_row = row + count
            region_id = qs_region.first().pk
            region_name = sheet.cell(column=2, row=row, value=qs_region.first().name)
            region_name.alignment = self.align
            region_name.font = Font(name='Times New Roman', bold=True, size=12)
            for key in range(length):
                column = col + key
                formula = f'=SUM({get_column_letter(column)}{begin_row}:{get_column_letter(column)}{end_row})'
                each_region_cell = sheet.cell(row=row, column=column, value=formula)
                each_region_cell.alignment = self.align
                each_region_cell.font = Font(name='Times New Roman', bold=True, size=12)
            return region_id

