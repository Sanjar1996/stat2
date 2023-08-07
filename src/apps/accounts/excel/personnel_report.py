from datetime import datetime
from django.db import connection
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter, column_index_from_string
from ...accounts.models import Region, Department


class PoultrySheet:
    def __init__(self, start=None, end=None):
        self.end_date = end
        self.start_date = start

    def generate_poultry_excel_report(self):
        now = datetime.now()
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f"attachment; filename=parranda-{now.strftime('%d-%m-%Y')}.xlsx"
        wb = Workbook()
        sheet = wb.active
        align = Alignment(wrapText=True, horizontal='center', vertical='center')
        sheet.column_dimensions['A'].width = 5
        sheet.column_dimensions['B'].width = 50
        sheet.row_dimensions[1].height = 50
        sheet.row_dimensions[2].height = 70
        sheet.row_dimensions[3].height = 25
        sheet.row_dimensions[4].height = 80

        sheet.merge_cells('A1:Y1')
        title = sheet['A1']
        title.value = f"Ўрмон хўжалиги давлат қўмитаси тизимидаги хўжалик субъектларида парранда бош сонини " \
                      f"кўпайтириш, тухум ва парранда гўшти ишлаб чиқариш бўйича {self.end_date} ҳолатига МАЪЛУМОТ"
        title.font = Font(name='Times New Roman', bold=True, size=13)
        title.alignment = align
        sheet.merge_cells('A2:A4')
        sheet.merge_cells('B2:B4')
        number = sheet.cell(column=1, row=2, value='Т/р')
        number.alignment = align
        number.font = Font(name='Times New Roman', bold=True, size=11)

        department_title = sheet.cell(column=2, row=2, value='Хўжалик номи')
        department_title.alignment = align
        department_title.font = Font(name='Times New Roman', bold=True, size=15)

        sheet.merge_cells('C2:F2')
        old_residue = sheet.cell(column=3, row=2, value=f'{int(self.end_date[:4]) - 1}-01-01 йил холатига мавжуд')
        old_residue.alignment = align
        old_residue.font = Font(name='Times New Roman', bold=True, size=12)
        sheet.merge_cells('D3:F3')
        total_poultry = sheet.cell(column=3, row=3, value='Жами парранда')
        total_poultry.alignment = align
        sheet.merge_cells('C3:C4')
        also = sheet.cell(column=4, row=3, value='Шундан')
        also.alignment = align

        turkey = sheet.cell(column=4, row=4, value='Курка сони')
        turkey.alignment = align
        roosters_chickens = sheet.cell(column=5, row=4, value='Жами хўроз ва товуқлар, сони')
        roosters_chickens.alignment = align
        egg = sheet.cell(column=6, row=4, value='Шундан тухум берадиган товуқлар сони')
        egg.alignment = align
        egg.font = Font(name='Times New Roman', size=10)

        # sheet.freeze_panes = 'A5'
        wb.save(response)
        return response

    def department_and_annual_residue(self):
        with connection.cursor() as cursor:
            old = int(self.end_date[:4]) - 1
            start = f'{old}-01-01'
            query = f"""
            select 
            d.sort, d.region_id, d.name, 
            sum(ch_actual_kurka.amount) jami_kurka, 
            sum(ch_actual_tovuq.amount) jami_tovuq
            from department d
            left join chorvachilik_types cht on cht.id = {self.type_name.pk} 
            left join chorvachilik_actual ch_actual_kurka ON ch_actual_kurka.chorvachilik_type_id = cht.id  and ch_actual_kurka.amount_type = 2 
            and ch_actual_kurka.date BETWEEN '{start}' AND '{self.start_date}' 
            left join chorvachilik_actual ch_actual_tovuq ON ch_actual_tovuq.chorvachilik_type_id = cht.id  and ch_actual_tovuq.amount_type = 2 
            and ch_actual_tovuq.date BETWEEN '{start}' AND '{self.start_date}' 
            where d.status = 1 
            group by d.id order by d.sort"""
            cursor.execute(query)
            row = cursor.fetchall()
            return row

    def get_value_by_chorva_type(self, parent):
        with connection.cursor() as cursor:
            start = self.start_date
            end = self.end_date
            query = f"""
            select d.region_id, 
            sum(ch_plan_kurka.amount) plan_kurka, 
            sum(ch_actual_kurka.amount) actual_kurka, 
            sum(ch_plan_tovuq.amount) plan_tovuq, 
            sum(ch_actual_tovuq.amount) actual_tovuq
            from department d
            left join chorvachilik ch on ch.parent_id = {parent} 
            left join chorvachilik_plan ch_plan_kurka ON ch_plan_kurka.chorvachilik_id = ch.id and ch_plan_kurka.amount_type = 2 
            and ch_plan_kurka.date BETWEEN '{start}' AND '{end}'
            left join chorvachilik_actual ch_actual_kurka ON ch_actual_kurka.chorvachilik_id = ch.id  and ch_plan_kurka.amount_type = 2
            and ch_actual_kurka.date BETWEEN '{start}' AND '{end}'
            left join chorvachilik_plan ch_plan_tovuq ON ch_plan_tovuq.chorvachilik_id = ch.id  and ch_plan_tovuq.amount_type = 2
            and ch_plan_tovuq.date BETWEEN '{start}' AND '{end}'
            left join chorvachilik_actual ch_actual_tovuq ON ch_actual_tovuq.chorvachilik_id = ch.id  and ch_plan_tovuq.amount_type = 2 
            and ch_actual_tovuq.date BETWEEN '{start}' AND '{end}'
            where d.status = 1 
            group by d.id order by d.sort"""
            cursor.execute(query)
            row = cursor.fetchall()
            return row

    def get_profit_by_chorva_type(self):
        with connection.cursor() as cursor:
            start = self.start_date
            end = self.end_date
            query = f"""
            select  d.region_id, 
            sum(ch_actual.profit) profit
            from department d
            left join chorvachilik_types cht on cht.id = {self.type_name.pk} 
            left join chorvachilik_actual ch_actual on ch_actual.chorvachilik_type_id = cht.id 
            and ch_actual.amount_type = 2 
            and ch_actual.date between '{start}' and '{end}' where d.status = 1
            group by d.id order by d.sort"""
            cursor.execute(query)
            row = cursor.fetchall()
            return row

