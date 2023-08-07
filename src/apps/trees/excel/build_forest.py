from datetime import datetime
from django.db import connection
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from .utils import set_border
from ...accounts.models import Region, Department


class ForestTreePlantSheetSheet:
    def __init__(self, user, start=None, end=None):
        self.user = user
        self.start_date = start
        self.end_date = end
        self.align = Alignment(wrapText=True, horizontal='center', vertical='center')

    def generate_forest_excel_report(self):
        now = datetime.now()
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f"attachment; filename=o'rmon_barpo_qilish-{now.strftime('%d-%m-%Y')}.xlsx"
        wb = Workbook()
        sheet = wb.active
        align = Alignment(wrapText=True, horizontal='center', vertical='center')
        sheet.column_dimensions['A'].width = 5
        sheet.merge_cells('A2:W2')
        main_title_value = f"Қарор, ташриф ва баёнлар ижроси бўйича (ПҚ-4424, 70-Баён ва ташрифлар) ўрмон фонди " \
                           f"ерларида {self.start_date[:4]} йил якунига қадар ўрмон барпо қилиш кўрсатгичларининг " \
                           f"таҳлили ва бажарилиши тўғрисида тезкор маълумот"
        main_title = sheet.cell(row=2, column=1, value=main_title_value)
        main_title.alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
        main_title.font = Font(name='Times New Roman', bold=True, size=13)
        sheet.row_dimensions[2].height = 35

        sheet.merge_cells('U3:W3')
        hectare_title = sheet.cell(row=3, column=21, value='гектар ҳисобида')
        hectare_title.alignment = align
        hectare_title.font = Font(name='Times New Roman', size=11)

        sheet.merge_cells('A4:A8')
        sheet.row_dimensions[4].height = 20
        sheet.row_dimensions[5].height = 20
        sheet.row_dimensions[6].height = 20
        sheet.row_dimensions[7].height = 25
        sheet.row_dimensions[8].height = 25

        number = sheet.cell(row=4, column=1, value='№')
        number.font = Font(name='Times New Roman', bold=True, size=10)
        number.alignment = align

        sheet.merge_cells('B4:B8')
        sheet.column_dimensions['B'].width = 55
        department_names = sheet.cell(row=4, column=2, value='Ўрмон хўжаликлари номи')
        department_names.font = Font(name='Times New Roman', bold=True, size=14)
        department_names.alignment = align

        sheet.merge_cells('C4:E6')
        big_trees_task = sheet.cell(row=4, column=3, value='Ўрмон барпо қилиш- йиллик жами')
        big_trees_task.font = Font(name='Times New Roman', bold=True, size=13)
        big_trees_task.alignment = align

        sheet.column_dimensions['C'].width = 15
        sheet.column_dimensions['D'].width = 14
        sheet.column_dimensions['E'].width = 13

        sheet.merge_cells('C7:C8')
        sheet.merge_cells('D7:D8')
        sheet.merge_cells('E7:E8')

        in_task = 'топшириқ'
        in_progress = 'амалда'

        task_title = sheet.cell(row=7, column=3, value=in_task.upper())
        task_title.font = Font(name='Times New Roman', bold=True, size=11)
        task_title.alignment = align

        progress_title = sheet.cell(row=7, column=4, value=in_progress.upper())
        progress_title.font = Font(name='Times New Roman', bold=True, size=11)
        progress_title.alignment = align

        percent_title = sheet.cell(row=7, column=5, value='%')
        percent_title.font = Font(name='Times New Roman', italic=True, bold=True, size=13)
        percent_title.alignment = align

        sheet.merge_cells('F4:W4')
        including_title = sheet.cell(row=4, column=6, value='Шу жумладан:')
        including_title.font = Font(name='Times New Roman', bold=True, size=12)
        including_title.alignment = align

        sheet.merge_cells('F5:G6')
        sheet.column_dimensions['F'].width = 11
        sheet.column_dimensions['G'].width = 11
        desert_plants_title = sheet.cell(row=5, column=6, value='чўл ўсимликлари')
        desert_plants_title.font = Font(name='Times New Roman', bold=True, underline="single", size=12)
        desert_plants_title.alignment = align

        sheet.merge_cells('F7:F8')
        f_7 = sheet.cell(row=7, column=6, value=in_task)
        f_7.font = Font(name='Times New Roman', bold=True, size=11)
        f_7.alignment = align

        sheet.merge_cells('G7:G8')
        g_7 = sheet.cell(row=7, column=7, value=in_progress)
        g_7.font = Font(name='Times New Roman', bold=True, size=11)
        g_7.alignment = align

        sheet.merge_cells('H6:I7')
        h_6 = sheet.cell(row=6, column=8, value='жами')
        h_6.font = Font(name='Times New Roman', bold=True, underline="single", size=13)
        h_6.alignment = align

        sheet.column_dimensions['H'].width = 10
        sheet.column_dimensions['I'].width = 10

        h_8_in_task = sheet.cell(row=8, column=8, value=in_task)
        h_8_in_task.font = Font(name='Times New Roman', bold=True, size=10)
        h_8_in_task.alignment = align

        h_8_in_progress = sheet.cell(row=8, column=9, value=in_progress)
        h_8_in_progress.font = Font(name='Times New Roman', bold=True, size=10)
        h_8_in_progress.alignment = align

        sheet.merge_cells('H5:O5')
        nuts_title = sheet.cell(row=5, column=8, value='ёнғоқмевалилар:')
        nuts_title.font = Font(name='Times New Roman', bold=True, underline="single", size=11)
        nuts_title.alignment = align

        sheet.merge_cells('J6:O6')
        nuts_title = sheet.cell(row=6, column=10, value='жумладан:')
        nuts_title.font = Font(name='Times New Roman', bold=True, underline="single", size=11)
        nuts_title.alignment = align

        sheet.column_dimensions['J'].width = 10
        sheet.column_dimensions['K'].width = 10
        sheet.column_dimensions['L'].width = 10
        sheet.column_dimensions['M'].width = 10
        sheet.column_dimensions['N'].width = 10
        sheet.column_dimensions['O'].width = 10
        sheet.column_dimensions['P'].width = 10
        sheet.column_dimensions['Q'].width = 10
        font = Font(name='Times New Roman', size=10)
        sheet.merge_cells('J7:K7')
        greece_nut_title = sheet.cell(row=7, column=10, value='грек ёнғоғи')
        greece_nut_title.font = Font(name='Times New Roman', bold=True, size=11)
        greece_nut_title.alignment = align
        j_8_in_task = sheet.cell(row=8, column=10, value=in_task)
        j_8_in_task.font = font
        j_8_in_task.alignment = align
        j_8_in_progress = sheet.cell(row=8, column=11, value=in_progress)
        j_8_in_progress.font = font
        j_8_in_progress.alignment = align

        sheet.merge_cells('L7:M7')
        greece_nut_title = sheet.cell(row=7, column=12, value='писта')
        greece_nut_title.font = Font(name='Times New Roman', bold=True, size=11)
        greece_nut_title.alignment = align
        l_8_in_task = sheet.cell(row=8, column=12, value=in_task)
        l_8_in_task.font = font
        l_8_in_task.alignment = align
        m_8_in_progress = sheet.cell(row=8, column=13, value=in_progress)
        m_8_in_progress.font = font
        m_8_in_progress.alignment = align

        sheet.merge_cells('N7:O7')
        uz_nut_title = sheet.cell(row=7, column=14, value='бодом')
        uz_nut_title.font = Font(name='Times New Roman', bold=True, size=11)
        uz_nut_title.alignment = align
        n_8_in_task = sheet.cell(row=8, column=14, value=in_task)
        n_8_in_task.font = font
        n_8_in_task.alignment = align
        o_8_in_progress = sheet.cell(row=8, column=15, value=in_progress)
        o_8_in_progress.font = font
        o_8_in_progress.alignment = align

        sheet.merge_cells('P5:U5')
        nuts_title = sheet.cell(row=5, column=16, value='тез ўсувчилар:')
        nuts_title.font = Font(name='Times New Roman', bold=True, underline="single", size=11)
        nuts_title.alignment = align

        sheet.merge_cells('P6:Q6')
        nuts_title = sheet.cell(row=6, column=16, value='жами')
        nuts_title.font = Font(name='Times New Roman', bold=True, underline="single", size=11)
        nuts_title.alignment = align

        sheet.merge_cells('R6:U6')
        nuts_title = sheet.cell(row=6, column=18, value='шундан:')
        nuts_title.font = Font(name='Times New Roman', bold=True, size=11)
        nuts_title.alignment = align

        sheet.merge_cells('P7:P8')
        nuts_title = sheet.cell(row=7, column=16, value=in_task)
        nuts_title.font = Font(name='Times New Roman', bold=True, size=10)
        nuts_title.alignment = align

        sheet.merge_cells('Q7:Q8')
        nuts_title = sheet.cell(row=7, column=17, value=in_progress)
        nuts_title.font = Font(name='Times New Roman', bold=True, size=10)
        nuts_title.alignment = align

        sheet.merge_cells('R7:S7')
        uz_nut_title = sheet.cell(row=7, column=18, value='терак')
        uz_nut_title.font = Font(name='Times New Roman', bold=True, size=11)
        uz_nut_title.alignment = align
        r_8_in_task = sheet.cell(row=8, column=18, value=in_task)
        r_8_in_task.font = font
        r_8_in_task.alignment = align
        s_8_in_progress = sheet.cell(row=8, column=19, value=in_progress)
        s_8_in_progress.font = font
        s_8_in_progress.alignment = align

        sheet.merge_cells('T7:U7')
        uz_nut_title = sheet.cell(row=7, column=20, value='павловния')
        uz_nut_title.font = Font(name='Times New Roman', bold=True, size=11)
        uz_nut_title.alignment = align
        t_8_in_task = sheet.cell(row=8, column=20, value=in_task)
        t_8_in_task.font = Font(name='Times New Roman', size=10)
        t_8_in_task.alignment = align
        u_8_in_progress = sheet.cell(row=8, column=21, value=in_progress)
        u_8_in_progress.font = Font(name='Times New Roman', size=10)
        u_8_in_progress.alignment = align

        sheet.merge_cells('V5:W6')
        nuts_title = sheet.cell(row=5, column=22, value='бошқа дарахт турлари')
        nuts_title.font = Font(name='Times New Roman', bold=True, underline="single", size=11)
        nuts_title.alignment = align

        sheet.merge_cells('V7:V8')
        nuts_title = sheet.cell(row=7, column=22, value=in_task)
        nuts_title.font = Font(name='Times New Roman', bold=True, size=10)
        nuts_title.alignment = align

        sheet.merge_cells('W7:W8')
        nuts_title = sheet.cell(row=7, column=23, value=in_progress)
        nuts_title.font = Font(name='Times New Roman', bold=True, size=10)
        nuts_title.alignment = align
        tree_ground_planting_list = self.forest_prepare_raw_sql()
        row_index = 9
        region_id = None
        region_row = []
        article_id = 0

        for index, value in enumerate(tree_ground_planting_list):
            if region_id is None or int(value[2]) != region_id:
                article_id = 0
                region_row.append(row_index)
                region_id = self.write_region_data(sheet=sheet, row=row_index, region_pk=value[2])
                row_index += 1
            article_id += 1
            department_id = sheet.cell(column=1, row=row_index, value=article_id)
            department_id.alignment = align
            department_id.font = Font(name='Times New Roman', size=11)

            department_name = sheet.cell(column=2, row=row_index, value=value[1])
            department_name.font = Font(name='Times New Roman', size=11)

            formula_1 = f'=+F{row_index}+H{row_index}+P{row_index}+V{row_index}'
            cell_task = sheet.cell(column=3, row=row_index, value=formula_1)
            cell_task.font = Font(name='Times New Roman', bold=True, size=11)
            cell_task.alignment = align

            formula_2 = f'=+G{row_index}+I{row_index}+Q{row_index}+W{row_index}'
            cell_progress = sheet.cell(column=4, row=row_index, value=formula_2)
            cell_progress.font = Font(name='Times New Roman', bold=True, size=11)
            cell_progress.alignment = align

            percent_formula = f'=IF(D{row_index}={0},{0},D{row_index}*{100}/C{row_index})'
            cell_percent = sheet.cell(column=5, row=row_index, value=percent_formula)
            cell_percent.font = Font(name='Times New Roman', bold=True, size=11)
            cell_percent.alignment = align
            cell_percent.number_format = '0.0'

            desert_plant_task = sheet.cell(column=6, row=row_index, value=value[3])
            desert_plant_task.font = Font(name='Times New Roman', size=11)
            desert_plant_task.alignment = align

            desert_plant_progress = sheet.cell(column=7, row=row_index, value=value[4])
            desert_plant_progress.font = Font(name='Times New Roman', size=11)
            desert_plant_progress.alignment = align

            formula_3 = f'=+J{row_index}+L{row_index}+N{row_index}'
            cell_nuts_total_task = sheet.cell(column=8, row=row_index, value=formula_3)
            cell_nuts_total_task.font = Font(name='Times New Roman', bold=True, size=10)
            cell_nuts_total_task.alignment = align

            formula_4 = f'=+K{row_index}+M{row_index}+O{row_index}'
            cell_nuts_total_progress = sheet.cell(column=9, row=row_index, value=formula_4)
            cell_nuts_total_progress.font = Font(name='Times New Roman', bold=True, size=10)
            cell_nuts_total_progress.alignment = align

            # greece nuts, task and progress
            cell_greece_nut_task = sheet.cell(column=10, row=row_index, value=value[5])
            cell_greece_nut_task.font = Font(name='Times New Roman', size=11)
            cell_greece_nut_task.alignment = align

            cell_greece_nut_progress = sheet.cell(column=11, row=row_index, value=value[6])
            cell_greece_nut_progress.font = Font(name='Times New Roman', size=11)
            cell_greece_nut_progress.alignment = align

            # pistachios, task and progress
            cell_pistachios_task = sheet.cell(column=12, row=row_index, value=value[7])
            cell_pistachios_task.font = Font(name='Times New Roman', size=11)
            cell_pistachios_task.alignment = align

            cell_pistachios_progress = sheet.cell(column=13, row=row_index, value=value[8])
            cell_pistachios_progress.font = Font(name='Times New Roman', size=11)
            cell_pistachios_progress.alignment = align

            # uz_nut, task and progress
            cell_uz_nut_task = sheet.cell(column=14, row=row_index, value=value[9])
            cell_uz_nut_task.font = Font(name='Times New Roman', size=11)
            cell_uz_nut_task.alignment = align

            cell_uz_nut_progress = sheet.cell(column=15, row=row_index, value=value[10])
            cell_uz_nut_progress.font = Font(name='Times New Roman', size=11)
            cell_uz_nut_progress.alignment = align

            # fast growing, task and progress
            cell_fast_task = sheet.cell(column=16, row=row_index, value=f'=+R{row_index}+T{row_index}')
            cell_fast_task.font = Font(name='Times New Roman', bold=True, size=10)
            cell_fast_task.alignment = align

            cell_fast_progress = sheet.cell(column=17, row=row_index, value=f'=+S{row_index}+U{row_index}')
            cell_fast_progress.font = Font(name='Times New Roman', bold=True, size=10)
            cell_fast_progress.alignment = align

            # poplar, task and progress
            cell_poplar_task = sheet.cell(column=18, row=row_index, value=value[11])
            cell_poplar_task.font = Font(name='Times New Roman', size=11)
            cell_poplar_task.alignment = align

            cell_poplar_progress = sheet.cell(column=19, row=row_index, value=value[12])
            cell_poplar_progress.font = Font(name='Times New Roman', size=11)
            cell_poplar_progress.alignment = align

            # paulownia, task and progress
            cell_paulownia_task = sheet.cell(column=20, row=row_index, value=value[13])
            cell_paulownia_task.font = Font(name='Times New Roman', size=11)
            cell_paulownia_task.alignment = align

            cell_paulownia_progress = sheet.cell(column=21, row=row_index, value=value[14])
            cell_paulownia_progress.font = Font(name='Times New Roman', size=11)
            cell_paulownia_progress.alignment = align

            # other plant, task and progress
            cell_other_task = sheet.cell(column=22, row=row_index, value=value[15])
            cell_other_task.font = Font(name='Times New Roman', size=11)
            cell_other_task.alignment = align

            cell_other_progress = sheet.cell(column=23, row=row_index, value=value[16])
            cell_other_progress.font = Font(name='Times New Roman', size=11)
            cell_other_progress.alignment = align
            row_index += 1

        total_formula = '='
        cell_finish_total_title = sheet.cell(row=row_index, column=2, value='Республика бўйича хаммаси')
        cell_finish_total_title.font = Font(name='Times New Roman', bold=True, size=13)
        cell_finish_total_title.alignment = align
        for c in range(3, sheet.max_column + 1):
            for cat in region_row:
                total_formula += f'+{get_column_letter(c)}{cat}'
                if c == 5:
                    formula_finish = f'=IF(D{row_index}={0},{0},D{row_index}*{100}/C{row_index})'
                    cell_total_percent = sheet.cell(row=row_index, column=c, value=formula_finish)
                    cell_total_percent.alignment = align
                    cell_total_percent.font = Font(name='Times New Roman', bold=True, size=13)
                    cell_total_percent.number_format = '0.0'
                else:
                    cell_total_c = sheet.cell(row=row_index, column=c, value=total_formula)
                    cell_total_c.alignment = align
                    cell_total_c.font = Font(name='Times New Roman', bold=True, size=13)
            total_formula = '='

        set_border(sheet=sheet, cell_range=f'A1:{get_column_letter(sheet.max_column)}{row_index}')
        sheet.freeze_panes = 'A9'
        wb.save(response)
        return response

    def write_region_data(self, sheet, row, region_pk):
        qs_region = Region.objects.filter(pk=region_pk, status=1)
        if qs_region.exists():
            count = Department.objects.filter(region_id=region_pk, status=1).count()
            begin_row = row + 1
            end_row = row + count
            region_id = qs_region.first().pk
            region_name = sheet.cell(column=2, row=row, value=qs_region.first().name)
            region_name.alignment = self.align
            region_name.font = Font(name='Times New Roman', bold=True, size=12)

            c = sheet.cell(row=row, column=3, value=f'=SUM(C{begin_row}:C{end_row})')
            c.alignment = self.align
            c.font = Font(name='Times New Roman', bold=True, size=12)

            d = sheet.cell(row=row, column=4, value=f'=SUM(D{begin_row}:D{end_row})')
            d.alignment = self.align
            d.font = Font(name='Times New Roman', bold=True, size=12)
            e_percent = sheet.cell(row=row, column=5, value=f'=IF(D{row}={0},{0},D{row}*{100}/C{row})')
            e_percent.alignment = self.align
            e_percent.font = Font(name='Times New Roman', bold=True, size=12)
            e_percent.number_format = '0.0'

            f = sheet.cell(row=row, column=6, value=f'=SUM(F{begin_row}:F{end_row})')
            f.alignment = self.align
            f.font = Font(name='Times New Roman', bold=True, size=11)

            g = sheet.cell(row=row, column=7, value=f'=SUM(G{begin_row}:G{end_row})')
            g.alignment = self.align
            g.font = Font(name='Times New Roman', bold=True, size=11)

            h = sheet.cell(row=row, column=8, value=f'=SUM(H{begin_row}:H{end_row})')
            h.alignment = self.align
            h.font = Font(name='Times New Roman', bold=True, size=11)

            i = sheet.cell(row=row, column=9, value=f'=SUM(I{begin_row}:I{end_row})')
            i.alignment = self.align
            i.font = Font(name='Times New Roman', bold=True, size=11)

            j = sheet.cell(row=row, column=10, value=f'=SUM(J{begin_row}:J{end_row})')
            j.alignment = self.align
            j.font = Font(name='Times New Roman', bold=True, size=11)

            k = sheet.cell(row=row, column=11, value=f'=SUM(K{begin_row}:K{end_row})')
            k.alignment = self.align
            k.font = Font(name='Times New Roman', bold=True, size=11)

            cell_l = sheet.cell(row=row, column=12, value=f'=SUM(L{begin_row}:L{end_row})')
            cell_l.alignment = self.align
            cell_l.font = Font(name='Times New Roman', bold=True, size=11)

            m = sheet.cell(row=row, column=13, value=f'=SUM(M{begin_row}:M{end_row})')
            m.alignment = self.align
            m.font = Font(name='Times New Roman', bold=True, size=11)

            n = sheet.cell(row=row, column=14, value=f'=SUM(N{begin_row}:N{end_row})')
            n.alignment = self.align
            n.font = Font(name='Times New Roman', bold=True, size=11)

            o = sheet.cell(row=row, column=15, value=f'=SUM(O{begin_row}:O{end_row})')
            o.alignment = self.align
            o.font = Font(name='Times New Roman', bold=True, size=11)

            p = sheet.cell(row=row, column=16, value=f'=SUM(P{begin_row}:P{end_row})')
            p.alignment = self.align
            p.font = Font(name='Times New Roman', bold=True, size=11)

            q = sheet.cell(row=row, column=17, value=f'=SUM(Q{begin_row}:Q{end_row})')
            q.alignment = self.align
            q.font = Font(name='Times New Roman', bold=True, size=11)

            r = sheet.cell(row=row, column=18, value=f'=SUM(R{begin_row}:R{end_row})')
            r.alignment = self.align
            r.font = Font(name='Times New Roman', bold=True, size=11)

            s = sheet.cell(row=row, column=19, value=f'=SUM(S{begin_row}:S{end_row})')
            s.alignment = self.align
            s.font = Font(name='Times New Roman', bold=True, size=11)

            t = sheet.cell(row=row, column=20, value=f'=SUM(T{begin_row}:T{end_row})')
            t.alignment = self.align
            t.font = Font(name='Times New Roman', bold=True, size=11)

            u = sheet.cell(row=row, column=21, value=f'=SUM(U{begin_row}:U{end_row})')
            u.alignment = self.align
            u.font = Font(name='Times New Roman', bold=True, size=11)

            v = sheet.cell(row=row, column=22, value=f'=SUM(V{begin_row}:V{end_row})')
            v.alignment = self.align
            v.font = Font(name='Times New Roman', bold=True, size=11)

            w = sheet.cell(row=row, column=23, value=f'=SUM(W{begin_row}:W{end_row})')
            w.alignment = self.align
            w.font = Font(name='Times New Roman', bold=True, size=11)
            return region_id


    def forest_prepare_raw_sql(self):
        with connection.cursor() as cursor:
            query = f"""select d.id, d.name, d.region_id, 
                sum(tgpp.desert_plant), sum(tgp.desert_plant), 
                sum(tgpp.walnut), sum(tgp.walnut), 
                sum(tgpp.pistachios), sum(tgp.pistachios), 
                sum(tgpp.nut), sum(tgp.nut), 
                sum(tgpp.poplar), sum(tgp.poplar), 
                sum(tgpp.paulownia), sum(tgp.paulownia), 
                sum(tgpp.other_plants), sum(tgp.other_plants) 
                from department d 
                left join user_departments_departments udd ON udd.department_id = d.id 
                left join user_departments ud on udd.userdepartment_id = ud.id
                left join tree_ground_planting tgp on tgp.department_id = d.id and tgp.date 
                between '{self.start_date[:4]}-01-01' and '{self.end_date[:4]}-12-31' 
                left join tree_ground_planting_plan tgpp on tgpp.department_id = d.id and tgpp.date 
                between '{self.start_date[:4]}-01-01' and '{self.end_date[:4]}-12-31' 
                where d.status = 1 and ud.user_id = {self.user.id} group by d.id order by d.sort
            """
            cursor.execute(query)
            row = cursor.fetchall()
            return row
