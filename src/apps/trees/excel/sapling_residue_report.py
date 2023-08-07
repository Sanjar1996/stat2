from datetime import datetime
from django.db import connection
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from .utils import set_border, write_republic_data
from ..models import TreeCategory


class AnnualReportSaplingResidueSheet:

    def __init__(self, start=None, end=None, obj=None):
        self.start_date = start
        self.end_date = end
        self.department = obj

    def generate_annual_sapling_excel_report(self):
        now = datetime.now()
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f"attachment; filename=ko'chat_kirim_chiqim-{now.strftime('%d-%m-%Y')}.xlsx"
        wb = Workbook()
        sheet = wb.active
        align = Alignment(wrapText=True, horizontal='center', vertical='center')
        sheet.column_dimensions['A'].width = 4
        sheet.column_dimensions['B'].width = 30
        sheet.column_dimensions['C'].width = 16
        sheet.column_dimensions['F'].width = 12
        sheet.column_dimensions['G'].width = 10
        sheet.column_dimensions['H'].width = 15
        sheet.column_dimensions['J'].width = 15
        sheet.column_dimensions['M'].width = 21
        sheet.column_dimensions['N'].width = 12
        sheet.row_dimensions[1].height = 35
        sheet.merge_cells('A1:N1')
        title = sheet['A1']

        title.value = f'{self.department.name} да КЎЧАТЛАР харакати тўғрисида {self.start_date}' \
                      f' ҳолатига бўлган қолдиқ хисобот'
        title.alignment = align
        title.font = Font(name='Times New Roman', bold=True, size=14)
        sheet.merge_cells('A3:A4')
        sheet.row_dimensions[3].height = 25
        sheet.row_dimensions[4].height = 80

        sheet.merge_cells('B3:B4')
        sheet.merge_cells('C3:C4')
        sheet.merge_cells('D3:D4')
        sheet.merge_cells('E3:E4')
        sheet.merge_cells('F3:F4')
        sheet.merge_cells('G3:G4')

        trees_id = sheet.cell(column=1, row=3, value='Т/р')
        trees_id.alignment = align
        sprout_type = sheet.cell(column=2, row=3, value='Кўчатлар тури')
        sprout_type.alignment = align
        sprout_type.font = Font(name='Times New Roman', bold=True, size=14)

        annual_residue = sheet.cell(column=3, row=3, value=f'01.01.{self.start_date[:4]} йил холатига қолдиқ (йиллик)')
        annual_residue.alignment = align
        annual_residue.font = Font(name='Times New Roman', bold=True, size=11)

        donate_title = sheet.cell(column=4, row=3, value='Беғараз олинди')
        donate_title.alignment = align
        donate_title.font = Font(name='Times New Roman', size=12)

        donate_title = sheet.cell(column=5, row=3, value='Сотиб олинди')
        donate_title.alignment = align
        donate_title.font = Font(name='Times New Roman', size=12)

        donate_title = sheet.cell(column=6, row=3, value='Янгидан униб чиқан нихоллар')
        donate_title.alignment = align
        donate_title.font = Font(name='Times New Roman', size=11)

        donate_title = sheet.cell(column=7, row=3, value='Жами кирим қилинди')
        donate_title.alignment = align
        donate_title.font = Font(name='Times New Roman', bold=True, size=11)

        sheet.merge_cells('H3:M3')
        donate_title = sheet.cell(column=8, row=3, value='ЧИҚИМ (йиллик)')
        donate_title.alignment = align
        donate_title.font = Font(name='Times New Roman', size=11)

        build_forest_title = sheet.cell(column=8, row=4, value='Ўрмон ва плантация барпо қилиш учун')
        build_forest_title.alignment = align
        build_forest_title.font = Font(name='Times New Roman', size=11)

        error_title = sheet.cell(column=9, row=4, value='Хатосига экилди')
        error_title.alignment = align
        error_title.font = Font(name='Times New Roman', size=11)

        move_title = sheet.cell(column=10, row=4, value='Кўчириб экилди (школка)')
        move_title.alignment = align
        move_title.font = Font(name='Times New Roman', size=11)

        buy_title = sheet.cell(column=11, row=4, value='Сотилди')
        buy_title.alignment = align
        buy_title.font = Font(name='Times New Roman', size=11)

        given_donate_title = sheet.cell(column=12, row=4, value='Беғараз берилди')
        given_donate_title.alignment = align
        given_donate_title.font = Font(name='Times New Roman', size=11)

        sprout_list_title = sheet.cell(
            column=13, row=4, value='Ҳисобдан чиқарилган нихоллар (қарор, фармойиш, буйруқ, акт)')
        sprout_list_title.alignment = align
        sprout_list_title.font = Font(name='Times New Roman', size=11)

        sheet.merge_cells('N3:N4')
        end_date = f"{self.end_date.replace('-', '.')}  йилга қолдиқ"
        annual_residue_title = sheet.cell(column=14, row=3, value=end_date)
        annual_residue_title.alignment = align
        annual_residue_title.font = Font(name='Times New Roman', bold=True, size=11)

        category_tree_index = 0
        row_index = 5
        letters_c_d_e_f = '='
        letters_g_h_i_j_k_l_m = ''
        tree_plant_list = self.prepare_sapling_residue_sql()
        category_id = 0
        category_row = []
        article_id = 0
        end_column = sheet.max_column
        for index, data in enumerate(tree_plant_list):
            article_id += 1
            category_id = data[1]
            tree_id = sheet.cell(row=row_index, column=1, value=article_id)
            tree_id.alignment = align
            tree_id.font = Font(name='Times New Roman')

            tree_name = sheet.cell(row=row_index, column=2, value=data[2])
            tree_name.font = Font(name='Times New Roman', size=11)

            tree_residue = sheet.cell(row=row_index, column=3, value=data[3])
            tree_residue.alignment = align
            tree_residue.font = Font(name='Times New Roman', size=11)
            letters_c_d_e_f += f'+{tree_residue.column_letter}{row_index}'

            tree_donate_got = sheet.cell(row=row_index, column=4, value=data[4])
            tree_donate_got.alignment = align
            tree_donate_got.font = Font(name='Times New Roman', size=11)
            letters_c_d_e_f += f'+{tree_donate_got.column_letter}{row_index}'

            tree_selling = sheet.cell(row=row_index, column=5, value=data[5])
            tree_selling.alignment = align
            tree_selling.font = Font(name='Times New Roman', size=11)
            letters_c_d_e_f += f'+{tree_selling.column_letter}{row_index}'

            tree_new_sprout = sheet.cell(row=row_index, column=6, value=data[6])
            tree_new_sprout.alignment = align
            tree_new_sprout.font = Font(name='Times New Roman', size=11)
            letters_c_d_e_f += f'+{tree_new_sprout.column_letter}{row_index}'

            tree_subtotal = sheet.cell(row=row_index, column=7, value=letters_c_d_e_f)
            tree_subtotal.alignment = align
            tree_subtotal.font = Font(name='Times New Roman', bold=True, size=12)
            letters_c_d_e_f = '='

            tree_new_forest = sheet.cell(row=row_index, column=8, value=data[7])
            tree_new_forest.alignment = align
            tree_new_forest.font = Font(name='Times New Roman', size=11)
            letters_g_h_i_j_k_l_m += f'=+G{row_index}-{tree_new_forest.column_letter}{row_index}'

            tree_error = sheet.cell(row=row_index, column=9, value=data[8])
            tree_error.alignment = align
            tree_error.font = Font(name='Times New Roman', size=11)
            letters_g_h_i_j_k_l_m += f'-{tree_error.column_letter}{row_index}'

            tree_placed = sheet.cell(row=row_index, column=10, value=data[9])
            tree_placed.alignment = align
            tree_placed.font = Font(name='Times New Roman', size=11)
            letters_g_h_i_j_k_l_m += f'-{tree_placed.column_letter}{row_index}'

            tree_selled = sheet.cell(row=row_index, column=11, value=data[10])
            tree_selled.alignment = align
            tree_selled.font = Font(name='Times New Roman', size=11)
            letters_g_h_i_j_k_l_m += f'-{tree_selled.column_letter}{row_index}'

            tree_donate_giv = sheet.cell(row=row_index, column=12, value=data[11])
            tree_donate_giv.alignment = align
            tree_donate_giv.font = Font(name='Times New Roman', size=11)
            letters_g_h_i_j_k_l_m += f'-{tree_donate_giv.column_letter}{row_index}'

            tree_out_of = sheet.cell(row=row_index, column=13, value=data[12])
            tree_out_of.alignment = align
            tree_out_of.font = Font(name='Times New Roman', size=11)
            letters_g_h_i_j_k_l_m += f'-{tree_out_of.column_letter}{row_index}'

            tree_annual_residue = sheet.cell(row=row_index, column=14, value=letters_g_h_i_j_k_l_m)
            tree_annual_residue.alignment = align
            tree_annual_residue.font = Font(name='Times New Roman', bold=True, size=13)
            letters_g_h_i_j_k_l_m = ''
            row_index += 1

            if index == len(tree_plant_list) - 1 or category_id != tree_plant_list[index + 1][1]:
                article_id = 0
                category_row.append(row_index)
                if category_tree_index == 0:
                    begin_row = row_index - index - 1
                    end_row = row_index - 1
                else:
                    begin_row = row_index - (index - category_tree_index)
                    end_row = row_index - 1
                qs_category = TreeCategory.objects.filter(id=category_id, status=1)
                if qs_category.exists():
                    category_name = sheet.cell(row=row_index, column=2, value=qs_category.first().name)
                    category_name.alignment = align
                    category_name.font = Font(name='Times New Roman', bold=True, size=13)
                    for start_col in range(3, end_column + 1):
                        letter = get_column_letter(start_col)
                        cell = sheet.cell(row=row_index, column=start_col,
                                          value=f'=SUM({letter}{begin_row}:{letter}{end_row})')
                        cell.font = Font(name='Times New Roman', bold=True, size=13)
                        cell.alignment = align
                    category_tree_index = index
                    row_index += 1
        write_republic_data(
            sheet=sheet, row=row_index, rep_title='Жами кўчатлар',
            begin=3, end=end_column + 1, array_row=category_row[:-1])
        set_border(sheet=sheet, cell_range=f'A1:{get_column_letter(end_column)}{row_index}')
        sheet.freeze_panes = 'A5'
        wb.save(response)
        return response

    def prepare_sapling_residue_sql(self):
        with connection.cursor() as cursor:
            query = f"""
            SELECT tp.id, tp.category_id, tp.name, (SELECT SUM (si.donation) + SUM (si.buying) + 
            SUM (si.new_sprout) - SUM (so.for_the_forest) + SUM (so.unsuccessful) + 
            SUM (so.place_changed)+ SUM (so.selling) + SUM (so.donation) 
            FROM tree_plant tp1 
            LEFT JOIN sapling_input si ON si.plant_id = tp1.id AND si.date < '{self.start_date}' 
            AND si.department_id = {self.department.id} 
            LEFT JOIN sapling_output so ON so.plant_id = tp1.id AND so.date < '{self.start_date}' 
            AND so.department_id = {self.department.id} 
            WHERE tp1.id = tp.id) qoldiq, 
            SUM (si.donation) donation, SUM (si.buying) buying, SUM (si.new_sprout) new_sprout, 
            SUM (so.for_the_forest),  SUM (so.unsuccessful), SUM (so.place_changed), 
            SUM (so.selling), SUM (so.donation), SUM (so.out_of_count) 
            FROM tree_plant tp 
            LEFT JOIN sapling_input si ON si.plant_id = tp.id 
            AND si.date BETWEEN '{self.start_date}' AND '{self.end_date}' AND si.department_id = {self.department.id} 
            LEFT JOIN sapling_output so ON so.plant_id = tp.id 
            AND so.date BETWEEN '{self.start_date}' AND '{self.end_date}' AND so.department_id = {self.department.id} 
            GROUP BY tp.id, tp.name ORDER BY tp.category_id"""
            cursor.execute(query)
            row = cursor.fetchall()
            return row
