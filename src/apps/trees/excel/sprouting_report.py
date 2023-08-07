from datetime import datetime
from django.db import connection
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter, column_index_from_string
from ..models import TreePlant
from ...accounts.models import Region, Department, UserDepartment


class SproutReportSheet:
    def __init__(self, user=None, start=None, end=None):
        self.user = user
        self.start_date = start
        self.end_date = end
        self.align = Alignment(wrapText=True, horizontal='center', vertical='center')

    def generate_sprout_excel_report(self):
        now = datetime.now()
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f"attachment; filename=nixol_reja_amal-{now.strftime('%d-%m-%Y')}.xlsx"
        wb = Workbook()
        sheet = wb.active
        align = Alignment(wrapText=True, horizontal='center', vertical='center')

        sheet.column_dimensions['A'].width = 4
        sheet.row_dimensions[1].height = 40

        sheet.merge_cells('A1:X2')
        sheet.merge_cells('F3:X3')
        head_title = f"Ўрмон хўжалиги давлат қўмитаси тизимидаги ўрмон хўжаликлари томонидан " \
                     f"{self.start_date[:4]} йилда мавжуд ёпиқ илдизли йирик ҳажмли нихоллар ва " \
                     f"{self.end_date[:4]} йилда етиштириш топшириғининг бажарилиши тўғрисида тезкор маълумот"
        cell_title = sheet.cell(row=1, column=1, value=head_title)
        cell_title.alignment = align
        cell_title.font = Font(name='Times New Roman', bold=True, size=12)
        sheet.merge_cells('A3:A5')
        sheet.row_dimensions[3].height = 45
        sheet.row_dimensions[4].height = 30
        sheet.row_dimensions[5].height = 30

        number = sheet['A3']
        number.value = '№'
        number.font = Font(name='Times New Roman', bold=True, size=10)
        number.alignment = align

        sheet.merge_cells('B3:B5')
        sheet.column_dimensions['B'].width = 50
        department_names = sheet['B3']
        department_names.value = 'Ўрмон хўжаликлари номи'
        department_names.font = Font(name='Times New Roman', bold=True, size=14)
        department_names.alignment = align

        sheet.merge_cells('C3:E3')
        big_trees_task = sheet['C3']
        big_trees_task.value = 'Жами ёпиқ илдизли йирик ҳажмли кўчатларни етиштириш топшириғи'
        big_trees_task.font = Font(name='Times New Roman', bold=True, size=8)
        big_trees_task.alignment = align

        sheet.column_dimensions['C'].width = 7
        sheet.column_dimensions['D'].width = 7
        sheet.column_dimensions['E'].width = 7
        trees_plant_title = sheet['F3']
        trees_plant_title.value = 'Шу жумладан, дарахт турлари бўйича бажарилиши:'
        trees_plant_title.font = Font(name='Times New Roman', bold=True, size=12)
        trees_plant_title.alignment = align

        #  топш  | амал  |  %
        sheet.merge_cells('C4:C5')
        sheet.merge_cells('D4:D5')
        sheet.merge_cells('E4:E5')

        big_trees_task_done = sheet['C4']
        big_trees_task_done.value = 'топш'
        big_trees_task_done.font = Font(name='Times New Roman', bold=True, size=10)
        big_trees_task_done.alignment = align

        big_trees_task_progress = sheet['D4']
        big_trees_task_progress.value = 'амал'
        big_trees_task_progress.font = Font(name='Times New Roman', bold=True, size=10)
        big_trees_task_progress.alignment = align

        big_trees_task_percent = sheet['E4']
        big_trees_task_percent.value = '%'
        big_trees_task_percent.font = Font(name='Times New Roman', bold=True, size=10)
        big_trees_task_percent.alignment = align

        tree_name_coordinate = 6
        plan_actual_title = 6
        active_plants = TreePlant.objects.filter(is_show_sprouting=True, status=1).order_by('id')
        for tree_index, tree_value in enumerate(active_plants):
            tree_name = sheet.cell(row=4, column=tree_name_coordinate, value=tree_value.name)
            merge_start_column = tree_name.column_letter
            merge_end_column = get_column_letter(tree_name.col_idx + 1)
            merge_coordinate = f'{merge_start_column}4:{merge_end_column}4'
            sheet.merge_cells(merge_coordinate)
            tree_name.alignment = align
            tree_name.font = Font(name='Times New Roman', bold=True, size=9)

            cell_plan_title = sheet.cell(row=5, column=plan_actual_title, value='топш')
            cell_plan_title.alignment = align
            cell_plan_title.font = Font(name='Times New Roman', bold=True, size=9)

            tree_status_progress = sheet.cell(row=5, column=plan_actual_title + 1, value='амал')
            tree_status_progress.alignment = align
            tree_status_progress.font = Font(name='Times New Roman', bold=True, size=9)

            plan_actual_title += 2
            tree_name_coordinate += 2

        qs_user_department = UserDepartment.objects.filter(user=self.user)
        if qs_user_department.exists():
            departments = qs_user_department[0].departments.filter(status=1).order_by('sort')
        else:
            departments = []
        region_id = None
        row_index = 6
        trees_done_row_c = '='
        trees_done_row_d = '='
        total_region_row = []
        trees_plan_letters = '='
        trees_actual_letters = '='
        total_plan_letters = []
        total_actual_letters = []
        article_id = 0
        for index, department_value in enumerate(departments):
            actual_data = self.get_sprout_plan_and_actual_data(department=department_value.pk)
            if region_id is None or department_value.region.id != region_id:
                article_id = 0
                total_region_row.append(row_index)
                trees_done_row_c += f'+C{row_index}'
                trees_done_row_d += f'+D{row_index}'
                region_id = self.write_region_data(
                    sheet=sheet, row=row_index, region_id=department_value.region.id, data=actual_data)
                row_index += 1
            article_id += 1
            cell_department_id = sheet.cell(column=1, row=row_index, value=article_id)
            cell_department_id.font = Font(name='Times New Roman', size=10)
            if len(department_value.name) > 50:
                name = f'{department_value.name[:48]}...'
            else:
                name = department_value.name
            cell_department_name = sheet.cell(column=2, row=row_index, value=name)
            cell_department_name.font = Font(name='Times New Roman', size=10)
            percent_formula = f'=IF(D{row_index}={0},{0},D{row_index}*100/C{row_index})'
            cell_department_name = sheet.cell(column=5, row=row_index, value=percent_formula)
            cell_department_name.number_format = '0.0'
            cell_department_name.font = Font(name='Times New Roman', bold=True, size=10)
            cell_department_name.alignment = align
            col_index = 6
            for s_index, s_value in enumerate(actual_data):
                cell_plan = sheet.cell(column=col_index, row=row_index, value=s_value[1])
                cell_plan.font = Font(name='Times New Roman', size=10)
                cell_plan.alignment = align
                if row_index == 7:
                    plan_letter_id = cell_plan.col_idx
                    total_plan_letters.append(get_column_letter(plan_letter_id))
                    total_actual_letters.append(get_column_letter(plan_letter_id + 1))
                col_index += 1
                cell_actual = sheet.cell(column=col_index, row=row_index, value=s_value[2])
                cell_actual.font = Font(name='Times New Roman', size=10)
                cell_actual.alignment = align
                trees_plan_letters += f'+{cell_plan.column_letter}{row_index}'
                cell_plan_formula = sheet.cell(column=3, row=row_index, value=trees_plan_letters)
                cell_plan_formula.font = Font(name='Times New Roman', bold=True, size=10)
                cell_plan_formula.alignment = align

                trees_actual_letters += f'+{cell_actual.column_letter}{row_index}'
                cell_actual_formula = sheet.cell(column=4, row=row_index, value=trees_actual_letters)
                cell_actual_formula.font = Font(name='Times New Roman', bold=True, size=10)
                cell_actual_formula.alignment = align
                col_index += 1
            trees_plan_letters = '='
            trees_actual_letters = '='
            row_index += 1
        finish_total = sheet.cell(row=row_index, column=2, value="жами")
        finish_total.font = Font(name='Times New Roman', bold=True, size=16)
        finish_total.alignment = align

        cell_total_done_formula_c = sheet.cell(row=row_index, column=3, value=trees_done_row_c)
        cell_total_done_formula_c.font = Font(name='Times New Roman', bold=True, size=14)
        cell_total_done_formula_c.alignment = align

        cell_total_done_formula_d = sheet.cell(row=row_index, column=4, value=trees_done_row_d)
        cell_total_done_formula_d.font = Font(name='Times New Roman', bold=True, size=14)
        cell_total_done_formula_d.alignment = align

        finish_total_percent_formula = f'=IF(D{row_index}={0},{0},D{row_index}*100/C{row_index})'
        cell_finish_total_percent_formula = sheet.cell(column=5, row=row_index, value=finish_total_percent_formula)
        cell_finish_total_percent_formula.number_format = '0.0'
        cell_finish_total_percent_formula.font = Font(name='Times New Roman', bold=True, size=12)
        cell_finish_total_percent_formula.alignment = align

        for a in total_plan_letters:
            build_done_total_formula = '='
            for b in total_region_row:
                build_done_total_formula += f'+{a}{b}'
            cell_finish_done = sheet.cell(
                row=row_index, column=column_index_from_string(str(a)), value=build_done_total_formula)
            cell_finish_done.font = Font(name='Times New Roman', bold=True, size=10)
            cell_finish_done.alignment = align

        for c in total_actual_letters:
            build_progress_total_formula = '='
            for d in total_region_row:
                build_progress_total_formula += f'+{c}{d}'
            cell_finish_progress = sheet.cell(
                row=row_index, column=column_index_from_string(str(c)), value=build_progress_total_formula)
            cell_finish_progress.font = Font(name='Times New Roman', bold=True, size=10)
            cell_finish_progress.alignment = align

        def set_border(ws, cell_range):
            thin = Side(border_style="thin", color="000000")
            for row in ws[cell_range]:
                for cell in row:
                    cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
        set_border(sheet, f'A1:{get_column_letter(sheet.max_column)}{row_index}')

        sheet.freeze_panes = 'A6'
        wb.save(response)
        return response

    def write_region_data(self, sheet, row, region_id, data):
        qs_region = Region.objects.filter(id=region_id, status=1)
        count = Department.objects.filter(region_id=region_id, status=1).count()
        if qs_region.exists():
            begin_row = row + 1
            end_row = row + count
            region_id = qs_region.first().pk
            region_name = sheet.cell(column=2, row=row, value=qs_region.first().name)
            region_name.alignment = self.align
            region_name.font = Font(name='Times New Roman', bold=True, size=12)
            formula_coordinate = sheet.cell(column=3, row=row, value=f'=SUM(C{begin_row}:C{end_row})')
            formula_coordinate.alignment = self.align
            formula_coordinate.font = Font(name='Times New Roman', bold=True, size=10)

            formula_coordinate = sheet.cell(column=4, row=row, value=f'=SUM(D{begin_row}:D{end_row})')
            formula_coordinate.alignment = self.align
            formula_coordinate.font = Font(name='Times New Roman', bold=True, size=10)
            cell_total_percent_formula = sheet.cell(column=5, row=row, value=f'=IF(D{row}={0},{0},D{row}*100/C{row})')
            cell_total_percent_formula.number_format = '0.0'
            cell_total_percent_formula.font = Font(name='Times New Roman', bold=True, size=10)
            cell_total_percent_formula.alignment = self.align
            col = 6
            for _, _ in enumerate(data):
                formula_coordinate = sheet.cell(
                    column=col, row=row, value=f'=SUM({get_column_letter(col)}{begin_row}:{get_column_letter(col)}{end_row})')
                formula_coordinate.alignment = self.align
                formula_coordinate.font = Font(name='Times New Roman', bold=True, size=10)
                col += 1
                formula_coordinate = sheet.cell(
                    column=col, row=row, value=f'=SUM({get_column_letter(col)}{begin_row}:{get_column_letter(col)}{end_row})')
                formula_coordinate.alignment = self.align
                formula_coordinate.font = Font(name='Times New Roman', bold=True, size=10)
                col += 1
            return region_id

    def get_sprout_plan_and_actual_data(self, department):
        start = self.start_date
        end = self.end_date
        with connection.cursor() as cursor:
            query = f"""select report.id, sum(reja), sum(amal) 
            from (
            (select tp.id, sum(sp.count) reja, 0 amal 
            from tree_plant tp 
            left join sprout_plan sp on sp.plant_id = tp.id and sp.date 
            between '{start}' and '{end}' and sp.department_id = {department} and sp.status = 1 
            where tp.is_show_sprouting=true group by tp.id order by tp.id
            )union all (
            select tp.id, 0 reja, sum(s.count) amal 
            from tree_plant tp 
            left join sprout s on s.plant_id = tp.id and s.date 
            between '{start}' and '{end}' and s.department_id = {department} and s.status = 1 
            where tp.is_show_sprouting=true group by tp.id order by tp.id)
            ) report group by report.id order by report.id"""
            cursor.execute(query)
            row = cursor.fetchall()
            return row
