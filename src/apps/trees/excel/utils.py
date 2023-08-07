from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


align = Alignment(wrapText=True, horizontal='center', vertical='center')


def write_region_title(sheet, row, count, value):
    begin_row = row + 1
    end_row = row + count
    region_name = sheet.cell(column=2, row=row, value=value)
    region_name.alignment = align
    region_name.font = Font(name='Times New Roman', bold=True, size=12)
    return dict(begin=begin_row, end=end_row)


def set_border(sheet, cell_range):
    thin = Side(border_style="thin", color="000000")
    for row in sheet[cell_range]:
        for cell in row:
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)


def write_republic_data(sheet, row, rep_title, begin, end, array_row):
    total_formula = '='
    republic = sheet.cell(row=row, column=2, value=rep_title)
    republic.alignment = align
    republic.font = Font(name='Times New Roman', bold=True, size=15)
    for column in range(begin, end):
        for x in array_row:
            total_formula += f'+{get_column_letter(column)}{x}'
            cell_total_c = sheet.cell(row=row, column=column, value=total_formula)
            cell_total_c.alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
            cell_total_c.font = Font(name='Times New Roman', bold=True, size=13)
        total_formula = '='


def write_trees_title(sheet, col, value):
    letter = get_column_letter(col)
    sheet.column_dimensions[letter].width = 12
    sheet.merge_cells(f'{letter}4:{letter}5')
    tree_name = sheet.cell(row=4, column=col, value=value)
    tree_name.alignment = align
    tree_name.font = Font(name='Times New Roman', bold=True, size=9)
