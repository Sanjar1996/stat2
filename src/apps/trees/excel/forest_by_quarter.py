from datetime import datetime
from django.db import connection
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from .utils import set_border
from ..models import LandCategory
from ...accounts.models import Region, Department, UserDepartment


class FinanceReportSheet:

    def __init__(self, user, start=None, end=None):
        self.user = user
        self.start_date = start
        self.end_date = end
        self.align = Alignment(wrapText=True, horizontal='center', vertical='center')

    def generate_finance_excel_report(self):
        now = datetime.now()
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f"attachment; filename=o'rmon_barpo_kvartal-{now.strftime('%d-%m-%Y')}.xlsx"
        wb = Workbook()
        sheet = wb.active

        quarter_title = ['Йил', 'I', 'II', 'III', 'IV']

        align = Alignment(wrapText=True, horizontal='center', vertical='center')

        sheet.column_dimensions['A'].width = 5
        sheet.column_dimensions['B'].width = 55
        sheet.row_dimensions[1].height = 35
        sheet.row_dimensions[2].height = 22
        sheet.row_dimensions[3].height = 45

        sheet.merge_cells('A1:W1')
        sheet.merge_cells('A3:A4')
        sheet.merge_cells('B3:B4')

        title = sheet['A1']
        title.value = f"Ўрмон хўжалиги бош бошқармаси тизимидаги ўрмон хўжаликлари томонидан " \
                      f"{self.start_date[:4]} йилда бажариладиган ўрмончилик тадбирларининг иш ҳажмлари"
        title.font = Font(name='Times New Roman', bold=True, size=14)
        title.alignment = align

        number = sheet['A3']
        number.value = '№'
        number.alignment = align
        number.font = Font(name='Times New Roman', bold=True, size=11)

        department_name_title = sheet['B3']
        department_name_title.value = 'Ўрмон хўжаликлари номи'
        department_name_title.alignment = align
        department_name_title.font = Font(name='Times New Roman', bold=True, size=14)

        land_categories = LandCategory.objects.filter(status=1)
        next_coordinate = 0
        col_index = 3
        year_col = 0
        for lc_index, lc_value in enumerate(land_categories.order_by('id')):
            if lc_index == 0:
                first_plan_title = sheet.cell(row=3, column=col_index, value=lc_value.name)
                first_plan_title.alignment = align
                first_plan_title.font = Font(name='Times New Roman', bold=True, size=10)
                for columns, item in zip(sheet.iter_cols(min_row=4, max_row=4, min_col=3, max_col=7), quarter_title):
                    for col in columns:
                        col.value = item
                        col.font = Font(name='Times New Roman', bold=True, size=9)
                        col.alignment = align
                next_col = sheet.cell(row=3, column=(col_index + 4))
                coordinate = f"{first_plan_title.coordinate}:{next_col.coordinate}"
                sheet.merge_cells(coordinate)
                next_coordinate = next_col.col_idx
            else:
                if lc_index == 3:
                    by_verdict_title = sheet.cell(row=3, column=next_coordinate + 1, value='Карор бўйича')
                    by_verdict_title.alignment = align
                    by_verdict_title.font = Font(name='Times New Roman', bold=True, size=10)
                    next_col = sheet.cell(row=3, column=(next_coordinate + 1))
                    next_coordinate = next_col.col_idx

                    for_build_forest = sheet.cell(row=3, column=next_coordinate + 1, value='Жами ўрмон барпо қилиш, га')
                    for_build_forest.alignment = align
                    for_build_forest.font = Font(name='Times New Roman', bold=True, size=10)
                    next_col = sheet.cell(row=3, column=(next_coordinate + 5))
                    for columns, item in zip(sheet.iter_cols(min_row=4, max_row=4, min_col=next_coordinate + 1,
                                                             max_col=next_coordinate + 5), quarter_title):
                        for col in columns:
                            col.value = item
                            col.font = Font(name='Times New Roman', bold=True, size=9)
                            col.alignment = align
                            if 'Йил' in item:
                                year_col = col.col_idx
                                # col.fill = color
                    coordinate = f"{for_build_forest.coordinate}:{next_col.coordinate}"
                    sheet.merge_cells(coordinate)
                    next_coordinate = next_col.col_idx
                next_plan_title = sheet.cell(row=3, column=next_coordinate + 1, value=lc_value.name)
                next_plan_title.alignment = align
                next_plan_title.font = Font(name='Times New Roman', bold=True, size=10)
                next_col = sheet.cell(row=3, column=(next_coordinate + 5))
                for columns, item in zip(
                        sheet.iter_cols(min_row=4, max_row=4, min_col=next_coordinate + 1, max_col=next_coordinate + 5),
                        quarter_title):
                    for col in columns:
                        col.value = item
                        col.font = Font(name='Times New Roman', bold=True, size=9)
                        col.alignment = align
                coordinate = f"{next_plan_title.coordinate}:{next_col.coordinate}"
                sheet.merge_cells(coordinate)
                next_coordinate = next_col.col_idx
        row_index = 5
        region_id = None
        qs_user_department = UserDepartment.objects.filter(user=self.user)
        if qs_user_department.exists():
            departments = qs_user_department[0].departments.filter(status=1).order_by('sort')
        else:
            departments = []
        department_total_formula = '='
        region_row = []
        article_id = 0
        for index, value in enumerate(departments):
            quarter_value_by_department = self.get_data_by_quarter(department=value.pk)
            get_total_by_verdict = self.get_total_by_department(department=value.pk)
            if region_id is None or value.region.id != region_id:
                article_id = 0
                region_row.append(row_index)
                region_id = self.write_region_data(
                    sheet=sheet, row=row_index, col=year_col,
                    region_id=value.region.id, data=quarter_value_by_department)
                row_index += 1
            article_id += 1
            department_id = sheet.cell(column=1, row=row_index, value=article_id)
            department_id.alignment = align
            if len(value.name) > 53:
                name = f'{value.name[:50]}...'
            else:
                name = value.name
            department_name = sheet.cell(column=2, row=row_index, value=name)
            department_name.font = Font(name='Times New Roman', size=11)
            height_col = 3
            for quarter, verdict in zip(quarter_value_by_department, get_total_by_verdict):
                if height_col + 1 == year_col:
                    sheet.cell(column=year_col-1, row=row_index, value=verdict[1])
                    begin_col_letter = sheet.cell(column=year_col + 1, row=row_index)
                    end_col_letter = sheet.cell(column=year_col + 4, row=row_index)
                    total_forest_formula_by_verdict = f'=SUM({begin_col_letter.column_letter}{row_index}:{end_col_letter.column_letter}{row_index})'
                    set_formula = sheet.cell(column=year_col, row=row_index, value=total_forest_formula_by_verdict)
                    set_formula.alignment = align
                    set_formula.font = Font(name='Times New Roman', bold=True, size=11)

                    formula_1 = f'=+D{row_index}+I{row_index}+N{row_index}'
                    sheet.cell(column=year_col+1, row=row_index, value=formula_1).alignment = align

                    formula_2 = f'=+E{row_index}+J{row_index}+O{row_index}'
                    sheet.cell(column=year_col+2, row=row_index, value=formula_2).alignment = align

                    formula_3 = f'=+F{row_index}+K{row_index}+P{row_index}'
                    sheet.cell(column=year_col+3, row=row_index, value=formula_3).alignment = align

                    formula_4 = f'=+G{row_index}+L{row_index}+Q{row_index}'
                    sheet.cell(column=year_col+4, row=row_index, value=formula_4).alignment = align

                    height_col += 6
                begin = sheet.cell(column=height_col + 1, row=row_index)
                end = sheet.cell(column=height_col + 4, row=row_index)
                formula = f'=SUM({begin.coordinate}:{end.coordinate})'
                formula_crd = sheet.cell(column=height_col, row=row_index, value=formula)
                formula_crd.alignment = align
                formula_crd.font = Font(name='Times New Roman', bold=True, size=11)

                height_col += 1
                quarter_1 = sheet.cell(column=height_col, row=row_index, value=quarter[1])
                quarter_1.font = Font(name='Times New Roman', size=10)
                quarter_1.alignment = align

                height_col += 1
                quarter_2 = sheet.cell(column=height_col, row=row_index, value=quarter[2])
                quarter_2.font = Font(name='Times New Roman', size=10)
                quarter_2.alignment = align

                height_col += 1
                quarter_3 = sheet.cell(column=height_col, row=row_index, value=quarter[3])
                quarter_3.font = Font(name='Times New Roman', size=10)
                quarter_3.alignment = align

                height_col += 1
                quarter_4 = sheet.cell(column=height_col, row=row_index, value=quarter[4])
                quarter_4.font = Font(name='Times New Roman', size=10)
                quarter_4.alignment = align

                height_col += 1
            department_total = sheet.cell(column=2, row=row_index)
            department_total_formula += f'{department_total.coordinate}+'
            row_index += 1
        total_formula = '='
        cell_b = sheet.cell(row=row_index, column=2, value='Жами:')
        cell_b.font = Font(name='Times New Roman', bold=True, size=13)
        cell_b.alignment = align

        for c in range(3, sheet.max_column + 1):
            column = get_column_letter(c)
            sheet.column_dimensions[str(column)].width = 8
            for cat in region_row:
                total_formula += f'+{get_column_letter(c)}{cat}'
                cell_total_c = sheet.cell(row=row_index, column=c, value=total_formula)
                cell_total_c.alignment = align
                cell_total_c.font = Font(name='Times New Roman', bold=True, size=13)
            total_formula = '='

        set_border(sheet=sheet, cell_range=f'A1:{get_column_letter(sheet.max_column)}{row_index}')
        sheet.freeze_panes = 'A5'
        wb.save(response)
        return response

    def write_region_data(self, sheet, row, col, region_id, data):
        qs_region = Region.objects.filter(id=region_id, status=1)
        count = Department.objects.filter(region_id=region_id, status=1).count()
        if qs_region.exists():
            begin_row = row + 1
            end_row = row + count
            region_id = qs_region.first().pk
            region_name = sheet.cell(column=2, row=row, value=qs_region.first().name)
            region_name.alignment = self.align
            region_name.font = Font(name='Times New Roman', bold=True, size=12)
            start_col = 3
            for _, _ in enumerate(data):
                if start_col + 1 == col:
                    begin_cell = sheet.cell(column=start_col, row=begin_row)
                    end_cell = sheet.cell(column=start_col, row=end_row)

                    formula_r = f'=SUM({begin_cell.coordinate}:{end_cell.coordinate})'
                    style_1 = sheet.cell(column=col - 1, row=row, value=formula_r)
                    style_1.alignment = self.align
                    style_1.font = Font(name='Times New Roman', bold=True, size=11)

                    begin_cell = sheet.cell(column=start_col + 1, row=begin_row)
                    end_cell = sheet.cell(column=start_col + 1, row=end_row)
                    formula_s = f'=SUM({begin_cell.coordinate}:{end_cell.coordinate})'
                    style_2 = sheet.cell(column=col, row=row, value=formula_s)
                    style_2.alignment = self.align
                    style_2.font = Font(name='Times New Roman', bold=True, size=11)

                    begin_cell = sheet.cell(column=start_col + 2, row=begin_row)
                    end_cell = sheet.cell(column=start_col + 2, row=end_row)
                    formula_t = f'=SUM({begin_cell.coordinate}:{end_cell.coordinate})'
                    style_3 = sheet.cell(column=col + 1, row=row, value=formula_t)
                    style_3.alignment = self.align
                    style_3.font = Font(name='Times New Roman', bold=True, size=11)

                    begin_cell = sheet.cell(column=start_col + 3, row=begin_row)
                    end_cell = sheet.cell(column=start_col + 3, row=end_row)
                    formula_u = f'=SUM({begin_cell.coordinate}:{end_cell.coordinate})'
                    style_4 = sheet.cell(column=col + 2, row=row, value=formula_u)
                    style_4.alignment = self.align
                    style_4.font = Font(name='Times New Roman', bold=True, size=11)

                    begin_cell = sheet.cell(column=start_col + 4, row=begin_row)
                    end_cell = sheet.cell(column=start_col + 4, row=end_row)
                    formula_v = f'=SUM({begin_cell.coordinate}:{end_cell.coordinate})'
                    style_5 = sheet.cell(column=col + 3, row=row, value=formula_v)
                    style_5.alignment = self.align
                    style_5.font = Font(name='Times New Roman', bold=True, size=11)

                    begin_cell = sheet.cell(column=start_col + 5, row=begin_row)
                    end_cell = sheet.cell(column=start_col + 5, row=end_row)
                    formula_w = f'=SUM({begin_cell.coordinate}:{end_cell.coordinate})'
                    style_6 = sheet.cell(column=col + 4, row=row, value=formula_w)
                    style_6.alignment = self.align
                    style_6.font = Font(name='Times New Roman', bold=True, size=11)
                    start_col += 6

                begin_cell = sheet.cell(column=start_col, row=begin_row)
                end_cell = sheet.cell(column=start_col, row=end_row)
                formula = f'=SUM({begin_cell.coordinate}:{end_cell.coordinate})'
                formula_coordinate = sheet.cell(column=start_col, row=row, value=formula)
                formula_coordinate.alignment = self.align
                formula_coordinate.font = Font(name='Times New Roman', bold=True, size=11)
                start_col += 1

                begin_cell = sheet.cell(column=start_col, row=begin_row)
                end_cell = sheet.cell(column=start_col, row=end_row)
                formula = f'=SUM({begin_cell.coordinate}:{end_cell.coordinate})'
                formula_coordinate = sheet.cell(column=start_col, row=row, value=formula)
                formula_coordinate.alignment = self.align
                formula_coordinate.font = Font(name='Times New Roman', bold=True, size=11)
                start_col += 1

                begin_cell = sheet.cell(column=start_col, row=begin_row)
                end_cell = sheet.cell(column=start_col, row=end_row)
                formula = f'=SUM({begin_cell.coordinate}:{end_cell.coordinate})'
                formula_coordinate = sheet.cell(column=start_col, row=row, value=formula)
                formula_coordinate.alignment = self.align
                formula_coordinate.font = Font(name='Times New Roman', bold=True, size=11)
                start_col += 1

                begin_cell = sheet.cell(column=start_col, row=begin_row)
                end_cell = sheet.cell(column=start_col, row=end_row)
                formula = f'=SUM({begin_cell.coordinate}:{end_cell.coordinate})'
                formula_coordinate = sheet.cell(column=start_col, row=row, value=formula)
                formula_coordinate.alignment = self.align
                formula_coordinate.font = Font(name='Times New Roman', bold=True, size=11)
                start_col += 1

                begin_cell = sheet.cell(column=start_col, row=begin_row)
                end_cell = sheet.cell(column=start_col, row=end_row)
                formula = f'=SUM({begin_cell.coordinate}:{end_cell.coordinate})'
                formula_coordinate = sheet.cell(column=start_col, row=row, value=formula)
                formula_coordinate.alignment = self.align
                formula_coordinate.font = Font(name='Times New Roman', bold=True, size=11)
                start_col += 1
            return region_id

    def get_data_by_quarter(self, department=None):
        with connection.cursor() as cursor:
            year = self.start_date[:4]
            query = f"select lc.id, " \
                    f"(select sum(pl.hectare) from prepair_land pl " \
                    f"where pl.department_id = {department} and lc.id = pl.categories_id and pl.date " \
                    f"BETWEEN  '{year}-01-01' and '{year}-04-01') bir," \
                    f"(select sum(pl.hectare) from prepair_land pl " \
                    f"where pl.department_id = {department} and lc.id = pl.categories_id and pl.date " \
                    f"BETWEEN  '{year}-04-01' and '{year}-07-01') ikki,(select sum(pl.hectare) from prepair_land pl " \
                    f"where pl.department_id = {department} and lc.id = pl.categories_id and pl.date " \
                    f"BETWEEN  '{year}-07-01' and '{year}-10-01') uch,(select sum(pl.hectare) " \
                    f"from prepair_land pl where pl.department_id = {department} and lc.id = pl.categories_id and pl.date " \
                    f"BETWEEN  '{year}-10-01' and '{year}-01-01') turt " \
                    f"from land_category lc group by lc.id order by lc.id"
            cursor.execute(query)
            row = cursor.fetchall()
            return row

    def get_total_by_department(self, department):
        with connection.cursor() as cursor:
            start = self.start_date
            end = self.end_date
            query = f"select d.id, sum(prepair_land_plan.hectare) " \
                    f"from department d " \
                    f"left join prepair_land_plan on d.id = prepair_land_plan.department_id and prepair_land_plan.date " \
                    f"BETWEEN '{start}' and '{end}' and d.id = {department} group by d.id order by d.sort"
            cursor.execute(query)
            row = cursor.fetchall()
            return row
