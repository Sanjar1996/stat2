from datetime import datetime
from django.db import connection
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from ..models import TreePlant
from .utils import write_region_title, set_border, write_republic_data
from ...accounts.models import Region, Department, UserDepartment


class TreeHeightReportSheet:

    def __init__(self, user=None, start=None, end=None):
        self.user = user
        self.start_date = start
        self.end_date = end
        self.align = Alignment(wrapText=True, horizontal='center', vertical='center')

    def generate_height_excel_report(self):
        now = datetime.now()
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f"attachment; filename=ko'chatlar_balandligi-{now.strftime('%d-%m-%Y')}.xlsx"
        wb = Workbook()
        sheet = wb.active
        # sheet.title = 'height_report'
        # sheet.sheet_properties.tabColor = '7b8dc9'
        align = Alignment(wrapText=True, horizontal='center', vertical='center')
        sheet.column_dimensions['A'].width = 6
        sheet.column_dimensions['B'].width = 55
        sheet.column_dimensions['C'].width = 10
        sheet.row_dimensions[1].height = 30
        sheet.row_dimensions[3].height = 30
        sheet.row_dimensions[4].height = 90
        sheet.merge_cells('A1:Z2')
        sheet.merge_cells('A3:A4')
        sheet.merge_cells('B3:B4')
        title_value = f"Ўрмон хужалиги давлат қўмитаси тизим ташкилотларида {self.start_date}" \
                      f" дан {self.end_date} йил кўчатлар баландлиги бўйича маълумот"
        title = sheet.cell(column=1, row=1, value=title_value)
        title.font = Font(name='Times New Roman', bold=True, size=13)
        title.alignment = align
        trees_type = sheet.cell(column=2, row=3, value='Кўчат тури')
        trees_type.font = Font(name='Times New Roman', bold=True, size=15)
        trees_type.alignment = align
        sheet.merge_cells('C3:C4')
        total_trees = sheet.cell(column=3, row=3, value='Жами экиш ва сотишга яроқли кўчатлар')
        total_trees.font = Font(name='Times New Roman', bold=True, size=10)
        total_trees.alignment = align
        height_title = [
            'жами', '0,2 м гача', '0,2 дан 0,5 м гача', '0,5 -1 м. гача', '1-1,5 м гача', '1,5-2 м гача', '2-дан юқори'
        ]
        start_col = 4
        active_trees = TreePlant.objects.filter(status=1, is_show_height=True).order_by('sort')
        for index, tree_value in enumerate(active_trees):
            sheet.merge_cells(f'{get_column_letter(start_col)}3:{get_column_letter(start_col + 6)}3')
            tree_name = sheet.cell(row=3, column=start_col,  value=tree_value.name)
            tree_name.alignment = align
            tree_name.font = Font(name='Times New Roman', bold=True, size=12)
            for ht_id, ht_value in zip(range(len(height_title)), height_title):
                if ht_value == 'жами':
                    height_title_cell = sheet.cell(row=4, column=ht_id + start_col, value=ht_value)
                    height_title_cell.alignment = Alignment(textRotation=90, horizontal='center', vertical='center')
                    height_title_cell.font = Font(name='Times New Roman', bold=True, size=14)
                else:
                    height_title_cell = sheet.cell(row=4, column=ht_id + start_col, value=ht_value)
                    height_title_cell.alignment = Alignment(textRotation=90, horizontal='center', vertical='center')
                    height_title_cell.font = Font(name='Times New Roman', size=10)
            start_col += 7

        qs_user_department = UserDepartment.objects.filter(user=self.user)
        if qs_user_department.exists():
            departments = qs_user_department[0].departments.filter(status=1).order_by('sort')
        else:
            departments = []
        row_index = 5
        region_id = None
        region_row = []
        article_id = 0
        for index, department_value in enumerate(departments):
            tree_height_data = self.height_prepare_raw_sql(department=department_value.pk)
            if region_id is None or department_value.region.id != region_id:
                article_id = 0
                region_row.append(row_index)
                region_id = self.write_region_data(sheet=sheet, row=row_index, region_id=department_value.region.id)
                row_index += 1
            article_id += 1
            sheet.cell(column=1, row=row_index, value=article_id).alignment = align
            if len(department_value.name) > 50:
                name = f'{department_value.name[:48]}...'
            else:
                name = department_value.name
            department_name = sheet.cell(column=2, row=row_index, value=name)
            department_name.font = Font(name='Times New Roman', size=11)

            start_col = 4
            trees_subtotal_letters = '='
            for _, height in enumerate(tree_height_data):
                trees_subtotal_letters += f'+{get_column_letter(start_col)}{row_index}'
                tree_total = sheet.cell(
                    column=start_col, row=row_index,
                    value=f'=SUM({get_column_letter(start_col + 1)}{row_index}:{get_column_letter(start_col + 6)}{row_index})')
                tree_total.alignment = align
                tree_total.font = Font(name='Times New Roman', bold=True)
                sheet.cell(column=start_col + 1, row=row_index, value=height[0]).alignment = align
                sheet.cell(column=start_col + 2, row=row_index, value=height[1]).alignment = align
                sheet.cell(column=start_col + 3, row=row_index, value=height[2]).alignment = align
                sheet.cell(column=start_col + 4, row=row_index, value=height[3]).alignment = align
                sheet.cell(column=start_col + 5, row=row_index, value=height[4]).alignment = align
                sheet.cell(column=start_col + 6, row=row_index, value=height[5]).alignment = align
                start_col += 7
            department_total_value = sheet.cell(column=3, row=row_index, value=trees_subtotal_letters)
            department_total_value.font = Font(name='Times New Roman', bold=True, size=12)
            department_total_value.alignment = align
            row_index += 1
        write_republic_data(
            sheet=sheet, row=row_index, rep_title='Республика бўйича жами',
            begin=3, end=sheet.max_column + 1, array_row=region_row)
        set_border(sheet=sheet, cell_range=f'A1:{get_column_letter(sheet.max_column)}{row_index}')
        sheet.freeze_panes = 'A5'
        wb.save(response)
        return response

    def write_region_data(self, sheet, row, region_id):
        qs_region = Region.objects.filter(id=region_id, status=1)
        count = Department.objects.filter(region_id=region_id, status=1).count()
        if qs_region.exists():
            from_to = write_region_title(sheet=sheet, row=row, count=count, value=qs_region.first().name)
            region_id = qs_region.first().pk
            region_total = sheet.cell(column=3, row=row, value=f"=SUM(C{from_to['begin']}:C{from_to['end']})")
            region_total.alignment = self.align
            region_total.font = Font(name='Times New Roman', bold=True, size=12)
            for col in range(4, sheet.max_column + 1):
                region_subtotal = sheet.cell(
                    row=row, column=col,
                    value=f"=SUM({get_column_letter(col)}{from_to['begin']}:{get_column_letter(col)}{from_to['end']})")
                region_subtotal.alignment = self.align
                region_subtotal.font = Font(name='Times New Roman', bold=True, size=11)
            return region_id

    def height_prepare_raw_sql(self, department):
        with connection.cursor() as cursor:
            start = self.start_date
            end = self.end_date
            query = f"""select 
                thr.height_0_0_2_count, 
                thr.height_0_2_5_count, 
                thr.height_0_5_1_count, 
                thr.height_1_1_5_count, 
                thr.height_1_5_2_count, 
                thr.height_2_count 
                from tree_height thr 
                right join tree_plant tp on thr.tree_plan_id = tp.id and thr.department_id = {department} 
                and thr.status=1 and thr.date between '{start}' and '{end}' 
                where tp.is_show_height=true order by tp.sort"""
            cursor.execute(query)
            row = cursor.fetchall()
            return row
