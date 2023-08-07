from datetime import datetime
from django.http import HttpResponse
from django.db import connection
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from ..models import TreePlant
from .db_service import TreePlantQuery
from .utils import write_region_title, write_republic_data, set_border, write_trees_title
from ...accounts.models import Department, UserDepartment, Region


class SaplingReportSheet(TreePlantQuery):
    def __init__(self, user=None, start=None, end=None):
        self.user = user
        self.start_date = start
        self.end_date = end
        self.align = Alignment(wrapText=True, horizontal='center', vertical='center')

    def generate_sapling_excel_report(self):
        now = datetime.now()
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f"attachment; filename=ko'chat_reja_amal-{now.strftime('%d-%m-%Y')}.xlsx"
        wb = Workbook()
        sheet = wb.active
        align = Alignment(wrapText=True, horizontal='center', vertical='center')

        sheet.column_dimensions['A'].width = 4
        sheet.row_dimensions[1].height = 40

        sheet.merge_cells('A1:X2')
        head_title = f"Ўрмон хўжалиги давлат қўмитаси тизимидаги ўрмон хўжаликлари томонидан {self.start_date[:4]} " \
                     f"йилда мавжуд ёпиқ илдизли йирик ҳажмли кўчатлар ва {self.end_date[:4]} " \
                     f"йилда етиштириш топшириғининг бажарилиши тўғрисида тезкор маълумот"
        cell_title = sheet.cell(row=1, column=1, value=head_title)
        cell_title.alignment = align
        cell_title.font = Font(name='Times New Roman', bold=True, size=12)

        sheet.merge_cells('A3:A5')
        sheet.row_dimensions[3].height = 45
        sheet.row_dimensions[4].height = 30
        sheet.row_dimensions[5].height = 30
        number_title = sheet.cell(row=3, column=1, value='№')
        number_title.font = Font(name='Times New Roman', bold=True, size=10)
        number_title.alignment = align
        sheet.merge_cells('B3:B5')
        sheet.column_dimensions['B'].width = 55
        department_name_title = sheet.cell(row=3, column=2, value='Ўрмон хўжаликлари номи')
        department_name_title.font = Font(name='Times New Roman', bold=True, size=14)
        department_name_title.alignment = align
        sheet.column_dimensions['C'].width = 14
        sheet.merge_cells('C3:D4')
        task_title = sheet.cell(row=3, column=3, value='Турлари буйича жами:')
        task_title.font = Font(name='Times New Roman', bold=True, size=10)
        task_title.alignment = align
        sheet.merge_cells('E3:X3')
        tree_type_title = sheet.cell(row=3, column=5, value='Шу жумладан дарахт турлари амалда бажарилиши')
        tree_type_title.font = Font(name='Times New Roman', bold=True, size=15)
        tree_type_title.alignment = align
        task = sheet.cell(row=5, column=3, value='топшириқ')
        task.font = Font(name='Times New Roman', bold=True, size=12)
        task.alignment = align
        act = sheet.cell(row=5, column=4, value='амалда')
        act.font = Font(name='Times New Roman', bold=True, size=12)
        act.alignment = align

        start_col = 5
        plants = TreePlant.objects.filter(is_show_sapling=True, status=1).order_by('id')
        for tree_index, tree_value in enumerate(plants):
            write_trees_title(sheet=sheet, col=start_col, value=tree_value.name)
            start_col += 1
        # sapling_data = self.get_tree_plan_data(user=self.user, start=self.start_date, end=self.end_date, qs=plants)
        qs_user_department = UserDepartment.objects.filter(user=self.user)
        if qs_user_department.exists():
            departments = qs_user_department[0].departments.filter(status=1).order_by('sort')
        else:
            departments = []
        region_id = None
        row_index = 6
        region_row = []
        article_id = 0
        end_col = plants.count() + 5
        for index, value in enumerate(departments):
            actual_data = self.get_sapling_actual_data(department=value.id)
            if region_id is None or value.region.id != region_id:
                article_id = 0
                region_row.append(row_index)
                region_id = self.write_region_data(
                    sheet=sheet, row=row_index, region_id=value.region.id, end_column=end_col)
                row_index += 1
            article_id += 1
            sheet.cell(column=1, row=row_index, value=article_id).alignment = align
            if len(value.name) > 53:
                name = f"{value.name[:50]}..."
            else:
                name = value.name
            department_name = sheet.cell(column=2, row=row_index, value=name)
            department_name.font = Font(name='Times New Roman', size=11)
            count__sum = self.get_sapling_year_plan(dep_id=value.id, start=self.start_date, end=self.end_date)
            year_plan = sheet.cell(column=3, row=row_index, value=count__sum)
            year_plan.font = Font(name='Times New Roman', bold=True, size=12)
            year_plan.alignment = align
            all_act = sheet.cell(
                column=4, row=row_index, value=f'=SUM(E{row_index}:{get_column_letter(end_col)}{row_index})')
            all_act.font = Font(name='Times New Roman', bold=True, size=12)
            all_act.alignment = align
            column = 5
            for col, item in enumerate(actual_data):
                act = sheet.cell(row=row_index, column=column + col, value=item[0])
                act.alignment = align
            row_index += 1
        write_republic_data(
            sheet=sheet, row=row_index, rep_title='Республика бўйича жами',
            begin=3, end=end_col, array_row=region_row)
        set_border(sheet=sheet, cell_range=f'A1:{get_column_letter(sheet.max_column)}{row_index}')
        sheet.freeze_panes = 'A6'
        wb.save(response)
        return response

    def write_region_data(self, sheet, row, region_id, end_column):
        qs_region = Region.objects.filter(id=region_id, status=1)
        count = Department.objects.filter(region_id=region_id, status=1).count()
        if qs_region.exists():
            from_to = write_region_title(sheet=sheet, row=row, count=count, value=qs_region.first().name)
            region_id = qs_region.first().pk
            for col in range(3, end_column):
                letter = get_column_letter(col)
                total_by_region = sheet.cell(column=col, row=row,
                                             value=f"=SUM({letter}{from_to['begin']}:{letter}{from_to['end']})")
                total_by_region.alignment = self.align
                total_by_region.font = Font(name='Times New Roman', bold=True, size=11)
            return region_id

    # def get_sapling_plan_and_actual_data(self, department):
    #     with connection.cursor() as cursor:
    #         query = f"""select report.id, sum(reja), sum(amal)
    #                 from (
    #                 (select tp.id, sum(sp.count) reja, 0 amal
    #                 from tree_plant tp
    #                 left join sapling_plan sp on sp.plant_id = tp.id and sp.date
    #                 between '{self.start_date}' and '{self.end_date}'
    #                 and sp.department_id = {department} and sp.status = 1
    #                 where tp.is_show_sapling=true group by tp.id order by tp.id
    #                 )union all (
    #                 select tp.id, 0 reja, sum(s.count) amal
    #                 from tree_plant tp
    #                 left join sapling s on s.plant_id = tp.id and s.date
    #                 between '{self.start_date}' and '{self.end_date}'
    #                 and s.department_id = {department} and s.status = 1
    #                 where tp.is_show_sapling=true group by tp.id order by tp.id)
    #                 ) report group by report.id order by report.id"""
    #         cursor.execute(query)
    #         row = cursor.fetchall()
    #         return row

    def get_sapling_actual_data(self, department):
        with connection.cursor() as cursor:
            query = f"""select sum(s.count) 
                        from tree_plant tp
                        left join sapling s on s.plant_id = tp.id and s.date
                        between '{self.start_date}' and '{self.end_date}'
                        and s.department_id = {department} and s.status = 1 
                        where tp.is_show_sapling = true 
                        group by tp.id, tp.sort order by tp.sort"""
            cursor.execute(query)
            row = cursor.fetchall()
            return row

