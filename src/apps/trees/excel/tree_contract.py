from datetime import datetime
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from ..models import TreeContractCategory
from .db_service import DBQuery
from .utils import write_region_title, set_border, write_republic_data
from ...accounts.models import Region


class TreeContractReportSheet(DBQuery):

    def __init__(self, user=None, start=None, end=None):
        self.user = user
        self.start_date = start
        self.end_date = end
        self.align = Alignment(wrapText=True, horizontal='center', vertical='center')

    def generate_tree_contract_excel_report(self):
        now = datetime.now()
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f"attachment; filename=ko'chatlar_shartnoma-{now.strftime('%d-%m-%Y')}.xlsx"
        wb = Workbook()
        sheet = wb.active
        align = Alignment(wrapText=True, horizontal='center', vertical='center')
        sheet.column_dimensions['A'].width = 5
        sheet.column_dimensions['B'].width = 55
        sheet.column_dimensions['C'].width = 16
        sheet.row_dimensions[1].height = 40
        sheet.row_dimensions[3].height = 20
        sheet.row_dimensions[4].height = 18
        sheet.row_dimensions[5].height = 90
        sheet.merge_cells('A1:Z1')
        title_value = f"Республикада ўрмон хўжаликлари томонидан дарахт ва бута кўчатларини харид қилиш учун" \
                      f" тузилган шартномалар ва амалга оширилган тўловлар тўғрисида " \
                      f"{self.end_date} ҳолатига тезкор маълумот"
        title = sheet.cell(column=1, row=1, value=title_value)
        title.font = Font(name='Times New Roman', bold=True, size=13)
        title.alignment = align
        sheet.merge_cells('A2:A5')
        sheet.merge_cells('B2:B5')
        sheet.merge_cells('C2:C5')
        trees_type = sheet.cell(column=2, row=2, value='Ўрмон хўжаликлари номи')
        trees_type.font = Font(name='Times New Roman', bold=True, size=15)
        trees_type.alignment = align
        contract_title = sheet.cell(row=2, column=3, value='Фармойиш бўйича кўчат етказиб бериш топшириғи, минг дона')
        contract_title.font = Font(name='Times New Roman', bold=True, size=10)
        contract_title.alignment = align
        sheet.merge_cells('D2:G3')
        contract_by_category_title = sheet.cell(row=2, column=4,
                                                value='Жами барча ташкилотлар билан тузилган шартномалар')
        contract_by_category_title.font = Font(name='Times New Roman', bold=True, size=11)
        contract_by_category_title.alignment = align
        sheet.merge_cells('D4:E4')
        contract_by_title = sheet.cell(row=4, column=4,  value='Жами шартнома бўйича:')
        contract_by_title.font = Font(name='Times New Roman', bold=True, size=11)
        contract_by_title.alignment = align
        sheet.merge_cells('F4:G4')
        contract_by_title = sheet.cell(row=4, column=6, value='шундан:')
        contract_by_title.font = Font(name='Times New Roman', bold=True, size=11)
        contract_by_title.alignment = align
        sheet.merge_cells('H2:AI2')
        sheet.row_dimensions[2].height = 24
        contract_by_title = sheet.cell(row=2, column=8, value='Шу жумладан етказиб берилган кўчатлар соҳалар кесимида:')
        contract_by_title.font = Font(name='Times New Roman', bold=True, size=14)
        contract_by_title.alignment = align

        qs_contract_category = TreeContractCategory.objects.filter(status=1).order_by('id')
        self.write_first_header(sheet=sheet, column=4)
        start_col = 8
        letters_1 = '='
        letters_2 = '='
        letters_3 = '='
        letters_4 = '='
        for contract in qs_contract_category:
            coordinate = f'{get_column_letter(start_col)}3:{get_column_letter(start_col + 3)}4'
            sheet.merge_cells(coordinate)
            contract_by_title = sheet.cell(row=3, column=start_col, value=contract.name)
            contract_by_title.font = Font(name='Times New Roman', bold=True, size=11)
            contract_by_title.alignment = align
            letters_collection = self.write_category_header(sheet=sheet, column=start_col)
            letters_1 += f'+{letters_collection[0][0]}'
            letters_2 += f'+{letters_collection[1][0]}'
            letters_3 += f'+{letters_collection[2][0]}'
            letters_4 += f'+{letters_collection[3][0]}'
            start_col += 4

        row_index = 6
        region_id = None
        region_row = []
        article_id = 0
        all_tree_contract = self.get_all_tree_contracts(user=self.user, start=self.start_date, end=self.end_date)
        for index, value in enumerate(all_tree_contract):
            if region_id is None or int(value['region_id']) != region_id:
                article_id = 0
                region_row.append(row_index)
                region_id = self.write_region_data(
                    sheet=sheet, row=row_index, region_id=value['region_id'], count=value['department_count'])
                row_index += 1
            article_id += 1
            department_id = sheet.cell(column=1, row=row_index, value=article_id)
            department_id.alignment = align
            if len(value['department']) > 53:
                name = f"{value['department'][:50]}..."
            else:
                name = value['department']
            department_name = sheet.cell(column=2, row=row_index, value=name)
            department_name.font = Font(name='Times New Roman', size=11)
            year_plan = sheet.cell(column=3, row=row_index, value=value['year_plan'])
            year_plan.font = Font(name='Times New Roman', bold=True, size=12)
            year_plan.alignment = align
            formula_d = self.format_formula(letters=letters_1, row=row_index)
            formula_e = self.format_formula(letters=letters_2, row=row_index)
            formula_f = self.format_formula(letters=letters_3, row=row_index)
            formula_g = self.format_formula(letters=letters_4, row=row_index)

            cell_d = sheet.cell(row=row_index, column=4, value=formula_d)
            cell_d.alignment = align
            cell_d.font = Font(name='Times New Roman', size=12)
            cell_e = sheet.cell(row=row_index, column=5, value=formula_e)
            cell_e.alignment = align
            cell_e.font = Font(name='Times New Roman', size=12)
            cell_f = sheet.cell(row=row_index, column=6, value=formula_f)
            cell_f.alignment = align
            cell_f.font = Font(name='Times New Roman', size=12)
            cell_g = sheet.cell(row=row_index, column=7, value=formula_g)
            cell_g.alignment = align
            cell_g.font = Font(name='Times New Roman', size=12)
            column = 8
            for col, item in enumerate(value['actual']):
                act = sheet.cell(row=row_index, column=column + col, value=item)
                act.alignment = align
            row_index += 1
        write_republic_data(
            sheet=sheet, row=row_index, rep_title='Республика бўйича жами',
            begin=3, end=sheet.max_column + 1, array_row=region_row)
        set_border(sheet=sheet, cell_range=f'A1:{get_column_letter(sheet.max_column)}{row_index}')
        sheet.freeze_panes = 'A6'
        wb.save(response)
        return response

    def write_first_header(self, sheet, column):
        title = ['кўчат сони, минг дона', 'суммаси млн. сўм', 'Тўланган маблағ', 'Олиб кетилган кўчат сони, минг дона']
        for item in title:
            title = sheet.cell(row=5, column=column, value=item)
            title.font = Font(name='Times New Roman', bold=True, size=12)
            title.alignment = self.align
            sheet.column_dimensions[title.column_letter].width = 13
            column += 1

    def write_category_header(self, sheet, column):
        title_name = ['минг дона', 'млн. сўм', 'тўланган маблағ', 'олиб кетилган кўчат']
        data = []
        for item in title_name:
            title = sheet.cell(row=5, column=column, value=item)
            letter = title.column_letter
            data.append(tuple([letter]))
            title.font = Font(name='Times New Roman', size=16)
            title.alignment = self.align
            sheet.column_dimensions[letter].width = 13
            column += 1
        return data

    def write_region_data(self, sheet, row, region_id, count):
        qs_region = Region.objects.filter(id=region_id, status=1)
        if qs_region.exists():
            from_to = write_region_title(sheet=sheet, row=row, count=count, value=qs_region.first().name)
            region_id = qs_region.first().pk
            for col in range(3, sheet.max_column + 1):
                letter = get_column_letter(col)
                total_by_region = sheet.cell(column=col, row=row,
                                             value=f"=SUM({letter}{from_to['begin']}:{letter}{from_to['end']})")
                total_by_region.alignment = self.align
                total_by_region.font = Font(name='Times New Roman', bold=True, size=11)
            return region_id

    @staticmethod
    def format_formula(letters, row):
        letters_list = letters.split('+')
        formula = '='
        for x in letters_list:
            if x.isalpha():
                formula += f'+{x}{row}'
        return formula

