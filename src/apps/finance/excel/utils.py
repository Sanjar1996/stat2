from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font


class QuarterUtils:
    align = Alignment(wrapText=True, horizontal='center', vertical='center')

    def get_quarter_months(self, date):
        months = [
            {1: 'Январ', 2: 'Феврал', 3: 'Март'},
            {4: 'Апрел', 5: 'Май', 6: 'Июн'},
            {7: 'Июл', 8: 'Август', 9: 'Сентябр'},
            {10: 'Октябр', 11: 'Ноябр', 12: 'Декабр'}
        ]
        month = int(date[5:-3])
        quarters = {"quarter": 0, "count": 0, "month": []}
        for index, item in enumerate(months, start=1):
            if int(month) in item:
                quarters['quarter'] = index
                for key in item.keys():
                    if key <= int(month):
                        quarters['month'].append({key: item[key]})
                        quarters['count'] = quarters['count'] + 1
        return quarters

    def write_months(self, sheet, column, value):
        sheet.merge_cells(f'{get_column_letter(column)}4:{get_column_letter(column + 2)}4')
        month = sheet.cell(row=4, column=column, value=value)
        month.alignment = self.align
        month.font = Font(name='Times New Roman', size=14)

    def write_last_month(self, sheet, column, value):
        sheet.merge_cells(f'{get_column_letter(column)}4:{get_column_letter(column + 4)}4')
        month = sheet.cell(row=4, column=column, value=value)
        month.alignment = self.align
        month.font = Font(name='Times New Roman', size=14)
        sheet.merge_cells(f'{get_column_letter(column)}5:{get_column_letter(column + 1)}5')
        profit = sheet.cell(row=5, column=column, value='Тушум')
        profit.alignment = self.align
        profit.font = Font(name='Times New Roman', size=11)
        start_month = sheet.cell(row=6, column=column, value='Кунлик')
        start_month.alignment = self.align
        start_month.font = Font(name='Times New Roman', size=11)
        start_month = sheet.cell(row=6, column=column + 1, value='Ой бошидан')
        start_month.alignment = self.align
        sheet.column_dimensions[f'{get_column_letter(column + 1)}'].width = len(start_month.value) + 2
        column += 2
        return column

    def write_month_title(self, sheet, column):
        sheet.merge_cells(f'{get_column_letter(column)}5:{get_column_letter(column)}6')
        task = sheet.cell(row=5, column=column, value='Топшириқ')
        task.alignment = self.align
        task.font = Font(name='Times New Roman', size=10)
        sheet.merge_cells(f'{get_column_letter(column + 1)}5:{get_column_letter(column + 1)}6')
        task = sheet.cell(row=5, column=column + 1, value='Амалда')
        task.alignment = self.align
        task.font = Font(name='Times New Roman', size=10)
        sheet.merge_cells(f'{get_column_letter(column + 2)}5:{get_column_letter(column + 2)}6')
        task = sheet.cell(row=5, column=column + 2, value='%')
        task.alignment = self.align
        task.font = Font(name='Times New Roman', size=10)

    def write_annual_data(self, sheet, end_column):
        build_coordinate = f'{get_column_letter(end_column)}3:{get_column_letter(end_column + 1)}4'
        sheet.merge_cells(build_coordinate)
        year_star_title = sheet.cell(row=3, column=end_column, value='Йил бошидан')
        year_star_title.alignment = self.align
        year_star_title.font = Font(name='Times New Roman', size=12)
        year_star_title.font = Font(name='Times New Roman', bold=True, size=10)
        sheet.merge_cells(f'{get_column_letter(end_column)}5:{get_column_letter(end_column)}6')
        sheet.merge_cells(f'{get_column_letter(end_column + 1)}5:{get_column_letter(end_column + 1)}6')
        income_title = sheet.cell(row=5, column=end_column, value='Тушум')
        income_title.alignment = self.align
        regarding_plan_title = sheet.cell(row=5, column=end_column + 1, value='Йиллик режага нисбатан %')
        regarding_plan_title.alignment = self.align
        regarding_plan_title.font = Font(name='Times New Roman', size=10)
