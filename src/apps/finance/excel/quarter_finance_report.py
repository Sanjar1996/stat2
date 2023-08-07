from datetime import datetime
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter, column_index_from_string
from ...accounts.models import Region
from .db_service import DBQuery
from .utils import QuarterUtils


class FinancePlanQuarterSheet(DBQuery, QuarterUtils):
    def __init__(self, user=None, amount_type=None, start=None, end=None):
        self.user = user
        self.at = amount_type
        self.start_date = start
        self.end_date = end
        self.align = Alignment(wrapText=True, horizontal='center', vertical='center')

    def generate_excel_finance_report(self):
        now = datetime.now()
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f"attachment; filename=daromad_kvartal_buyicha-{now.strftime('%d-%m-%Y')}.xlsx"
        wb = Workbook()
        sheet = wb.active
        align = Alignment(wrapText=True, horizontal='center', vertical='center')
        sheet.column_dimensions['A'].width = 5
        sheet.column_dimensions['B'].width = 55
        sheet.column_dimensions['C'].width = 14
        sheet.column_dimensions['D'].width = 14
        sheet.column_dimensions['E'].width = 14
        sheet.column_dimensions['F'].width = 14
        sheet.column_dimensions['G'].width = 14
        sheet.column_dimensions['H'].width = 13
        sheet.column_dimensions['I'].width = 13
        sheet.column_dimensions['J'].width = 13
        sheet.column_dimensions['K'].width = 13
        sheet.column_dimensions['M'].width = 13
        sheet.column_dimensions['N'].width = 13
        sheet.column_dimensions['P'].width = 13
        sheet.column_dimensions['S'].width = 13
        sheet.row_dimensions[1].height = 40
        sheet.row_dimensions[3].height = 22
        sheet.row_dimensions[4].height = 22
        sheet.row_dimensions[5].height = 25
        sheet.row_dimensions[6].height = 25

        sheet.merge_cells('A1:O2')
        sheet.merge_cells('A3:A6')
        sheet.merge_cells('B3:B6')
        year = f'{self.end_date[:4]} йил'
        title = sheet['A1']
        title.value = f"Ўрмон хўжалиги давлат қўмитаси тизим корхоналарида {year} ҳолатига даромад тушуми тўғрисида маълумот"
        title.font = Font(name='Times New Roman', bold=True, size=14)
        title.alignment = align

        number = sheet['A3']
        number.value = '№'
        number.alignment = align
        number.font = Font(name='Times New Roman', bold=True, size=11)

        department_name_title = sheet['B3']
        department_name_title.value = 'Ўрмон хўжаликлари'
        department_name_title.alignment = align
        department_name_title.font = Font(name='Times New Roman', bold=True, size=14)
        sheet.merge_cells('C3:C6')
        year_plan = sheet.cell(row=3, column=3, value='Йиллик режа')
        year_plan.alignment = align
        year_plan.font = Font(name='Times New Roman', bold=True, size=12)
        finance_db_data = []
        column = 4
        quarters = self.get_quarter_months(date=self.end_date)
        if quarters['quarter'] == 1:
            plan_merge_coordinate = f'{get_column_letter(4)}3:{get_column_letter(4)}4'
            sheet.merge_cells(plan_merge_coordinate)
            plan_title = sheet.cell(row=3, column=4, value='I - чорак режа')
            plan_title.alignment = align
            plan_title.font = Font(name='Times New Roman', bold=True, size=10)
            sheet.merge_cells('D5:D6')
            plan = sheet.cell(row=5, column=4, value='Режа')
            plan.alignment = align
            plan.font = Font(name='Times New Roman', bold=True, size=10)
            column += 1
            end_column = (quarters['count'] * 3) + column + 2
            quarter_title_coordinate = f'{get_column_letter(column)}3:{get_column_letter(end_column - 1)}3'
            sheet.merge_cells(quarter_title_coordinate)
            quarter_title = sheet.cell(row=3, column=column, value='I - чорак')
            quarter_title.alignment = align
            quarter_title.font = Font(name='Times New Roman', bold=True, size=14)
            for month_id, value in enumerate(quarters['month'], start=1):
                last_month = quarters['month'][-1]
                if last_month == value or last_month.get(month_id) == value.get(month_id):
                    column = self.write_last_month(sheet=sheet, column=column, value=value[month_id])
                    self.write_month_title(sheet=sheet, column=column)
                    column += 3
                    self.write_annual_data(sheet=sheet, end_column=end_column)
                    if month_id == 1:
                        finance_db_data = self.quarter_1_month_1(user=self.user, end=self.end_date, at=self.at)
                    elif month_id == 2:
                        finance_db_data = self.quarter_1_month_1_2(user=self.user, end=self.end_date, at=self.at)
                    elif month_id == 3:
                        finance_db_data = self.quarter_1_month_1_2_3(user=self.user, end=self.end_date, at=self.at)
                    else:
                        finance_db_data = []
                else:
                    self.write_months(sheet=sheet, column=column, value=value[month_id])
                    self.write_month_title(sheet=sheet, column=column)
                    column += 3
        elif quarters['quarter'] == 2:
            column = 4
            sheet.merge_cells('D3:F4')
            quarter_title = sheet.cell(row=3, column=column, value='I - чорак')
            quarter_title.alignment = align
            quarter_title.font = Font(name='Times New Roman', bold=True, size=14)
            sheet.merge_cells('D5:D6')
            sheet.merge_cells('E5:E6')
            sheet.merge_cells('F5:F6')
            task = sheet.cell(row=5, column=column, value='Топшириқ')
            task.alignment = align
            task.font = Font(name='Times New Roman', bold=True, size=10)
            task = sheet.cell(row=5, column=column + 1, value='Тушум')
            task.alignment = align
            task.font = Font(name='Times New Roman', bold=True, size=10)
            task = sheet.cell(row=5, column=column + 2, value='I-чорак режага нисбатан %')
            task.alignment = align
            task.font = Font(name='Times New Roman', bold=True, size=10)
            column += 3
            sheet.merge_cells(f'{get_column_letter(column)}3:{get_column_letter(column)}4')
            plan_title = sheet.cell(row=3, column=column, value='II - чорак режа')
            plan_title.alignment = align
            plan_title.font = Font(name='Times New Roman', bold=True, size=10)
            sheet.merge_cells('G5:G6')
            plan = sheet.cell(row=5, column=column, value='Режа')
            plan.alignment = align
            plan.font = Font(name='Times New Roman', bold=True, size=10)
            column += 1
            end_column = (quarters['count'] * 3) + column + 2
            sheet.merge_cells(f'{get_column_letter(column)}3:{get_column_letter(end_column - 1)}3')
            quarter_title = sheet.cell(row=3, column=column, value='II - чорак')
            quarter_title.alignment = align
            quarter_title.font = Font(name='Times New Roman', bold=True, size=14)
            for month_id, value in enumerate(quarters['month'], start=4):
                last_month = quarters['month'][-1]
                if last_month == value or last_month.get(month_id) == value.get(month_id):
                    column = self.write_last_month(sheet=sheet, column=column, value=value[month_id])
                    self.write_month_title(sheet=sheet, column=column)
                    column += 3
                    self.write_annual_data(sheet=sheet, end_column=end_column)
                    if month_id == 4:
                        finance_db_data = self.quarter_2_month_1(user=self.user, end=self.end_date, at=self.at)
                    elif month_id == 5:
                        finance_db_data = self.quarter_2_month_1_2(user=self.user, end=self.end_date, at=self.at)
                    elif month_id == 6:
                        finance_db_data = self.quarter_2_month_1_2_3(user=self.user, end=self.end_date, at=self.at)
                    else:
                        finance_db_data = []
                else:
                    self.write_months(sheet=sheet, column=column, value=value[month_id])
                    self.write_month_title(sheet=sheet, column=column)
                    column += 3
        elif quarters['quarter'] == 3:
            column = 4
            sheet.merge_cells('D3:F4')
            quarter_title = sheet.cell(row=3, column=column, value='6 - ойлик')
            quarter_title.alignment = align
            quarter_title.font = Font(name='Times New Roman', bold=True, size=14)
            sheet.merge_cells('D5:D6')
            sheet.merge_cells('E5:E6')
            sheet.merge_cells('F5:F6')
            task = sheet.cell(row=5, column=column, value='Топшириқ')
            task.alignment = align
            task.font = Font(name='Times New Roman', bold=True, size=10)
            task = sheet.cell(row=5, column=column + 1, value='Тушум')
            task.alignment = align
            task.font = Font(name='Times New Roman', bold=True, size=10)
            task = sheet.cell(row=5, column=column + 2, value='6-ойлик режага нисбатан %')
            task.alignment = align
            task.font = Font(name='Times New Roman', bold=True, size=10)
            column += 3
            sheet.merge_cells(f'{get_column_letter(column)}3:{get_column_letter(column)}4')
            plan_title = sheet.cell(row=3, column=column, value='III - чорак режа')
            plan_title.alignment = align
            plan_title.font = Font(name='Times New Roman', bold=True, size=10)
            sheet.merge_cells('G5:G6')
            plan = sheet.cell(row=5, column=column, value='Режа')
            plan.alignment = align
            plan.font = Font(name='Times New Roman', bold=True, size=10)
            column += 1
            end_column = (quarters['count'] * 3) + column + 2
            sheet.merge_cells(f'{get_column_letter(column)}3:{get_column_letter(end_column - 1)}3')
            quarter_title = sheet.cell(row=3, column=column, value='III - чорак')
            quarter_title.alignment = align
            quarter_title.font = Font(name='Times New Roman', bold=True, size=14)
            for month_id, value in enumerate(quarters['month'], start=7):
                last_month = quarters['month'][-1]
                if last_month == value or last_month.get(month_id) == value.get(month_id):
                    column = self.write_last_month(sheet=sheet, column=column, value=value[month_id])
                    self.write_month_title(sheet=sheet, column=column)
                    column += 3
                    self.write_annual_data(sheet=sheet, end_column=end_column)
                    if month_id == 7:
                        finance_db_data = self.quarter_3_month_1(user=self.user, end=self.end_date, at=self.at)
                    elif month_id == 8:
                        finance_db_data = self.quarter_3_month_1_2(user=self.user, end=self.end_date, at=self.at)
                    elif month_id == 9:
                        finance_db_data = self.quarter_3_month_1_2_3(user=self.user, end=self.end_date, at=self.at)
                    else:
                        finance_db_data = []
                else:
                    self.write_months(sheet=sheet, column=column, value=value[month_id])
                    self.write_month_title(sheet=sheet, column=column)
                    column += 3
        elif quarters['quarter'] == 4:
            column = 4
            sheet.merge_cells('D3:F4')
            quarter_title = sheet.cell(row=3, column=column, value='9 - ойлик')
            quarter_title.alignment = align
            quarter_title.font = Font(name='Times New Roman', bold=True, size=14)
            sheet.merge_cells('D5:D6')
            sheet.merge_cells('E5:E6')
            sheet.merge_cells('F5:F6')
            task = sheet.cell(row=5, column=column, value='Топшириқ')
            task.alignment = align
            task.font = Font(name='Times New Roman', bold=True, size=10)
            task = sheet.cell(row=5, column=column + 1, value='Тушум')
            task.alignment = align
            task.font = Font(name='Times New Roman', bold=True, size=10)
            task = sheet.cell(row=5, column=column + 2, value='9-ойлик режага нисбатан %')
            task.alignment = align
            task.font = Font(name='Times New Roman', bold=True, size=10)
            column += 3
            sheet.merge_cells(f'{get_column_letter(column)}3:{get_column_letter(column)}4')
            plan_title = sheet.cell(row=3, column=column, value='IV - чорак режа')
            plan_title.alignment = align
            plan_title.font = Font(name='Times New Roman', bold=True, size=10)
            sheet.merge_cells('G5:G6')
            plan = sheet.cell(row=5, column=column, value='Режа')
            plan.alignment = align
            plan.font = Font(name='Times New Roman', bold=True, size=10)
            column += 1
            end_column = (quarters['count'] * 3) + column + 2
            sheet.merge_cells(f'{get_column_letter(column)}3:{get_column_letter(end_column - 1)}3')
            quarter_title = sheet.cell(row=3, column=column, value='IV - чорак')
            quarter_title.alignment = align
            quarter_title.font = Font(name='Times New Roman', bold=True, size=14)
            for month_id, value in enumerate(quarters['month'], start=10):
                last_month = quarters['month'][-1]
                if last_month == value or last_month.get(month_id) == value.get(month_id):
                    column = self.write_last_month(sheet=sheet, column=column, value=value[month_id])
                    self.write_month_title(sheet=sheet, column=column)
                    column += 3
                    self.write_annual_data(sheet=sheet, end_column=end_column)
                    if month_id == 10:
                        finance_db_data = self.quarter_4_month_1(user=self.user, end=self.end_date, at=self.at)
                    elif month_id == 11:
                        finance_db_data = self.quarter_4_month_1_2(user=self.user, end=self.end_date, at=self.at)
                    elif month_id == 12:
                        finance_db_data = self.quarter_4_month_1_2_3(user=self.user, end=self.end_date, at=self.at)
                    else:
                        finance_db_data = []
                else:
                    self.write_months(sheet=sheet, column=column, value=value[month_id])
                    self.write_month_title(sheet=sheet, column=column)
                    column += 3
        region_row = []
        row_index = 7
        region_id = None
        article_id = 0
        for index, value in enumerate(finance_db_data):
            if region_id is None or int(value['region_id']) != region_id:
                article_id = 0
                region_row.append(row_index)
                region_id = self.write_region_data(
                    sheet=sheet, row=row_index, region_pk=value['region_id'], count=value['department_count'], quarter=quarters)
                row_index += 1
            article_id += 1
            department_id = sheet.cell(column=1, row=row_index, value=article_id)
            department_id.alignment = align
            department_id.font = Font(name='Times New Roman', size=11)
            if len(value['department']) > 50:
                name = f"{value['department'][:50]}..."
            else:
                name = value['department']
            department_name = sheet.cell(column=2, row=row_index, value=name)
            department_name.font = Font(name='Times New Roman', size=11)
            for value_id, val in enumerate(value.values()):
                if value_id > 2:
                    sheet.cell(column=value_id, row=row_index, value=val).alignment = align
            row_index += 1
        cell_finish_total_title = sheet.cell(row=row_index, column=2, value='Республика бўйича жами')
        cell_finish_total_title.font = Font(name='Times New Roman', bold=True, size=13)
        cell_finish_total_title.alignment = align

        if quarters['quarter'] == 1:
            month_index = quarters['count']
            if month_index == 1:
                for col in range(3, 11):
                    col_letter = get_column_letter(col).upper()
                    formula = '='
                    if col_letter == 'I':
                        self.calculate_region_percent(sheet=sheet, row=row_index, columns=['I'])
                        continue
                    for x in region_row:
                        formula += f'+{col_letter}{x}'
                    republic_cell = sheet.cell(column=col, row=row_index, value=formula)
                    republic_cell.alignment = self.align
                    republic_cell.font = Font(name='Times New Roman', bold=True, size=12)
                annual_percent = sheet.cell(
                    column=11, row=row_index,
                    value=f'=IF(OR(J{row_index}={0},C{row_index}={0}),{0},J{row_index}*100/C{row_index})')
                annual_percent.alignment = self.align
                annual_percent.font = Font(name='Times New Roman', bold=True, size=12)
                annual_percent.number_format = '0.0'
            elif month_index == 2:
                percent_columns = ['G', 'L']
                for col in range(3, 14):
                    col_letter = get_column_letter(col).upper()
                    formula = '='
                    if col_letter in percent_columns:
                        self.calculate_region_percent(sheet=sheet, row=row_index, columns=percent_columns)
                        continue
                    for x in region_row:
                        formula += f'+{col_letter}{x}'
                    republic_cell = sheet.cell(column=col, row=row_index, value=formula)
                    republic_cell.alignment = self.align
                    republic_cell.font = Font(name='Times New Roman', bold=True, size=12)
                annual_percent = sheet.cell(
                    column=14, row=row_index,
                    value=f'=IF(OR(M{row_index}={0},C{row_index}={0}),{0},M{row_index}*100/C{row_index})')
                annual_percent.alignment = self.align
                annual_percent.font = Font(name='Times New Roman', bold=True, size=12)
                annual_percent.number_format = '0.0'
            elif month_index == 3:
                percent_columns = ['G', 'J', 'O']
                for col in range(3, 17):
                    col_letter = get_column_letter(col).upper()
                    formula = '='
                    if col_letter in percent_columns:
                        self.calculate_region_percent(sheet=sheet, row=row_index, columns=percent_columns)
                        continue
                    for x in region_row:
                        formula += f'+{col_letter}{x}'
                    republic_cell = sheet.cell(column=col, row=row_index, value=formula)
                    republic_cell.alignment = self.align
                    republic_cell.font = Font(name='Times New Roman', bold=True, size=12)
                annual_percent = sheet.cell(
                    column=17, row=row_index,
                    value=f'=IF(OR(P{row_index}={0},C{row_index}={0}),{0},P{row_index}*100/C{row_index})')
                annual_percent.alignment = self.align
                annual_percent.font = Font(name='Times New Roman', bold=True, size=12)
                annual_percent.number_format = '0.0'
        else:
            month_index = quarters['count']
            if month_index == 1:
                percent_columns = ['F', 'L']
                for col in range(3, 14):
                    col_letter = get_column_letter(col).upper()
                    formula = '='
                    if col_letter in percent_columns:
                        self.calculate_region_percent(sheet=sheet, row=row_index, columns=percent_columns)
                        continue
                    for x in region_row:
                        formula += f'+{col_letter}{x}'
                    republic_cell = sheet.cell(column=col, row=row_index, value=formula)
                    republic_cell.alignment = self.align
                    republic_cell.font = Font(name='Times New Roman', bold=True, size=12)
                annual_percent = sheet.cell(
                    column=14, row=row_index,
                    value=f'=IF(OR(M{row_index}={0},C{row_index}={0}),{0},M{row_index}*100/C{row_index})')
                annual_percent.alignment = self.align
                annual_percent.font = Font(name='Times New Roman', bold=True, size=12)
                annual_percent.number_format = '0.0'
            elif month_index == 2:
                percent_columns = ['F', 'J', 'O']
                for col in range(3, 17):
                    col_letter = get_column_letter(col).upper()
                    formula = '='
                    if col_letter in percent_columns:
                        self.calculate_region_percent(sheet=sheet, row=row_index, columns=percent_columns)
                        continue
                    for x in region_row:
                        formula += f'+{col_letter}{x}'
                    republic_cell = sheet.cell(column=col, row=row_index, value=formula)
                    republic_cell.alignment = self.align
                    republic_cell.font = Font(name='Times New Roman', bold=True, size=12)
                annual_percent = sheet.cell(
                    column=17, row=row_index,
                    value=f'=IF(OR(P{row_index}={0},C{row_index}={0}),{0},P{row_index}*100/C{row_index})')
                annual_percent.alignment = self.align
                annual_percent.font = Font(name='Times New Roman', bold=True, size=12)
                annual_percent.number_format = '0.0'
            elif month_index == 3:
                percent_columns = ['F', 'J', 'M', 'R']
                for col in range(3, 20):
                    col_letter = get_column_letter(col).upper()
                    formula = '='
                    if col_letter in percent_columns:
                        self.calculate_region_percent(sheet=sheet, row=row_index, columns=percent_columns)
                        continue
                    for x in region_row:
                        formula += f'+{col_letter}{x}'
                    republic_cell = sheet.cell(column=col, row=row_index, value=formula)
                    republic_cell.alignment = self.align
                    republic_cell.font = Font(name='Times New Roman', bold=True, size=12)
                annual_percent = sheet.cell(
                    column=20, row=row_index,
                    value=f'=IF(OR(S{row_index}={0},C{row_index}={0}),{0},S{row_index}*100/C{row_index})')
                annual_percent.alignment = self.align
                annual_percent.font = Font(name='Times New Roman', bold=True, size=12)
                annual_percent.number_format = '0.0'

        def set_border(ws, cell_range):
            thin = Side(border_style="thin", color="000000")
            for row in ws[cell_range]:
                for cell in row:
                    cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
        set_border(sheet, f'A1:{get_column_letter(sheet.max_column)}{row_index}')
        sheet.freeze_panes = 'A7'
        wb.save(response)
        return response

    def write_region_data(self, sheet, row, region_pk, count, quarter):
        qs_region = Region.objects.filter(pk=region_pk, status=1)
        if qs_region.exists():
            begin_row = row + 1
            end_row = row + count
            region_id = qs_region.first().pk
            region_name = sheet.cell(column=2, row=row, value=qs_region.first().name)
            region_name.alignment = self.align
            region_name.font = Font(name='Times New Roman', bold=True, size=12)
            cell_c = sheet.cell(column=3, row=row, value=f'=SUM(C{begin_row}:C{end_row})')
            cell_c.alignment = self.align
            cell_c.font = Font(name='Times New Roman', bold=True, size=12)
            cell_d = sheet.cell(column=4, row=row, value=f'=SUM(D{begin_row}:D{end_row})')
            cell_d.alignment = self.align
            cell_d.font = Font(name='Times New Roman', bold=True, size=12)
            cell_e = sheet.cell(column=5, row=row, value=f'=SUM(E{begin_row}:E{end_row})')
            cell_e.alignment = self.align
            cell_e.font = Font(name='Times New Roman', bold=True, size=12)
            if quarter['quarter'] == 1:
                if quarter['count'] == 1:
                    percent_columns = ['I']
                    for col in range(6, 11):
                        col_letter = get_column_letter(col).upper()
                        if col_letter == 'I':
                            self.calculate_region_percent(sheet=sheet, row=row, columns=percent_columns)
                            continue
                        region_cell = sheet.cell(column=col, row=row, value=f'=SUM({col_letter}{begin_row}:{col_letter}{end_row})')
                        region_cell.alignment = self.align
                        region_cell.font = Font(name='Times New Roman', bold=True, size=12)
                    annual_percent = sheet.cell(
                        column=11, row=row, value=f'=IF(OR(J{row}={0},C{row}={0}),{0},J{row}*100/C{row})')
                    annual_percent.alignment = self.align
                    annual_percent.font = Font(name='Times New Roman', bold=True, size=12)
                    annual_percent.number_format = '0.0'
                elif quarter['count'] == 2:
                    percent_columns = ['G', 'L']
                    for col in range(6, 14):
                        col_letter = get_column_letter(col).upper()
                        if col_letter in percent_columns:
                            self.calculate_region_percent(sheet=sheet, row=row, columns=percent_columns)
                            continue
                        region_cell = sheet.cell(column=col, row=row, value=f'=SUM({col_letter}{begin_row}:{col_letter}{end_row})')
                        region_cell.alignment = self.align
                        region_cell.font = Font(name='Times New Roman', bold=True, size=12)
                    annual_percent = sheet.cell(
                        column=14, row=row, value=f'=IF(OR(M{row}={0},C{row}={0}),{0},M{row}*100/C{row})')
                    annual_percent.alignment = self.align
                    annual_percent.font = Font(name='Times New Roman', bold=True, size=12)
                    annual_percent.number_format = '0.0'
                elif quarter['count'] == 3:
                    percent_columns = ['G', 'J', 'O']
                    for col in range(6, 17):
                        col_letter = get_column_letter(col).upper()
                        if col_letter in percent_columns:
                            self.calculate_region_percent(sheet=sheet, row=row, columns=percent_columns)
                            continue
                        region_cell = sheet.cell(column=col, row=row, value=f'=SUM({col_letter}{begin_row}:{col_letter}{end_row})')
                        region_cell.alignment = self.align
                        region_cell.font = Font(name='Times New Roman', bold=True, size=12)
                    annual_percent = sheet.cell(
                        column=17, row=row, value=f'=IF(OR(P{row}={0},C{row}={0}),{0},P{row}*100/C{row})')
                    annual_percent.alignment = self.align
                    annual_percent.font = Font(name='Times New Roman', bold=True, size=12)
                    annual_percent.number_format = '0.0'
            else:
                if quarter['count'] == 1:
                    percent_columns = ['F', 'L']
                    for col in range(6, 14):
                        col_letter = get_column_letter(col).upper()
                        if col_letter in percent_columns:
                            self.calculate_region_percent(sheet=sheet, row=row, columns=percent_columns)
                            continue
                        region_cell = sheet.cell(column=col, row=row,
                                                 value=f'=SUM({col_letter}{begin_row}:{col_letter}{end_row})')
                        region_cell.alignment = self.align
                        region_cell.font = Font(name='Times New Roman', bold=True, size=12)
                    annual_percent = sheet.cell(
                        column=14, row=row, value=f'=IF(OR(M{row}={0},C{row}={0}),{0},M{row}*100/C{row})')
                    annual_percent.alignment = self.align
                    annual_percent.font = Font(name='Times New Roman', bold=True, size=12)
                    annual_percent.number_format = '0.0'
                elif quarter['count'] == 2:
                    percent_columns = ['F', 'J', 'O']
                    for col in range(6, 17):
                        col_letter = get_column_letter(col).upper()
                        if col_letter in percent_columns:
                            self.calculate_region_percent(sheet=sheet, row=row, columns=percent_columns)
                            continue
                        region_cell = sheet.cell(column=col, row=row,
                                                 value=f'=SUM({col_letter}{begin_row}:{col_letter}{end_row})')
                        region_cell.alignment = self.align
                        region_cell.font = Font(name='Times New Roman', bold=True, size=12)
                    annual_percent = sheet.cell(
                        column=17, row=row, value=f'=IF(OR(P{row}={0},C{row}={0}),{0},P{row}*100/C{row})')
                    annual_percent.alignment = self.align
                    annual_percent.font = Font(name='Times New Roman', bold=True, size=12)
                    annual_percent.number_format = '0.0'
                elif quarter['count'] == 3:
                    percent_columns = ['F', 'J', 'M', 'R']
                    for col in range(6, 20):
                        col_letter = get_column_letter(col).upper()
                        if col_letter in percent_columns:
                            self.calculate_region_percent(sheet=sheet, row=row, columns=percent_columns)
                            continue
                        region_cell = sheet.cell(column=col, row=row,
                                                 value=f'=SUM({col_letter}{begin_row}:{col_letter}{end_row})')
                        region_cell.alignment = self.align
                        region_cell.font = Font(name='Times New Roman', bold=True, size=12)
                    annual_percent = sheet.cell(
                        column=20, row=row, value=f'=IF(OR(S{row}={0},C{row}={0}),{0},S{row}*100/C{row})')
                    annual_percent.alignment = self.align
                    annual_percent.font = Font(name='Times New Roman', bold=True, size=12)
                    annual_percent.number_format = '0.0'
            return region_id

    def calculate_region_percent(self, sheet,  row, columns):
        for column in columns:
            idx = column_index_from_string(column)
            act = get_column_letter(idx-1)
            plan = get_column_letter(idx-2)
            percent_cell = sheet.cell(column=idx, row=row,
                                      value=f'=IF(OR({act}{row}={0},{plan}{row}={0}),{0},{act}{row}*100/{plan}{row})')
            percent_cell.alignment = self.align
            percent_cell.font = Font(name='Times New Roman', bold=True, size=12)
            percent_cell.number_format = '0.0'
