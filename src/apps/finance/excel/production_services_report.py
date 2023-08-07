from datetime import datetime
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter, column_index_from_string
from .db_service import ProductionDBQuery
from ...accounts.models import Region


class ProductionServicesSheet(ProductionDBQuery):
    def __init__(self, user=None, start=None, end=None):
        self.start_date = start
        self.end_date = end
        self.align = Alignment(wrapText=True, horizontal='center', vertical='center')
        self.user = user
        self.sheet = None
        self.align = None

    def generate_production_services_excel(self):
        now = datetime.now()
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f"attachment; filename=daromad_ishlab_chiqarish-{now.strftime('%d-%m-%Y')}.xlsx"
        wb = Workbook()
        sheet = wb.active
        align = Alignment(wrapText=True, horizontal='center', vertical='center')
        self.align = align
        sheet.column_dimensions['A'].width = 5
        sheet.column_dimensions['B'].width = 52
        sheet.row_dimensions[1].height = 40
        sheet.row_dimensions[2].height = 50
        sheet.row_dimensions[3].height = 40

        sheet.merge_cells('A1:J1')
        title = sheet['A1']

        title.value = f"Ўрмон маҳсулотлари ишлаб чиқариш ва аҳолига пуллик хизмат кўрсатиш тўғрисидаги " \
                      f"{self.end_date[:4]} йил бўйича МАЪЛУМОТ"
        title.font = Font(name='Times New Roman', bold=True, size=12)
        title.alignment = align

        sheet.merge_cells('A2:A3')
        number = sheet.cell(column=1, row=2, value='№')
        number.alignment = align
        number.font = Font(name='Times New Roman', bold=True, size=12)

        sheet.merge_cells('B2:B3')
        department_name = sheet.cell(column=2, row=2, value='Хўжалик номи')
        department_name.alignment = align
        department_name.font = Font(name='Times New Roman', bold=True, size=12)
        self.sheet = sheet

        self.write_header_title(col=3, value='Ўрмон маҳсулотларини ишлаб чиқариш')
        self.write_header_title(col=5, value='Пуллик хизмат кўсатиш')
        self.write_header_title(col=7, value='Умумий')
        sheet.merge_cells('I2:I3')
        total_percent_title = self.sheet.cell(column=9, row=2, value='Жами %')
        total_percent_title.alignment = self.align
        total_percent_title.font = Font(name='Times New Roman', bold=True, size=12)

        region_id = None
        region_row = []
        row_index = 4
        production_services_data = self.get_db_production_data(user=self.user, start=self.start_date, end=self.end_date)
        article_id = 0
        for index, value in enumerate(production_services_data):
            if region_id is None or int(value['region_id']) != region_id:
                article_id = 0
                region_row.append(row_index)
                region_id = self.write_region_data(
                    sheet=sheet, row=row_index, region_pk=value['region_id'], count=value['department_count'])
                row_index += 1
            article_id += 1
            department_id = sheet.cell(column=1, row=row_index, value=article_id)
            department_id.alignment = align
            department_id.font = Font(name='Times New Roman', size=12)
            if len(value['department']) > 48:
                name = f"{value['department'][:40]}..."
            else:
                name = value['department']
            department_name = sheet.cell(column=2, row=row_index, value=name)
            department_name.alignment = Alignment(wrapText=True, vertical='center')
            department_name.font = Font(name='Times New Roman', size=12)
            for value_id, val in enumerate(value.values()):
                if value_id > 2:
                    sheet.cell(column=value_id, row=row_index, value=val).alignment = align

            percent_formula = f'=IF(OR(H{row_index}={0},G{row_index}={0}),{0},H{row_index}*100/G{row_index})'
            general_percent_title = sheet.cell(column=9, row=row_index, value=percent_formula)
            general_percent_title.alignment = align
            general_percent_title.font = Font(name='Times New Roman', bold=True, size=12)
            general_percent_title.number_format = '0.0'
            row_index += 1

        cell_finish_total_title = sheet.cell(row=row_index, column=2, value='Республика бўйича жами')
        cell_finish_total_title.font = Font(name='Times New Roman', bold=True, size=13)
        cell_finish_total_title.alignment = align

        column_letters = []
        for item in range(3, 9):
            column_letters.append(get_column_letter(item))
        total_formula = '='
        for col in column_letters:
            for row in region_row:
                total_formula += f'+{col}{row}'
                cell_republic = sheet.cell(row=row_index, column=column_index_from_string(col), value=total_formula)
                cell_republic.alignment = align
                cell_republic.font = Font(name='Times New Roman', bold=True, size=13)
            total_formula = '='

        percent_formula = f'=IF(OR(H{row_index}={0},G{row_index}={0}),{0},H{row_index}*100/G{row_index})'
        general_percent_title = sheet.cell(column=9, row=row_index, value=percent_formula)
        general_percent_title.alignment = align
        general_percent_title.font = Font(name='Times New Roman', bold=True, size=12)
        general_percent_title.number_format = '0.0'

        def set_border(ws, cell_range):
            thin = Side(border_style="thin", color="000000")
            for rows in ws[cell_range]:
                for cell in rows:
                    cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
        set_border(sheet, f'A1:{get_column_letter(sheet.max_column)}{row_index}')
        sheet.freeze_panes = 'A4'
        wb.save(response)
        return response

    def write_header_title(self, col, value):
        begin_letter = get_column_letter(col)
        end_letter = get_column_letter(col + 1)
        self.sheet.merge_cells(f'{begin_letter}2:{end_letter}2')
        header_title = self.sheet.cell(column=col, row=2, value=value)
        header_title.alignment = self.align
        header_title.font = Font(name='Times New Roman', bold=True, size=12)
        self.sheet.column_dimensions[begin_letter].width = 18
        annual_plan_title = self.sheet.cell(column=col, row=3, value='Йиллик режа')
        annual_plan_title.alignment = self.align
        annual_plan_title.font = Font(name='Times New Roman', bold=True, size=11)
        self.sheet.column_dimensions[end_letter].width = 18
        annual_act_title = self.sheet.cell(column=col + 1, row=3, value='Амалда')
        annual_act_title.alignment = self.align
        annual_act_title.font = Font(name='Times New Roman', bold=True, size=11)

    def write_region_data(self, sheet, row, region_pk, count):
        qs_region = Region.objects.filter(pk=region_pk, status=1)
        if qs_region.exists():
            begin_row = row + 1
            end_row = row + count
            region_id = qs_region.first().pk
            region_name = sheet.cell(column=2, row=row, value=qs_region.first().name)
            region_name.alignment = self.align
            region_name.font = Font(name='Times New Roman', bold=True, size=12)
            for col in range(3, 9):
                letter = get_column_letter(col)
                total_by_region = sheet.cell(column=col, row=row, value=f'=SUM({letter}{begin_row}:{letter}{end_row})')
                total_by_region.alignment = self.align
                total_by_region.font = Font(name='Times New Roman', bold=True, size=11)

            general_percent_title = sheet.cell(column=9, row=row, value=f'=IF(OR(H{row}={0},G{row}={0}),{0},H{row}*100/G{row})')
            general_percent_title.alignment = self.align
            general_percent_title.font = Font(name='Times New Roman', bold=True, size=12)
            general_percent_title.number_format = '0.0'
            return region_id
