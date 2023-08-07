from datetime import datetime
from django.db import connection
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter, column_index_from_string
from ...accounts.models import Region, Department, UserDepartment


class FinanceGeneralSheet:
    def __init__(self, user=None, start=None, end=None, amount_type=None):
        self.start_date = start
        self.end_date = end
        self.at = amount_type
        self.user_id = user
        self.align = Alignment(wrapText=True, horizontal='center', vertical='center')

    def generate_excel_general_finance_report(self):
        now = datetime.now()
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f"attachment; filename=daromad_reja_amal-{now.strftime('%d-%m-%Y')}.xlsx"
        wb = Workbook()
        sheet = wb.active
        align = Alignment(wrapText=True, horizontal='center', vertical='center')
        sheet.column_dimensions['A'].width = 5
        sheet.column_dimensions['B'].width = 80
        sheet.column_dimensions['C'].width = 15
        sheet.column_dimensions['D'].width = 17
        sheet.column_dimensions['E'].width = 20
        sheet.row_dimensions[1].height = 40
        sheet.row_dimensions[2].height = 40

        sheet.merge_cells('A1:E1')
        title = sheet['A1']
        title.value = f"Ўрмон хўжалиги давлат қўмитаси ташкилотларининг {self.start_date[:4]} йил режаси тўғрисида МАЪЛУМОТ"
        title.font = Font(name='Times New Roman', bold=True, size=12)
        title.alignment = align

        number = sheet.cell(column=1, row=2, value='№')
        number.alignment = align
        number.font = Font(name='Times New Roman', bold=True, size=12)

        department_name = sheet.cell(column=2, row=2, value='Хўжалик номи')
        department_name.alignment = align
        department_name.font = Font(name='Times New Roman', bold=True, size=12)

        annual_plan_title = sheet.cell(column=3, row=2, value='Йиллик режа')
        annual_plan_title.alignment = align
        annual_plan_title.font = Font(name='Times New Roman', size=11)

        actual_title = sheet.cell(column=4, row=2, value='Амалда')
        actual_title.alignment = align
        actual_title.font = Font(name='Times New Roman', size=11)

        percent_title = sheet.cell(column=5, row=2, value='Режага нисбатан %')
        percent_title.alignment = align
        percent_title.font = Font(name='Times New Roman', size=11)

        region_id = None
        region_row = []
        row_index = 3
        departments = self.prepare_finance_by_department_raw_sql()
        article_id = 0
        qs_user_department = UserDepartment.objects.filter(user_id=self.user_id)
        if qs_user_department.exists() and departments:
            for index, value in enumerate(departments):
                if region_id is None or int(value[2]) != region_id:
                    article_id = 0
                    region_row.append(row_index)
                    region_id = self.write_region_data(sheet=sheet, row=row_index, region_pk=value[2])
                    row_index += 1
                article_id += 1
                department_id = sheet.cell(column=1, row=row_index, value=article_id)
                department_id.alignment = align
                department_id.font = Font(name='Times New Roman', size=12)
                if len(value[1]) > 65:
                    name = f'{value[1][:65]}...'
                else:
                    name = value[1]
                department_name = sheet.cell(column=2, row=row_index, value=name)
                department_name.alignment = Alignment(wrapText=True, vertical='center')
                department_name.font = Font(name='Times New Roman', size=12)

                annual_plan_cell = sheet.cell(column=3, row=row_index, value=value[3])
                annual_plan_cell.alignment = align
                annual_plan_cell.font = Font(name='Times New Roman', size=11)

                actual_cell = sheet.cell(column=4, row=row_index, value=value[4])
                actual_cell.alignment = align
                actual_cell.font = Font(name='Times New Roman', size=11)
                # percent_formula = f'=IF(D{row_index}={0},{0},D{row_index}*100/C{row_index})'
                percent_formula = f'=IF(OR(C{row_index}={0},D{row_index}={0}),{0},D{row_index}*100/C{row_index})'
                percent_cell = sheet.cell(column=5, row=row_index, value=percent_formula)
                percent_cell.alignment = align
                percent_cell.font = Font(name='Times New Roman', size=11)
                percent_cell.number_format = '0.0'
                row_index += 1

        cell_finish_total_title = sheet.cell(row=row_index, column=2, value='Республика бўйича жами')
        cell_finish_total_title.font = Font(name='Times New Roman', bold=True, size=13)
        cell_finish_total_title.alignment = align
        column_letters = []
        for item in range(3, 5):
            column_letters.append(get_column_letter(item))
        total_formula_by_region = '='
        for col in column_letters:
            for row in region_row:
                total_formula_by_region += f'+{col}{row}'
                _idx = column_index_from_string(col)
                cell_republic_task_formula = sheet.cell(row=row_index, column=_idx, value=total_formula_by_region)
                cell_republic_task_formula.alignment = align
                cell_republic_task_formula.font = Font(name='Times New Roman', bold=True, size=13)
            total_formula_by_region = '='

        republic_formula = f'=IF(D{row_index}={0},{0},D{row_index}*100/C{row_index})'
        percent_republic = sheet.cell(row=row_index, column=5, value=republic_formula)
        percent_republic.alignment = align
        percent_republic.font = Font(name='Times New Roman', bold=True, size=13)
        percent_republic.number_format = '0.0'

        def set_border(ws, cell_range):
            thin = Side(border_style="thin", color="000000")
            for row in ws[cell_range]:
                for cell in row:
                    cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
        set_border(sheet, f'A1:{get_column_letter(sheet.max_column)}{row_index}')
        sheet.freeze_panes = 'A3'
        wb.save(response)
        return response

    def write_region_data(self, sheet, row, region_pk):
        qs_region = Region.objects.filter(pk=region_pk, status=1)
        if qs_region.exists():
            qs_user_department = UserDepartment.objects.filter(user_id=self.user_id)
            count = qs_user_department[0].departments.filter(region_id=region_pk, status=1).count()
            # count = Department.objects.filter(region_id=region_pk, status=1).count()
            begin_row = row + 1
            end_row = row + count
            region_id = qs_region.first().pk
            region_name = sheet.cell(column=2, row=row, value=qs_region.first().name)
            region_name.alignment = self.align
            region_name.font = Font(name='Times New Roman', bold=True, size=12)
            plan_by_region = sheet.cell(column=3, row=row, value=f'=SUM(C{begin_row}:C{end_row})')
            plan_by_region.alignment = self.align
            plan_by_region.font = Font(name='Times New Roman', bold=True, size=12)
            act_by_region = sheet.cell(column=4, row=row, value=f'=SUM(D{begin_row}:D{end_row})')
            act_by_region.alignment = self.align
            act_by_region.font = Font(name='Times New Roman', bold=True, size=12)
            percent_by_region = sheet.cell(column=5, row=row, value=f'=IF(D{row}={0},{0},D{row}*100/C{row})')
            percent_by_region.alignment = self.align
            percent_by_region.number_format = '0.0'
            percent_by_region.font = Font(name='Times New Roman', bold=True, size=12)
            return region_id

    def prepare_finance_by_department_raw_sql(self):
        with connection.cursor() as cursor:
            start = datetime.strptime(self.start_date, '%Y-%m-%d')
            end = datetime.strptime(self.end_date, '%Y-%m-%d')
            amount_type_id = self.at
            query = f"""select report.id, report.name, report.region_id, sum(y_reja), sum(amalda) 
                        from 
                            ((select d.id, d.name, d.region_id, sum(reja.amount) y_reja, 0 amalda, d.sort
                            from department d
                            left join user_departments_departments udd ON udd.department_id = d.id 
                            left join user_departments ud on udd.userdepartment_id = ud.id
                            left join finance_plan reja on reja.department_id = d.id 
                            and reja.date BETWEEN '{start}' and '{end}' 
                            and reja.amount_type = {amount_type_id}
                            and reja.status = 1 
                            where d.status = 1 and ud.user_id = {self.user_id} 
                            group by d.id order by d.sort)
                            UNION ALL
                            (select d.id, d.name, d.region_id, 0 y_reja, sum(actual.amount) amalda, d.sort
                            from department d
                            left join user_departments_departments udd ON udd.department_id = d.id 
                            left join user_departments ud on udd.userdepartment_id = ud.id
                            left join finance actual on actual.department_id = d.id 
                            and actual.date BETWEEN '{start}' and '{end}' 
                            and actual.amount_type = {amount_type_id}
                            and actual.status = 1 
                            and actual.state = 2 
                            where d.status = 1 and ud.user_id = {self.user_id} 
                            group by d.id order by d.sort)) 
                        report
                        group by report.id, report.name, report.region_id,  report.sort order by report.sort"""
            cursor.execute(query)
            row = cursor.fetchall()
            return row
