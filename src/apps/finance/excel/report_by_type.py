from datetime import datetime
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter, column_index_from_string
from .db_service import FinanceDBQuery
from ..models import FinanceType
from ...accounts.models import Region


class FinanceByTypeSheet(FinanceDBQuery):
    def __init__(self, start=None, end=None, amount_type=None, user=None):
        self.start_date = start
        self.end_date = end
        self.user = user
        self.at = amount_type
        self.align = Alignment(wrapText=True, horizontal='center', vertical='center')

    def generate_excel_finance_report(self):
        now = datetime.now()
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f"attachment; filename=daromad_turlari-{now.strftime('%d-%m-%Y')}.xlsx"
        wb = Workbook()
        sheet = wb.active
        align = Alignment(wrapText=True, horizontal='center', vertical='center')
        sheet.column_dimensions['A'].width = 5
        sheet.column_dimensions['B'].width = 55
        sheet.row_dimensions[1].height = 40
        sheet.row_dimensions[2].height = 100

        sheet.merge_cells('A1:Q1')
        if self.start_date[:4] == self.end_date[:4]:
            year = f'{self.start_date[:4]} йил'
        else:
            year = f'{self.start_date[:4]}-{self.end_date[:4]} йиллар'
        title = sheet['A1']
        title.value = f"Ўрмон хўжалиги давлат қўмитаси ташкилотларининг {year} давомида топилган даромадлари тўғрисида МАЪЛУМОТ"
        title.font = Font(name='Times New Roman', bold=True, size=14)
        title.alignment = align

        number = sheet.cell(column=1, row=2, value='№')
        number.alignment = align
        number.font = Font(name='Times New Roman', bold=True, size=12)

        department_name = sheet.cell(column=2, row=2, value='Хўжалик номи')
        department_name.alignment = align
        department_name.font = Font(name='Times New Roman', bold=True, size=14)
        all_types = FinanceType.objects.filter(status=1).order_by('id')
        header_col = 3
        for ty_index, ty_item in enumerate(all_types):
            ty_name = sheet.cell(row=2, column=header_col, value=ty_item.name)
            ty_name.alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
            ty_name.font = Font(name='Times New Roman', bold=True, size=11)
            header_col += 1
        ty_name = sheet.cell(row=2, column=header_col, value='Жами')
        ty_name.alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
        ty_name.font = Font(name='Times New Roman', bold=True, size=11)

        region_id = None
        region_row = []
        row_index = 3
        profit_by_type = self.get_finance_profit(
            user=self.user, qs=all_types, start=self.start_date, end=self.end_date, at=self.at)
        article_id = 0
        for index, value in enumerate(profit_by_type):
            start_col = 3
            if region_id is None or int(value['region_id']) != region_id:
                article_id = 0
                region_row.append(row_index)
                region_id = self.write_region_data(
                    sheet=sheet, row=row_index, region_pk=int(value['region_id']),
                    count=value['department_count'], length=value['profits'])
                row_index += 1
            article_id += 1
            department_id = sheet.cell(column=1, row=row_index, value=article_id)
            department_id.alignment = align
            department_id.font = Font(name='Times New Roman', size=12)
            if len(value['department']) > 45:
                name = f"{value['department'][:43]}..."
            else:
                name = value['department']
            department_name = sheet.cell(column=2, row=row_index, value=name)
            department_name.alignment = Alignment(wrapText=True, vertical='center')
            department_name.font = Font(name='Times New Roman', size=12)
            for profit in value['profits']:
                finance_type_cell = sheet.cell(column=start_col, row=row_index, value=profit)
                finance_type_cell.alignment = align
                start_col += 1
            end_col = get_column_letter(header_col - 1)
            total_row = sheet.cell(column=header_col, row=row_index, value=f'=SUM(C{row_index}:{end_col}{row_index})')
            total_row.alignment = align
            total_row.font = Font(name='Times New Roman', bold=True, size=11)
            row_index += 1

        cell_finish_total_title = sheet.cell(row=row_index, column=2, value='Республика бўйича жами')
        cell_finish_total_title.font = Font(name='Times New Roman', bold=True, size=13)
        cell_finish_total_title.alignment = align
        column_letters = []
        for item in range(3, header_col + 1):
            column_letters.append(get_column_letter(item))
        total_formula_by_region = '='
        for col in column_letters:
            sheet.column_dimensions[f'{col}'].width = 11
            for row in region_row:
                total_formula_by_region += f'+{col}{row}'
                _idx = column_index_from_string(col)
                cell_republic_task_formula = sheet.cell(row=row_index, column=_idx, value=total_formula_by_region)
                cell_republic_task_formula.alignment = align
                cell_republic_task_formula.font = Font(name='Times New Roman', bold=True, size=13)
            total_formula_by_region = '='

        def set_border(ws, cell_range):
            thin = Side(border_style="thin", color="000000")
            for row in ws[cell_range]:
                for cell in row:
                    cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

        set_border(sheet, f'A1:{get_column_letter(sheet.max_column)}{row_index}')
        sheet.freeze_panes = 'A3'
        wb.save(response)
        return response

    def write_region_data(self, sheet, row, region_pk, count, length):
        qs_region = Region.objects.filter(pk=region_pk, status=1)
        start_col = 3
        if qs_region.exists():
            begin_row = row + 1
            end_row = row + count
            region_id = qs_region.first().pk
            region_name = sheet.cell(column=2, row=row, value=qs_region.first().name)
            region_name.alignment = self.align
            region_name.font = Font(name='Times New Roman', bold=True, size=12)
            for x in range(0, len(length)):
                column_letter = get_column_letter(start_col)
                formula_by_region = f'=SUM({column_letter}{begin_row}:{column_letter}{end_row})'
                finance_type_cell = sheet.cell(column=start_col, row=row, value=formula_by_region)
                finance_type_cell.alignment = self.align
                finance_type_cell.font = Font(name='Times New Roman', bold=True, size=12)
                start_col += 1
            end_col = get_column_letter(start_col - 1)
            total_row = sheet.cell(column=start_col, row=row,
                                   value=f'=SUM(C{row}:{end_col}{row})')
            total_row.alignment = self.align
            total_row.font = Font(name='Times New Roman', bold=True, size=13)
            return region_id



