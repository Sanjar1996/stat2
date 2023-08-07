from datetime import datetime
from django.db import connection
from django.http import HttpResponse
from openpyxl import Workbook
from .utils import *
from ..models import Animal, AnimalCategory, ChorvaInputOutputCategory


class ChorvaInputOutputSheet:
    def __init__(self, user, end=None, obj=None):
        self.user = user
        self.end_date = end
        self.department = obj
        self.align = Alignment(wrapText=True, horizontal='center', vertical='center')
        self.head_col = 0
        self.scales_col = 0

    def generate_input_output_excel_report(self):
        now = datetime.now()
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f"attachment; filename=chorva_kirim_chiqim-{now.strftime('%d-%m-%Y')}.xlsx"
        wb = Workbook()
        sheet = wb.active
        align = Alignment(wrapText=True, horizontal='center', vertical='center')
        sheet.column_dimensions['A'].width = 5
        sheet.column_dimensions['B'].width = 25
        sheet.row_dimensions[1].height = 50
        sheet.row_dimensions[3].height = 25
        sheet.row_dimensions[4].height = 120

        sheet.merge_cells('A1:AF1')
        quarter = format_date_by_quarter(end=self.end_date)
        title = sheet.cell(row=1, column=1,
                           value=f"{self.department.name} нинг {self.end_date[:4]} йил {quarter['quarter']}"
                                 f" - квартал холатига бўлган давридаги пода харакати ва туғим якуни "
                                 f"тўғрисида МАЪЛУМОТНОМА.")
        title.font = Font(name='Times New Roman', bold=True, size=14)
        title.alignment = align
        sheet.merge_cells('A3:A5')
        number = sheet.cell(column=1, row=3, value='№')
        number.alignment = align
        number.font = Font(name='Times New Roman', bold=True, size=12)
        sheet.merge_cells('B3:B5')
        chorva_name_title = sheet.cell(column=2, row=3, value='Чорва ҳайвонлари')
        chorva_name_title.alignment = align
        chorva_name_title.font = Font(name='Times New Roman', bold=True, size=14)
        sheet.merge_cells('C3:D4')
        residue_title = sheet.cell(column=3, row=3, value=f"{self.end_date[:4]}-01-01 йил ҳолатига мавжуд")
        residue_title.alignment = align
        residue_title.font = Font(name='Times New Roman', bold=True, size=12)
        set_column_width(sheet=sheet, index=3)
        write_top_panel_footer(sheet=sheet, col=3)

        input_category = ChorvaInputOutputCategory.objects.filter(type=1, status=1)
        output_category = ChorvaInputOutputCategory.objects.filter(type=2, status=1)

        input_title_end_col = (input_category.count() * 2) + 6
        merge_columns(sheet=sheet, begin=5, end=input_title_end_col, row=3)

        input_title = sheet.cell(row=3, column=5, value='К И Р И М')
        input_title.alignment = align
        input_title.font = Font(name='Times New Roman', bold=True, size=12)

        output_title_start_col = input_title_end_col + 1
        output_title_end_col = output_category.count() * 2 + input_title_end_col + 2
        merge_columns(sheet=sheet, begin=output_title_start_col, end=output_title_end_col, row=3)

        output_title = sheet.cell(row=3, column=output_title_start_col, value='Ч И Қ И М ')
        output_title.alignment = align
        output_title.font = Font(name='Times New Roman', bold=True, size=12)

        column_start = chorva_inp_out_category(
            sheet=sheet, category=input_category, col=5, value='жами кирим қилинган чорва ҳайвонлар')
        self.head_col = column_start - 2
        column_end = chorva_inp_out_category(
            sheet=sheet, category=output_category, col=column_start, value='жами чиким қилинган чорва ҳайвонлар')
        self.scales_col = column_end - 2
        merge_columns(sheet=sheet, begin=column_end, end=column_end + 1, row=4)
        cell_input_output_total = sheet.cell(row=4, column=column_end, value=f"{self.end_date} йил ҳолатига мавжуд")
        cell_input_output_total.alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
        cell_input_output_total.font = Font(name='Times New Roman', bold=True, size=11)
        set_column_width(sheet=sheet, index=column_end)
        sheet.merge_cells(f'{get_column_letter(column_end - 2)}2:{get_column_letter(column_end + 1)}2')
        centner_total = sheet.cell(row=2, column=column_end - 2, value='Центнер хисобида')
        centner_total.alignment = align
        centner_total.font = Font(name='Times New Roman', bold=True, size=12)
        write_top_panel_footer(sheet=sheet, col=column_end)

        # # INPUT CATEGORY
        col_index = 5
        input_last_pk = 0
        output_last_pk = 0
        row_index = 6
        animal_row = 6
        category_id = 0
        input_head_letters = ''
        input_scales_letters = ''
        output_head_letters = ''
        output_scales_letters = ''
        mshm_head_row = '='
        mshm_scales_row = '='
        mshm_row = set()
        article_id = 0
        for ic_index, ic_value in enumerate(input_category):
            input_last_pk = ic_value.pk
            if ic_index == 0:
                animal_name_data = self.animal_name_raw_sql()
                for animal_id, animal_data in enumerate(animal_name_data):
                    article_id += 1
                    category_id = animal_data[1]
                    sheet.cell(row=animal_row, column=1, value=article_id)
                    sheet.cell(row=animal_row, column=2, value=animal_data[2])
                    old_year_head = sheet.cell(row=animal_row, column=3, value=animal_data[3])
                    old_year_head.alignment = align
                    old_year_scales = sheet.cell(row=animal_row, column=4, value=animal_data[4])
                    old_year_scales.alignment = align
                    animal_row += 1
                    if animal_id == len(animal_name_data) - 1 or category_id != animal_name_data[animal_id + 1][1]:
                        article_id = 0
                        count = Animal.objects.filter(category_id=category_id).count()
                        begin_row = (animal_row - count)
                        end_row = animal_row - 1
                        qs = AnimalCategory.objects.filter(id=category_id)
                        if qs.exists():
                            category_name = sheet.cell(column=2, row=animal_row, value=qs.first().name)
                            cell_head_of_category = sheet.cell(
                                column=3, row=animal_row, value=f'=SUM(C{begin_row}:C{end_row})')
                            cell_head_of_category.alignment = align
                            cell_head_of_category.font = Font(name='Times New Roman', bold=True, size=11)
                            mshm_head_row += f'+{cell_head_of_category.column_letter}{animal_row}'

                            cell_scales_of_category = sheet.cell(
                                column=4, row=animal_row, value=f'=SUM(D{begin_row}:D{end_row})')
                            cell_scales_of_category.alignment = align
                            cell_scales_of_category.font = Font(name='Times New Roman', bold=True, size=11)
                            mshm_scales_row += f'+{cell_scales_of_category.column_letter}{animal_row}'

                            category_name.alignment = align
                            category_name.font = Font(name='Times New Roman', bold=True, size=12)
                            animal_row += 1
                            if qs.first().name == "Эчкилар":
                                root_total_mshm = sheet.cell(column=2, row=animal_row, value="Жами МШМ")
                                root_total_mshm.alignment = align
                                root_total_mshm.font = Font(name='Times New Roman', bold=True, size=12)
                                mshm_head_total = sheet.cell(column=3, row=animal_row, value=mshm_head_row)
                                mshm_head_total.alignment = align
                                mshm_head_total.font = Font(name='Times New Roman', bold=True, size=11)
                                mshm_scales_total = sheet.cell(column=4, row=animal_row, value=mshm_scales_row)
                                mshm_scales_total.alignment = align
                                mshm_scales_total.font = Font(name='Times New Roman', bold=True, size=11)
                                animal_row += 1
            chorva_input_data = self.chorva_input_raw_sql(category=ic_value.pk)
            for input_id, input_data in enumerate(chorva_input_data):
                category_id = input_data[0]
                inp_head_cell = sheet.cell(row=row_index, column=col_index, value=input_data[1])
                inp_head_cell.alignment = align
                inp_scales_cell = sheet.cell(row=row_index, column=col_index + 1, value=input_data[2])
                inp_scales_cell.alignment = align
                if inp_head_cell.column_letter not in input_head_letters:
                    input_head_letters += f'{inp_head_cell.column_letter}-'
                if inp_scales_cell.column_letter not in input_scales_letters:
                    input_scales_letters += f'{inp_scales_cell.column_letter}-'
                row_index += 1
                if input_id == len(chorva_input_data) - 1 or category_id != chorva_input_data[input_id + 1][0]:
                    count = Animal.objects.filter(category_id=category_id).count()
                    begin_row = (row_index - count)
                    end_row = row_index - 1
                    qs = AnimalCategory.objects.filter(id=category_id)
                    if qs.exists():
                        begin_letter = get_column_letter(col_index)
                        head_formula = f'=SUM({begin_letter}{begin_row}:{begin_letter}{end_row})'
                        cell_total_head = sheet.cell(row=row_index, column=col_index, value=head_formula)
                        cell_total_head.alignment = align
                        cell_total_head.font = Font(name='Times New Roman', bold=True, size=11)
                        if qs.first().name == "Қойлар" or qs.first().name == "Эчкилар":
                            mshm_row.add(row_index)
                        end_letter = get_column_letter(col_index + 1)
                        scales_formula = f'=SUM({end_letter}{begin_row}:{end_letter}{end_row})'
                        cell_total_scales = sheet.cell(row=row_index, column=col_index + 1, value=scales_formula)
                        cell_total_scales.alignment = align
                        cell_total_scales.font = Font(name='Times New Roman', bold=True, size=11)
                        row_index += 1
                        if qs.first().name == "Эчкилар":
                            input_mshm_head_total_formula = '='
                            for row in mshm_row:
                                input_mshm_head_total_formula += f'+{get_column_letter(col_index)}{row}'
                            mshm_head_cell = sheet.cell(row=row_index, column=col_index,
                                                        value=input_mshm_head_total_formula)
                            mshm_head_cell.alignment = align
                            mshm_head_cell.font = Font(name='Times New Roman', bold=True, size=11)
                            input_mshm_scales_total_formula = '='
                            for row in mshm_row:
                                input_mshm_scales_total_formula += f'+{get_column_letter(col_index + 1)}{row}'
                            mshm_scales_cell = sheet.cell(row=row_index, column=col_index + 1,
                                                          value=input_mshm_scales_total_formula)
                            mshm_scales_cell.alignment = align
                            mshm_scales_cell.font = Font(name='Times New Roman', bold=True, size=11)
                            row_index += 1
            col_index += 2
            row_index = 6
        # WRITE EACH INPUT TOTAL
        total_input_data = self.chorva_input_raw_sql(category=input_last_pk)
        for total_id, total_data in enumerate(total_input_data):
            category_id = total_data[0]

            input_head_formula = input_head_letters.replace('-', f'{row_index}+')
            cell_total_head = sheet.cell(row=row_index, column=col_index, value=f'=+{input_head_formula[:-1]}')
            cell_total_head.alignment = align
            cell_total_head.font = Font(name='Times New Roman', bold=True)

            input_scales_formula = input_scales_letters.replace('-', f'{row_index}+')
            cell_total_scales = sheet.cell(row=row_index, column=col_index + 1, value=f'=+{input_scales_formula[:-1]}')
            cell_total_scales.alignment = align
            cell_total_scales.font = Font(name='Times New Roman', bold=True)

            row_index += 1
            if total_id == len(total_input_data) - 1 or category_id != total_input_data[total_id + 1][0]:
                count = Animal.objects.filter(category_id=category_id).count()
                begin_row = (row_index - count)
                end_row = row_index - 1
                qs = AnimalCategory.objects.filter(id=category_id)
                if qs.exists():
                    begin_letter = get_column_letter(col_index)
                    total_head_formula = f'=SUM({begin_letter}{begin_row}:{begin_letter}{end_row})'
                    total_category_head = sheet.cell(row=row_index, column=col_index, value=total_head_formula)
                    total_category_head.alignment = align
                    total_category_head.font = Font(name='Times New Roman', bold=True, size=11)
                    end_letter = get_column_letter(col_index + 1)
                    total_scales_formula = f'=SUM({end_letter}{begin_row}:{end_letter}{end_row})'
                    total_category_scales = sheet.cell(row=row_index, column=col_index + 1, value=total_scales_formula)
                    total_category_scales.alignment = align
                    total_category_scales.font = Font(name='Times New Roman', bold=True, size=11)
                    row_index += 1
                    if qs.first().name == "Эчкилар":
                        total_inp_mshm_head_formula = '='
                        for row in mshm_row:
                            total_inp_mshm_head_formula += f'+{get_column_letter(col_index)}{row}'
                        total_inp_mshm_head = sheet.cell(
                            row=row_index, column=col_index, value=total_inp_mshm_head_formula)
                        total_inp_mshm_head.alignment = align
                        total_inp_mshm_head.font = Font(name='Times New Roman', bold=True, size=11)

                        total_inp_mshm_scales_formula = '='
                        for row in mshm_row:
                            total_inp_mshm_scales_formula += f'+{get_column_letter(col_index + 1)}{row}'
                        total_inp_mshm_scales = sheet.cell(
                            row=row_index, column=col_index + 1, value=total_inp_mshm_scales_formula)
                        total_inp_mshm_scales.alignment = align
                        total_inp_mshm_scales.font = Font(name='Times New Roman', bold=True, size=11)
                        row_index += 1
        col_index += 2
        row_index = 6
        # OUTPUT CATEGORY
        category_id = 0
        for oc_index, oc_value in enumerate(output_category):
            output_last_pk = oc_value.pk
            chorva_output_data = self.chorva_output_raw_sql(category=oc_value.pk)
            for output_id, output_data in enumerate(chorva_output_data):
                category_id = output_data[0]
                output_head_cell = sheet.cell(row=row_index, column=col_index, value=output_data[1])
                output_head_cell.alignment = align
                output_scales_cell = sheet.cell(row=row_index, column=col_index + 1, value=output_data[2])
                output_scales_cell.alignment = align
                if output_head_cell.column_letter not in output_head_letters:
                    output_head_letters += f'{output_head_cell.column_letter}-'
                if output_scales_cell.column_letter not in output_scales_letters:
                    output_scales_letters += f'{output_scales_cell.column_letter}-'
                row_index += 1
                if output_id == len(chorva_output_data) - 1 or category_id != chorva_output_data[output_id + 1][0]:
                    count = Animal.objects.filter(category_id=category_id).count()
                    begin_row = (row_index - count)
                    end_row = row_index - 1
                    qs = AnimalCategory.objects.filter(id=category_id)
                    if qs.exists():
                        begin_letter = get_column_letter(col_index)
                        output_head_formula = f'=SUM({begin_letter}{begin_row}:{begin_letter}{end_row})'
                        output_category_head = sheet.cell(row=row_index, column=col_index, value=output_head_formula)
                        output_category_head.alignment = align
                        output_category_head.font = Font(name='Times New Roman', bold=True, size=11)
                        end_letter = get_column_letter(col_index + 1)
                        output_scales_formula = f'=SUM({end_letter}{begin_row}:{end_letter}{end_row})'
                        output_category_scales = sheet.cell(
                            row=row_index, column=col_index + 1, value=output_scales_formula)
                        output_category_scales.alignment = align
                        output_category_scales.font = Font(name='Times New Roman', bold=True, size=11)
                        row_index += 1
                        if qs.first().name == "Эчкилар":
                            total_out_mshm_head_formula = '='
                            for row in mshm_row:
                                total_out_mshm_head_formula += f'+{get_column_letter(col_index)}{row}'
                            mshm_out_head = sheet.cell(row=row_index, column=col_index,
                                                       value=total_out_mshm_head_formula)
                            mshm_out_head.alignment = align
                            mshm_out_head.font = Font(name='Times New Roman', bold=True, size=11)
                            total_out_mshm_scales_formula = '='
                            for row in mshm_row:
                                total_out_mshm_scales_formula += f'+{get_column_letter(col_index + 1)}{row}'
                            mshm_out_scales = sheet.cell(
                                row=row_index, column=col_index + 1, value=total_out_mshm_scales_formula)
                            mshm_out_scales.alignment = align
                            mshm_out_scales.font = Font(name='Times New Roman', bold=True, size=11)
                            row_index += 1
            col_index += 2
            row_index = 6
        # WRITE EACH OUTPUT TOTAL
        total_output_data = self.chorva_output_raw_sql(category=output_last_pk)
        for total_id, total_data in enumerate(total_output_data):
            category_id = total_data[0]
            output_head_formula = output_head_letters.replace('-', f'{row_index}+')
            cell_total_head = sheet.cell(row=row_index, column=col_index, value=f'=+{output_head_formula[:-1]}')
            cell_total_head.alignment = align
            cell_total_head.font = Font(name='Times New Roman', bold=True)
            output_scales_formula = output_scales_letters.replace('-', f'{row_index}+')
            cell_total_scales = sheet.cell(row=row_index, column=col_index + 1, value=f'=+{output_scales_formula[:-1]}')
            cell_total_scales.alignment = align
            cell_total_scales.font = Font(name='Times New Roman', bold=True)
            row_index += 1
            if total_id == len(total_output_data) - 1 or category_id != total_output_data[total_id + 1][0]:
                count = Animal.objects.filter(category_id=category_id).count()
                begin_row = (row_index - count)
                end_row = row_index - 1
                qs = AnimalCategory.objects.filter(id=category_id)
                if qs.exists():
                    begin_letter = get_column_letter(col_index)
                    total_head_formula = f'=SUM({begin_letter}{begin_row}:{begin_letter}{end_row})'
                    total_category_head = sheet.cell(row=row_index, column=col_index, value=total_head_formula)
                    total_category_head.alignment = align
                    total_category_head.font = Font(name='Times New Roman', bold=True, size=11)

                    end_letter = get_column_letter(col_index + 1)
                    total_head_formula = f'=SUM({end_letter}{begin_row}:{end_letter}{end_row})'
                    total_category_head = sheet.cell(row=row_index, column=col_index + 1, value=total_head_formula)
                    total_category_head.alignment = align
                    total_category_head.font = Font(name='Times New Roman', bold=True, size=11)
                    row_index += 1
                    if qs.first().name == "Эчкилар":
                        total_out_mshm_head_formula = '='
                        for row in mshm_row:
                            total_out_mshm_head_formula += f'+{get_column_letter(col_index)}{row}'
                        total_out_mshm_head = sheet.cell(
                            row=row_index, column=col_index, value=total_out_mshm_head_formula)
                        total_out_mshm_head.alignment = align
                        total_out_mshm_head.font = Font(name='Times New Roman', bold=True, size=11)
                        total_out_mshm_scales_formula = '='
                        for row in mshm_row:
                            total_out_mshm_scales_formula += f'+{get_column_letter(col_index + 1)}{row}'
                        total_out_mshm_scales = sheet.cell(
                            row=row_index, column=col_index + 1, value=total_out_mshm_scales_formula)
                        total_out_mshm_scales.alignment = align
                        total_out_mshm_scales.font = Font(name='Times New Roman', bold=True, size=11)
                        row_index += 1
        col_index += 2
        row_index = 6
        # WRITE EACH FINISH TOTAL
        total_output_data = self.chorva_output_raw_sql(category=output_last_pk)
        for total_id, total_data in enumerate(total_output_data):
            category_id = total_data[0]
            head_col_1 = get_column_letter(self.head_col)
            scales_col_1 = get_column_letter(self.scales_col)
            head_col_2 = get_column_letter(self.head_col + 1)
            scales_col_2 = get_column_letter(self.scales_col + 1)
            t_h_f = f'=+C{row_index}+{head_col_1}{row_index}-{scales_col_1}{row_index}'
            t_s_f = f'=+D{row_index}+{head_col_2}{row_index}-{scales_col_2}{row_index}'
            cell_total_head = sheet.cell(row=row_index, column=col_index, value=t_h_f)
            cell_total_head.alignment = align
            cell_total_head.font = Font(name='Times New Roman', bold=True)
            cell_total_scales = sheet.cell(row=row_index, column=col_index + 1, value=t_s_f)
            cell_total_scales.alignment = align
            cell_total_scales.font = Font(name='Times New Roman', bold=True)
            row_index += 1
            if total_id == len(total_output_data) - 1 or category_id != total_output_data[total_id + 1][0]:
                count = Animal.objects.filter(category_id=category_id).count()
                begin_row = (row_index - count)
                end_row = row_index - 1
                qs = AnimalCategory.objects.filter(id=category_id)
                if qs.exists():
                    begin_letter = get_column_letter(col_index)
                    total_head_formula = f'=SUM({begin_letter}{begin_row}:{begin_letter}{end_row})'
                    total_category_head = sheet.cell(row=row_index, column=col_index, value=total_head_formula)
                    total_category_head.alignment = align
                    total_category_head.font = Font(name='Times New Roman', bold=True, size=11)
                    end_letter = get_column_letter(col_index + 1)
                    total_head_formula = f'=SUM({end_letter}{begin_row}:{end_letter}{end_row})'
                    total_category_head = sheet.cell(row=row_index, column=col_index + 1, value=total_head_formula)
                    total_category_head.alignment = align
                    total_category_head.font = Font(name='Times New Roman', bold=True, size=11)
                    row_index += 1
                    if qs.first().name == "Эчкилар":
                        total_out_mshm_head_formula = '='
                        for row in mshm_row:
                            total_out_mshm_head_formula += f'+{get_column_letter(col_index)}{row}'
                        total_out_mshm_head = sheet.cell(
                            row=row_index, column=col_index, value=total_out_mshm_head_formula)
                        total_out_mshm_head.alignment = align
                        total_out_mshm_head.font = Font(name='Times New Roman', bold=True, size=11)
                        total_out_mshm_scales_formula = '='
                        for row in mshm_row:
                            total_out_mshm_scales_formula += f'+{get_column_letter(col_index + 1)}{row}'
                        total_out_mshm_scales = sheet.cell(
                            row=row_index, column=col_index + 1, value=total_out_mshm_scales_formula)
                        total_out_mshm_scales.alignment = align
                        total_out_mshm_scales.font = Font(name='Times New Roman', bold=True, size=11)
                        row_index += 1
        set_border(sheet=sheet, cell_range=f'A1:{get_column_letter(sheet.max_column)}{row_index - 1}')
        row_index += 3
        write_footer(sheet=sheet, row=row_index)
        sheet.freeze_panes = 'A6'
        wb.save(response)
        return response

    def animal_name_raw_sql(self):
        with connection.cursor() as cursor:
            year = self.end_date[:4]
            query = f"""
            select anml.id, anml.category_id, anml.name, 
            sum(inp_1.amount) - sum(out_2.amount) qoldiq_bosh, 
            sum(inp_1.weight) - sum(out_2.weight) qoldiq_t_vazn
            from animal anml
            left join chorva_input_output inp_1 on inp_1.animal_id = anml.id and inp_1.type = 1 
            and inp_1.department_id = {self.department.pk} and inp_1.date BETWEEN '{year}-01-01' and '{year}-12-31' 
            left join chorva_input_output out_2 on out_2.animal_id = anml.id and out_2.type = 2 
            and out_2.department_id = {self.department.pk} and out_2.date BETWEEN '{year}-01-01' and '{year}-12-31' 
            group by anml.id order by anml.id"""
            cursor.execute(query)
            row = cursor.fetchall()
            return row

    def chorva_input_raw_sql(self, category):
        with connection.cursor() as cursor:
            year = self.end_date[:4]
            query = f"""
            select anml.category_id, sum(inp.amount) bosh, sum(inp.weight) t_vazn
            from animal anml
            left join chorva_input_output inp on inp.type = 1 
            and inp.animal_id = anml.id and inp.department_id = {self.department.pk} and inp.category_id = {category} 
            and inp.date BETWEEN '{year}-01-01' and '{year}-12-31' 
            group by anml.id order by anml.id"""
            cursor.execute(query)
            row = cursor.fetchall()
            return row

    def chorva_output_raw_sql(self, category):
        with connection.cursor() as cursor:
            year = self.end_date[:4]
            query = f"""
            select anml.category_id, sum(out_2.amount) bosh, sum(out_2.weight) t_vazn
            from animal anml
            left join chorva_input_output out_2 on out_2.type = 2 
            and out_2.animal_id = anml.id and out_2.department_id = {self.department.pk} and out_2.category_id = {category} 
            and out_2.date BETWEEN '{year}-01-01' and '{year}-12-31' 
            group by anml.id order by anml.id"""
            cursor.execute(query)
            row = cursor.fetchall()
            return row

