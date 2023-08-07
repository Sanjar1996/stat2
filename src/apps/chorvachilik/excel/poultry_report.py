from datetime import datetime
from django.db import connection
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter, column_index_from_string
from ..models import Chorvachilik
from .utils import set_border, write_footer
from ...accounts.models import Region, Department


class PoultrySheet:
    def __init__(self, user=None, obj=None, start=None, end=None):
        self.user = user
        self.end_date = end
        self.start_date = start
        self.type_name = obj
        self.align = Alignment(wrapText=True, horizontal='center', vertical='center')
        self.meat_letter = 0
        self.for_new = 0

    def generate_poultry_excel_report(self):
        now = datetime.now()
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f"attachment; filename=parranda-amalda-{now.strftime('%d-%m-%Y')}.xlsx"
        wb = Workbook()
        sheet = wb.active
        align = Alignment(wrapText=True, horizontal='center', vertical='center')
        sheet.column_dimensions['A'].width = 5
        sheet.column_dimensions['B'].width = 50
        sheet.row_dimensions[1].height = 50
        sheet.row_dimensions[2].height = 70
        sheet.row_dimensions[3].height = 25
        sheet.row_dimensions[4].height = 80

        sheet.merge_cells('A1:Y1')
        title = sheet['A1']
        title.value = f"Ўрмон хўжалиги давлат қўмитаси тизимидаги хўжалик субъектларида парранда бош сонини " \
                      f"кўпайтириш, тухум ва парранда гўшти ишлаб чиқариш бўйича {self.end_date} ҳолатига МАЪЛУМОТ"
        title.font = Font(name='Times New Roman', bold=True, size=13)
        title.alignment = align
        sheet.merge_cells('A2:A4')
        sheet.merge_cells('B2:B4')
        number = sheet.cell(column=1, row=2, value='Т/р')
        number.alignment = align
        number.font = Font(name='Times New Roman', bold=True, size=11)

        department_title = sheet.cell(column=2, row=2, value='Хўжалик номи')
        department_title.alignment = align
        department_title.font = Font(name='Times New Roman', bold=True, size=15)

        sheet.merge_cells('C2:F2')
        old_residue = sheet.cell(column=3, row=2, value=f'{int(self.end_date[:4]) - 1}-01-01 йил холатига мавжуд')
        old_residue.alignment = align
        old_residue.font = Font(name='Times New Roman', bold=True, size=12)
        sheet.merge_cells('D3:F3')
        total_poultry = sheet.cell(column=3, row=3, value='Жами парранда')
        total_poultry.alignment = align
        sheet.merge_cells('C3:C4')
        also = sheet.cell(column=4, row=3, value='Шундан')
        also.alignment = align

        turkey = sheet.cell(column=4, row=4, value='Курка сони')
        turkey.alignment = align
        roosters_chickens = sheet.cell(column=5, row=4, value='Жами хўроз ва товуқлар, сони')
        roosters_chickens.alignment = align
        egg = sheet.cell(column=6, row=4, value='Шундан тухум берадиган товуқлар сони')
        egg.alignment = align
        egg.font = Font(name='Times New Roman', size=10)

        # get all chorva category with filter ( type__pk=self.type_name.pk, status=Active, parent=None)
        categories = Chorvachilik.objects.filter(type__pk=self.type_name.pk, status=1, parent=None)
        self.write_all_category(sheet=sheet, obj=categories)

        # departments_and_old_poultry -> A, B, C, D, E, F
        d_old_poultry = self.department_and_annual_residue()
        region_id = None
        region_row = []
        article_id = 0
        row_index = 5
        for d_index, d_value in enumerate(d_old_poultry):
            if region_id is None or int(d_value[1]) != region_id:
                article_id = 0
                region_row.append(row_index)
                region_id = self.write_region_data(sheet=sheet, row=row_index, region_pk=d_value[1])
                row_index += 1
            article_id += 1
            department_id = sheet.cell(column=1, row=row_index, value=article_id)
            department_id.alignment = align
            department_id.font = Font(name='Times New Roman', size=12)
            if len(d_value[2]) > 45:
                name = f'{d_value[2][:39]}...'
            else:
                name = d_value[2]
            department_name = sheet.cell(column=2, row=row_index, value=name)
            department_name.alignment = Alignment(wrapText=True, vertical='center')
            department_name.font = Font(name='Times New Roman', size=12)

            total_birds = sheet.cell(column=3, row=row_index, value=f'=+D{row_index}+E{row_index}')
            total_birds.alignment = align

            total_turkey = sheet.cell(column=4, row=row_index, value=d_value[3])
            total_turkey.alignment = align

            total_chickens = sheet.cell(column=4, row=row_index, value=d_value[4])
            total_chickens.alignment = align
            row_index += 1

        #  write all task and actual by category types -> G, H, I, J, K, L, M, N, O, P, K, R, S, T, U, V
        col_index = 7
        region_id = None
        row_index = 5
        all_types = Chorvachilik.objects.filter(type__pk=self.type_name.pk, status=1).exclude(parent=None)
        for types in all_types:
            turkey_chickens = self.get_value_by_chorva_type(parent=types.parent.id)
            for p_index, p_value in enumerate(turkey_chickens):
                if region_id is None or int(p_value[0]) != region_id:
                    qs_region = Region.objects.filter(pk=p_value[0], status=1)
                    count = Department.objects.filter(region_id=p_value[0], status=1).count()
                    begin_row = row_index + 1
                    end_row = row_index + count
                    region_id = qs_region.first().pk
                    if qs_region.exists():
                        for x in range(len(p_value) - 1):
                            letter = get_column_letter(col_index + x)
                            region_total = sheet.cell(
                                column=col_index + x, row=row_index,
                                value=f'=SUM({letter}{begin_row}:{letter}{end_row})')
                            region_total.alignment = align
                            region_total.font = Font(name='Times New Roman', bold=True, size=12)
                        row_index += 1
                turkey_task = sheet.cell(column=col_index, row=row_index, value=p_value[1])
                turkey_task.alignment = align
                turkey_act = sheet.cell(column=col_index + 1, row=row_index, value=p_value[2])
                turkey_act.alignment = align
                chicken_task = sheet.cell(column=col_index + 2, row=row_index, value=p_value[3])
                chicken_task.alignment = align
                chicken_act = sheet.cell(column=col_index + 3, row=row_index, value=p_value[4])
                chicken_act.alignment = align
                row_index += 1
            col_index += 4
            row_index = 5

        # currently available data and poultry profits
        profits = self.get_profit_by_chorva_type()
        region_id = None
        region_row = []
        row_index = 5
        col_index += 1
        letter_l = get_column_letter(self.meat_letter)
        letter_n = get_column_letter(self.meat_letter + 2)
        letter_g = get_column_letter(self.for_new)
        letter_h = get_column_letter(self.for_new + 1)
        letter_i = get_column_letter(self.for_new + 2)
        letter_j = get_column_letter(self.for_new + 3)
        letter_m = get_column_letter(self.meat_letter + 1)
        task_in_total = f'=+C{row_index}+{letter_g}{row_index}+{letter_i}{row_index}-{letter_m}{row_index}'
        turkey_count_total = f'=+D{row_index}+{letter_h}{row_index}'
        also_in_total = f'=+E{row_index}+{letter_j}{row_index}'
        for r_index, r_value in enumerate(profits):
            if region_id is None or int(r_value[0]) != region_id:
                region_row.append(row_index)
                qs_region = Region.objects.filter(pk=r_value[0], status=1)
                count = Department.objects.filter(region_id=r_value[0], status=1).count()
                begin_row = row_index + 1
                end_row = row_index + count
                region_id = qs_region.first().pk
                if qs_region.exists():
                    for x in range(8):
                        letter = get_column_letter(col_index + x)
                        available_data = sheet.cell(
                            column=col_index + x, row=row_index,
                            value=f'=SUM({letter}{begin_row}:{letter}{end_row})')
                        available_data.alignment = align
                        available_data.font = Font(name='Times New Roman', bold=True, size=12)
                    row_index += 1
            cell_1 = sheet.cell(column=col_index, row=row_index, value=f'=+{letter_l}{row_index}')    # =+L5
            cell_1.alignment = align
            cell_2 = sheet.cell(column=col_index + 1, row=row_index, value=f'=+{letter_n}{row_index}')  # =+N5
            cell_2.alignment = align
            cell_3 = sheet.cell(column=col_index + 2, row=row_index, value=task_in_total)  # =+C5+G5+I5-M5
            cell_3.alignment = align
            cell_5 = sheet.cell(column=col_index + 4, row=row_index,
                                value=f'{turkey_count_total}-{cell_1.column_letter}{row_index}')  # =+D5+H5-X5
            cell_5.alignment = align
            cell_6 = sheet.cell(column=col_index + 5, row=row_index,
                                value=f'{also_in_total}-{cell_2.column_letter}{row_index}')  # =+E5+J5-Y5
            cell_6.alignment = align
            formula = f'=+{cell_5.column_letter}{row_index}+{cell_6.column_letter}{row_index}'
            cell_4 = sheet.cell(column=col_index + 3, row=row_index, value=formula)
            cell_4.alignment = align
            cell_7 = sheet.cell(column=col_index + 6, row=row_index, value=f'=+F{row_index}')
            cell_7.alignment = align
            cell_8 = sheet.cell(column=col_index + 7, row=row_index, value=r_value[1])
            cell_8.alignment = align
            row_index += 1
        col_index += 7
        total_title = sheet.cell(row=row_index, column=2, value='Республика бўйича жами')
        total_title.font = Font(name='Times New Roman', bold=True, size=13)
        total_title.alignment = align

        column_letters = []
        for c in range(3, col_index + 1):
            column_letters.append(get_column_letter(c))
        for col in column_letters:
            republic_formula = f'='
            for row in region_row:
                republic_formula += f'+{col}{row}'
            idx = column_index_from_string(col)
            republic_cell = sheet.cell(column=idx, row=row_index, value=republic_formula)
            republic_cell.alignment = align
            republic_cell.font = Font(name='Times New Roman', bold=True, size=12)
        set_border(sheet=sheet, cell_range=f'A1:{get_column_letter(sheet.max_column)}{row_index}')
        row_index += 3
        write_footer(sheet=sheet, row=row_index)
        sheet.freeze_panes = 'A5'
        wb.save(response)
        return response

    def write_region_data(self, sheet, row, region_pk):
        qs_region = Region.objects.filter(pk=region_pk, status=1)
        count = Department.objects.filter(region_id=region_pk, status=1).count()
        begin_row = row + 1
        end_row = row + count
        region_id = qs_region.first().pk
        if qs_region.exists():
            region_name = sheet.cell(column=2, row=row, value=qs_region.first().name)
            region_name.alignment = self.align
            region_name.font = Font(name='Times New Roman', bold=True, size=13)
            c_total_by_reg = sheet.cell(column=3, row=row, value=f'=SUM(C{begin_row}:C{end_row})')
            c_total_by_reg.alignment = self.align
            c_total_by_reg.font = Font(name='Times New Roman', bold=True, size=12)
            d_total_by_reg = sheet.cell(column=4, row=row, value=f'=SUM(D{begin_row}:D{end_row})')
            d_total_by_reg.alignment = self.align
            d_total_by_reg.font = Font(name='Times New Roman', bold=True, size=12)
            e_total_by_reg = sheet.cell(column=5, row=row, value=f'=SUM(E{begin_row}:E{end_row})')
            e_total_by_reg.alignment = self.align
            e_total_by_reg.font = Font(name='Times New Roman', bold=True, size=12)
            return region_id

    def write_all_category(self, sheet, obj):
        begin_col = 7
        for item in obj:
            if str(item.name) == "Гўштга сўйилади (25%)" or "Гўштга сўйилади" in str(item.name):
                self.meat_letter = begin_col + 1
            if str(item.name) == "Янги ташкил этилади (70%)" or "Янги ташкил этилади" in str(item.name):
                self.for_new = begin_col
            sheet.merge_cells(f'{get_column_letter(begin_col)}2:{get_column_letter(begin_col + 3)}2')
            cat_name = sheet.cell(column=begin_col, row=2, value=str(item.name))
            cat_name.alignment = self.align
            cat_name.font = Font(name='Times New Roman', bold=True, size=12)
            sheet.merge_cells(f'{get_column_letter(begin_col)}3:{get_column_letter(begin_col + 1)}3')
            turkey = sheet.cell(column=begin_col, row=3, value='Курка')
            turkey.alignment = self.align
            sheet.merge_cells(f'{get_column_letter(begin_col + 2)}3:{get_column_letter(begin_col + 3)}3')
            chicken = sheet.cell(column=begin_col + 2, row=3, value='Товуқ')
            chicken.alignment = self.align

            task = sheet.cell(column=begin_col, row=4, value='Топшириқ')
            task.alignment = self.align
            task.font = Font(name='Times New Roman', size=10)
            act = sheet.cell(column=begin_col + 1, row=4, value='Амалда')
            act.alignment = self.align
            task = sheet.cell(column=begin_col + 2, row=4, value='Топшириқ')
            task.alignment = self.align
            task.font = Font(name='Times New Roman', size=10)
            act = sheet.cell(column=begin_col + 3, row=4, value='Амалда')
            act.alignment = self.align
            begin_col += 4

        sheet.merge_cells(f'{get_column_letter(begin_col)}2:{get_column_letter(begin_col)}4')
        gift = sheet.cell(column=begin_col, row=2, value='Кам таъминланган оилаларга тарқатилди')
        gift.alignment = self.align
        gift.font = Font(name='Times New Roman', bold=True, size=12)
        begin_col += 1
        sheet.merge_cells(f'{get_column_letter(begin_col)}2:{get_column_letter(begin_col + 1)}3')
        sheet.column_dimensions[f'{get_column_letter(begin_col)}'].width = 7
        sheet.column_dimensions[f'{get_column_letter(begin_col + 1)}'].width = 7
        total_head = sheet.cell(column=begin_col, row=2, value='Жами чиқим бош')
        total_head.alignment = self.align
        total_head.font = Font(name='Times New Roman', bold=True, size=12)
        turkey = sheet.cell(column=begin_col, row=4, value='Курка')
        turkey.alignment = self.align
        chicken = sheet.cell(column=begin_col + 1, row=4, value='Товуқ')
        chicken.alignment = self.align
        begin_col += 2
        sheet.merge_cells(f'{get_column_letter(begin_col)}2:{get_column_letter(begin_col + 4)}2')
        available = sheet.cell(column=begin_col, row=2, value=f'{self.end_date} йил холатига мавжуд')
        available.alignment = self.align
        available.font = Font(name='Times New Roman', bold=True, size=12)

        sheet.merge_cells(f'{get_column_letter(begin_col)}3:{get_column_letter(begin_col + 1)}3')
        total_poultry = sheet.cell(column=begin_col, row=3, value='Жами парранда')
        total_poultry.alignment = self.align
        total_poultry.font = Font(name='Times New Roman', bold=True, size=12)

        task = sheet.cell(column=begin_col, row=4, value='Топшириқ')
        task.alignment = self.align
        task.font = Font(name='Times New Roman', size=10)
        act = sheet.cell(column=begin_col + 1, row=4, value='Амалда')
        act.alignment = self.align
        begin_col += 2

        sheet.merge_cells(f'{get_column_letter(begin_col)}3:{get_column_letter(begin_col + 2)}3')
        total_poultry = sheet.cell(column=begin_col, row=3, value='шундан')
        total_poultry.alignment = self.align
        total_poultry.font = Font(name='Times New Roman', bold=True, size=12)

        turkey_count = sheet.cell(column=begin_col, row=4, value='Куркалар сони')
        turkey_count.alignment = self.align
        begin_col += 1
        roosters_chickens = sheet.cell(column=begin_col, row=4, value='Жами товуқ ва хўрозлар сони')
        roosters_chickens.alignment = self.align
        roosters_chickens.font = Font(name='Times New Roman', size=10)
        begin_col += 1
        egg_chickens = sheet.cell(column=begin_col, row=4, value='Тухум берадиган товуқлар сони')
        egg_chickens.alignment = self.align
        egg_chickens.font = Font(name='Times New Roman', size=10)
        begin_col += 1

        sheet.merge_cells(f'{get_column_letter(begin_col)}2:{get_column_letter(begin_col)}4')
        poultry_profit = sheet.cell(column=begin_col, row=2, value='Паррандачиликдан олинган жами даромад млн сўм')
        poultry_profit.alignment = self.align
        poultry_profit.font = Font(name='Times New Roman', bold=True)
        begin_col += 1

    @staticmethod
    def set_column_width(sheet, index):
        col_1 = get_column_letter(index)
        col_2 = get_column_letter(index + 1)
        sheet.column_dimensions[col_1].width = 6
        sheet.column_dimensions[col_2].width = 6

    def department_and_annual_residue(self):
        with connection.cursor() as cursor:
            old = int(self.end_date[:4]) - 1
            start = f'{old}-01-01'
            query = f"""
            select 
            d.sort, d.region_id, d.name, 
            sum(ch_actual_kurka.amount) jami_kurka, 
            sum(ch_actual_tovuq.amount) jami_tovuq
            from department d
            left join user_departments_departments udd ON udd.department_id = d.id 
            left join user_departments ud on udd.userdepartment_id = ud.id
            left join chorvachilik_types cht on cht.id = {self.type_name.pk} 
            left join chorvachilik_actual ch_actual_kurka ON ch_actual_kurka.chorvachilik_type_id = cht.id  and ch_actual_kurka.amount_type = 2 
            and ch_actual_kurka.date BETWEEN '{start}' AND '{self.start_date}' 
            left join chorvachilik_actual ch_actual_tovuq ON ch_actual_tovuq.chorvachilik_type_id = cht.id  and ch_actual_tovuq.amount_type = 2 
            and ch_actual_tovuq.date BETWEEN '{start}' AND '{self.start_date}' 
            where d.status = 1 and ud.user_id = {self.user.pk} 
            group by d.id order by d.sort"""
            cursor.execute(query)
            row = cursor.fetchall()
            return row

    def get_value_by_chorva_type(self, parent):
        with connection.cursor() as cursor:
            start = self.start_date
            end = self.end_date
            query = f"""
            select d.region_id, 
            sum(ch_plan_kurka.amount) plan_kurka, 
            sum(ch_actual_kurka.amount) actual_kurka, 
            sum(ch_plan_tovuq.amount) plan_tovuq, 
            sum(ch_actual_tovuq.amount) actual_tovuq
            from department d
            left join user_departments_departments udd ON udd.department_id = d.id 
            left join user_departments ud on udd.userdepartment_id = ud.id
            left join chorvachilik ch on ch.parent_id = {parent} 
            left join chorvachilik_plan ch_plan_kurka ON ch_plan_kurka.chorvachilik_id = ch.id and ch_plan_kurka.amount_type = 2 
            and ch_plan_kurka.date BETWEEN '{start}' AND '{end}'
            left join chorvachilik_actual ch_actual_kurka ON ch_actual_kurka.chorvachilik_id = ch.id  and ch_plan_kurka.amount_type = 2
            and ch_actual_kurka.date BETWEEN '{start}' AND '{end}'
            left join chorvachilik_plan ch_plan_tovuq ON ch_plan_tovuq.chorvachilik_id = ch.id  and ch_plan_tovuq.amount_type = 2
            and ch_plan_tovuq.date BETWEEN '{start}' AND '{end}'
            left join chorvachilik_actual ch_actual_tovuq ON ch_actual_tovuq.chorvachilik_id = ch.id  and ch_plan_tovuq.amount_type = 2 
            and ch_actual_tovuq.date BETWEEN '{start}' AND '{end}'
            where d.status = 1 and ud.user_id = {self.user.pk} 
            group by d.id order by d.sort"""
            cursor.execute(query)
            row = cursor.fetchall()
            return row

    def get_profit_by_chorva_type(self):
        with connection.cursor() as cursor:
            start = self.start_date
            end = self.end_date
            query = f"""
            select  d.region_id, 
            sum(ch_actual.profit) profit
            from department d
            left join user_departments_departments udd ON udd.department_id = d.id 
            left join user_departments ud on udd.userdepartment_id = ud.id
            left join chorvachilik_types cht on cht.id = {self.type_name.pk} 
            left join chorvachilik_actual ch_actual on ch_actual.chorvachilik_type_id = cht.id 
            and ch_actual.amount_type = 2 
            and ch_actual.date between '{start}' and '{end}' 
            where d.status = 1 and ud.user_id = {self.user.pk} 
            group by d.id order by d.sort"""
            cursor.execute(query)
            row = cursor.fetchall()
            return row
