from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, Side, Border


def format_date_by_quarter(end):
    year = end[:4]
    next_year = int(year) + 1
    month = int(end[5:-3])
    if month <= 3:
        return dict(start=f'{year}-01-01', end=f'{year}-04-01', quarter=1)
    elif month <= 6:
        return dict(start=f'{year}-04-01', end=f'{year}-07-01', quarter=2)
    elif month <= 9:
        return dict(start=f'{year}-07-01', end=f'{year}-10-01', quarter=3)
    elif month <= 12:
        return dict(start=f'{year}-10-01', end=f'{next_year}-01-01', quarter=4)
    else:
        return dict(start=f'{year}-01-01', end=f'{next_year}-01-01', quarter=0)


def set_border(sheet, cell_range):
    thin = Side(border_style="thin", color="000000")
    for row in sheet[cell_range]:
        for cell in row:
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)


def merge_columns(sheet, begin, end, row):
    begin = get_column_letter(begin)
    end = get_column_letter(end)
    output_title_crd = f'{begin}{row}:{end}{row}'
    sheet.merge_cells(output_title_crd)


def set_column_width(sheet, index):
    col_1 = get_column_letter(index)
    col_2 = get_column_letter(index + 1)
    sheet.column_dimensions[col_1].width = 6
    sheet.column_dimensions[col_2].width = 6


def write_top_panel_footer(sheet, col):
    head_title = sheet.cell(row=5, column=col, value='бош')
    head_title.alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
    head_title.font = Font(name='Times New Roman', bold=True, size=11)

    scales_title = sheet.cell(row=5, column=col + 1, value='т/в')
    scales_title.alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
    scales_title.font = Font(name='Times New Roman', bold=True, size=11)


def chorva_inp_out_category(sheet, category, col, value):
    for item in category:
        merge_columns(sheet=sheet, begin=col, end=col + 1, row=4)
        cell_input_name = sheet.cell(row=4, column=col, value=item.name)
        cell_input_name.alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
        cell_input_name.font = Font(name='Times New Roman', bold=True, size=11)
        set_column_width(sheet=sheet, index=col)
        write_top_panel_footer(sheet=sheet, col=col)
        col += 2
    merge_columns(sheet=sheet, begin=col, end=col + 1, row=4)
    cell_total_input = sheet.cell(row=4, column=col, value=value)
    cell_total_input.alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
    cell_total_input.font = Font(name='Times New Roman', bold=True, size=11)
    set_column_width(sheet=sheet, index=col)
    write_top_panel_footer(sheet=sheet, col=col)
    col += 2
    return col


def write_footer(sheet, row):
    sheet.merge_cells(f'D{row}:F{row}')
    cell_1 = sheet.cell(column=4, row=row, value='Хўжалик директори:')
    cell_1.font = Font(name='Times New Roman', bold=True, size=13)
    sheet.merge_cells(f'D{row + 2}:F{row + 2}')
    cell_2 = sheet.cell(column=4, row=row + 2, value='Бош ўрмончи:')
    cell_2.font = Font(name='Times New Roman', bold=True, size=13)
    sheet.merge_cells(f'D{row + 4}:F{row + 4}')
    cell_3 = sheet.cell(column=4, row=row + 4, value='Бош хисобчи:')
    cell_3.font = Font(name='Times New Roman', bold=True, size=13)
