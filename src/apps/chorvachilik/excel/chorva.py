from datetime import datetime
from django.db import connection
from .db_service import CattleDBQuery
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from ..models import Chorvachilik
from .utils import set_border, write_footer
from ...accounts.models import Region


class ChorvaSheet(CattleDBQuery):
    def __init__(self, user, obj=None, start=None, end=None):
        self.user = user
        self.start_date = start
        self.end_date = end
        self.type_name = obj
        self.align = Alignment(wrapText=True, horizontal='center', vertical='center')

    def generate_cattle_report_by_type(self):
        now = datetime.now()
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f"attachment; filename=chorva_hisoboti-{now.strftime('%d-%m-%Y')}.xlsx"
        wb = Workbook()
        sheet = wb.active

        align = Alignment(wrapText=True, horizontal='center', vertical='center')
        sheet.merge_cells('A1:T1')
        sheet.column_dimensions['A'].width = 5
        sheet.column_dimensions['B'].width = 55
        sheet.column_dimensions['C'].width = 14
        sheet.row_dimensions[1].height = 45  # height size in the first row
        year = f'{self.end_date[:4]} йил'
        title = sheet['A1']
        title.value = f"Ўрмон хўжалиги давлат қўмитаси тизимидаги-давлат ўрмон хўжалигида " \
                      f"{self.type_name.name} бўйича {year} йил ҳолатига МАЬЛУМОТ"
        title.alignment = align
        title.font = Font(name='Times New Roman', bold=True, size=15)
        sheet.merge_cells('A2:A4')
        sheet.merge_cells('B2:B4')
        sheet.row_dimensions[2].height = 40
        sheet.row_dimensions[3].height = 60
        sheet.row_dimensions[4].height = 35

        number = sheet.cell(column=1, row=2, value='№')
        number.alignment = align
        number.font = Font(name='Times New Roman', bold=True, size=12)

        department_name = sheet.cell(column=2, row=2, value='Хўжалик номи')
        department_name.alignment = align
        department_name.font = Font(name='Times New Roman', bold=True, size=15)

        sheet.merge_cells('C2:E3')
        cereals_title = sheet.cell(column=3, row=2, value=self.type_name.name)
        cereals_title.alignment = align
        cereals_title.font = Font(name='Times New Roman', bold=True, italic=True, size=12)

        to_title = sheet.cell(column=3, row=4, value='Топшириқ')
        to_title.alignment = align
        to_title.font = Font(name='Times New Roman', bold=True, italic=True, size=10)

        to_centner_title = sheet.cell(column=4, row=4, value='Амалда')
        to_centner_title.alignment = align
        to_centner_title.font = Font(name='Times New Roman', italic=True, bold=True, size=10)

        ton_title = sheet.cell(column=5, row=4, value='%')
        ton_title.alignment = align
        ton_title.font = Font(name='Times New Roman', italic=True, bold=True, size=10)
        qs_category = Chorvachilik.objects.filter(type__pk=self.type_name.pk)
        print("Qs.....", qs_category, type(qs_category))
        sheet.merge_cells('F2:T2')
        total_ton_title = sheet.cell(column=6, row=2, value='Шу жумладан')
        total_ton_title.alignment = align
        total_ton_title.font = Font(name='Times New Roman', italic=True, bold=True, size=13)
        col_index = 6
        for index, product in enumerate(qs_category):
            start_letter = get_column_letter(col_index)
            end_letter = get_column_letter(col_index + 3)
            distance = f'{start_letter}3:{end_letter}3'
            sheet.merge_cells(distance)
            col = sheet.cell(row=3, column=col_index, value=product.name)
            col.alignment = align
            col.font = Font(name='Times New Roman', bold=True, italic=True, size=13)

            units_1 = sheet.cell(row=4, column=col_index, value='Топшириқ')
            units_1.font = Font(name='Times New Roman', bold=True, italic=True, size=10)
            units_1.alignment = align

            units_2 = sheet.cell(row=4, column=col_index + 1, value='Амалда')
            units_2.font = Font(name='Times New Roman', bold=True, italic=True, size=10)
            units_2.alignment = align

            units_3 = sheet.cell(row=4, column=col_index + 2, value='%')
            units_3.font = Font(name='Times New Roman', bold=True, italic=True, size=10)
            units_3.alignment = align

            units_4 = sheet.cell(row=4, column=col_index + 3, value='даромад')
            units_4.font = Font(name='Times New Roman', bold=True, italic=True, size=10)
            units_4.alignment = align
            col_index += 4
        region_id = None
        row_index = 5
        region_row = []
        task_each_letter_row = '='
        act_each_letter_row = '='
        article_id = 0
        cattle_db_data = self.get_cattle_db_data(
            user=self.user, type_id=self.type_name.pk, qs_category=qs_category, start=self.start_date, end=self.end_date)

        for value in cattle_db_data:
            # print("1......", value)
            if region_id is None or value['region_id'] != region_id:
                article_id = 0
                region_row.append(row_index)
                region_id = self.write_region_data(
                    sheet=sheet, row=row_index, region_id=value['region_id'],
                    count=value['department_count'], data=value['plan_act'])
                row_index += 1
            article_id += 1
            sheet.cell(column=1, row=row_index, value=article_id)
            if len(value['department']) > 50:
                name = f" {value['department'][:45]}..."
            else:
                name = f" {value['department']}"
            department_name = sheet.cell(column=2, row=row_index, value=name)
            department_name.font = Font(name='Times New Roman', size=11)

            col_index = 6
            for ch_index, ch_value in enumerate(value['plan_act']):
                task_cell = sheet.cell(column=col_index, row=row_index, value=ch_value[0])
                task_cell.alignment = align
                if task_cell.column_letter not in task_each_letter_row:
                    task_each_letter_row += f'+{task_cell.column_letter}{row_index}'
                act_cell = sheet.cell(column=col_index + 1, row=row_index, value=ch_value[1])
                act_cell.alignment = align
                if act_cell.column_letter not in act_each_letter_row:
                    act_each_letter_row += f'+{act_cell.column_letter}{row_index}'
                plan_letter = task_cell.column_letter
                act_letter = act_cell.column_letter
                percent_cell = sheet.cell(
                    column=col_index + 2, row=row_index,
                    value=f'=IF(OR({plan_letter}{row_index}={0}, {act_letter}{row_index}={0}),{0},{act_letter}{row_index}*{100}/{plan_letter}{row_index})')
                percent_cell.number_format = '0.0'
                percent_cell.alignment = align
                profit_cell = sheet.cell(column=col_index + 3, row=row_index, value=0)
                profit_cell.alignment = align
                col_index += 4

            department_task = sheet.cell(column=3, row=row_index, value=task_each_letter_row)
            department_task.alignment = align
            department_task.font = Font(name='Times New Roman', size=11)

            department_progress = sheet.cell(column=4, row=row_index, value=act_each_letter_row)
            department_progress.alignment = align
            department_progress.font = Font(name='Times New Roman', size=11)
            department_percent = sheet.cell(
                column=5, row=row_index,
                value=f'=IF(OR(C{row_index}={0}, D{row_index}={0}),{0},D{row_index}*{100}/C{row_index})')
            department_percent.alignment = align
            department_percent.number_format = '0.0'
            department_percent.font = Font(name='Times New Roman', size=11)
            task_each_letter_row = '='
            act_each_letter_row = '='
            row_index += 1

        cell_finish_total_title = sheet.cell(row=row_index, column=2, value='Республика бўйича жами')
        cell_finish_total_title.font = Font(name='Times New Roman', bold=True, size=13)
        cell_finish_total_title.alignment = align

        total_task_formula = '='
        for row in region_row:
            total_task_formula += f'+C{row}'
        total_act_formula = '='
        for row in region_row:
            total_act_formula += f'+D{row}'

        cell_total_task = sheet.cell(row=row_index, column=3, value=total_task_formula)
        cell_total_task.alignment = align
        cell_total_task.font = Font(name='Times New Roman', bold=True, size=13)

        cell_total_act = sheet.cell(row=row_index, column=4, value=total_act_formula)
        cell_total_act.alignment = align
        cell_total_act.font = Font(name='Times New Roman', bold=True, size=13)

        cell_total_percent = sheet.cell(
            row=row_index, column=5,
            value=f'=IF(OR(C{row_index}={0}, D{row_index}={0}),{0},D{row_index}*{100}/C{row_index})')
        cell_total_percent.alignment = align
        cell_total_percent.number_format = '0.0'
        cell_total_percent.font = Font(name='Times New Roman', bold=True, size=13)
        length = qs_category.count() * 4
        for x in range(length):
            col = x + 6
            letter = get_column_letter(col)
            if col > 7 and col % 4 == 0:
                start = get_column_letter(col - 1)
                end = get_column_letter(col - 2)
                percent_cell = sheet.cell(
                    column=col, row=row_index,
                    value=f'=IF(OR({start}{row_index}={0}, {end}{row_index}={0}),{0},{start}{row_index}*{100}/{end}{row_index})')
                percent_cell.alignment = self.align
                percent_cell.number_format = '0.0'
                percent_cell.font = Font(name='Times New Roman', bold=True, size=11)
            else:
                formula = '='
                for row in region_row:
                    formula += f'+{letter}{row}'
                cell = sheet.cell(column=col, row=row_index, value=formula)
                cell.alignment = self.align
                cell.font = Font(name='Times New Roman', bold=True, size=11)

        set_border(sheet=sheet, cell_range=f'A1:{get_column_letter(sheet.max_column)}{row_index}')
        row_index += 3
        write_footer(sheet=sheet, row=row_index)
        sheet.freeze_panes = 'A5'
        wb.save(response)
        return response

    def write_region_data(self, sheet, row, region_id, count, data):
        qs_region = Region.objects.filter(id=region_id, status=1)
        if qs_region.exists():
            begin_row = row + 1
            end_row = row + count
            region_id = qs_region.first().pk
            region_name = sheet.cell(column=2, row=row, value=qs_region.first().name)
            region_name.alignment = self.align
            region_name.font = Font(name='Times New Roman', bold=True, size=12)

            region_task_each_letter_row = '='
            region_act_each_letter_row = '='
            length = len(data) * 4
            for x in range(length):
                col = x + 6
                letter = get_column_letter(col)
                if col > 7 and col % 4 == 0:
                    if get_column_letter(col - 1) not in region_act_each_letter_row:
                        region_act_each_letter_row += f'+{get_column_letter(col - 1)}{row}'
                    if get_column_letter(col - 2) not in region_task_each_letter_row:
                        region_task_each_letter_row += f'+{get_column_letter(col - 2)}{row}'
                    start = get_column_letter(col - 1)
                    end = get_column_letter(col - 2)
                    percent_cell = sheet.cell(
                        column=col, row=row,
                        value=f'=IF(OR({start}{row}={0}, {end}{row}={0}),{0},{start}{row}*{100}/{end}{row})')
                    percent_cell.alignment = self.align
                    percent_cell.number_format = '0.0'
                    percent_cell.font = Font(name='Times New Roman', bold=True, size=11)
                else:
                    task_act_profit = sheet.cell(column=col, row=row, value=f'=SUM({letter}{begin_row}:{letter}{end_row})')
                    task_act_profit.alignment = self.align
                    task_act_profit.font = Font(name='Times New Roman', bold=True, size=11)

            region_task = sheet.cell(column=3, row=row, value=region_task_each_letter_row)
            region_task.alignment = self.align
            region_task.font = Font(name='Times New Roman', bold=True, size=11)
            region_progress = sheet.cell(column=4, row=row, value=region_act_each_letter_row)
            region_progress.alignment = self.align
            region_progress.font = Font(name='Times New Roman', bold=True, size=11)
            region_percent = sheet.cell(
                column=5, row=row, value=f'=IF(OR(C{row}={0}, D{row}={0}),{0},D{row}*{100}/C{row})')
            region_percent.alignment = self.align
            region_percent.number_format = '0.0'
            region_percent.font = Font(name='Times New Roman', bold=True, size=11)
            return region_id

    # def chorva_prepare_raw_sql(self, department):
    #     start = self.start_date  # date format '2021-01-01'
    #     end = self.end_date
    #     with connection.cursor() as cursor:
    #         query = f"""
    #         SELECT
    #         sum(chp.amount) topshiriq,
    #         sum(cha.amount) amalda,
    #         sum(cha.amount)*100/sum(chp.amount) foiz,
    #         sum(cha.profit) foyda
    #         FROM chorvachilik ch
    #         LEFT JOIN chorvachilik_plan chp ON chp.chorvachilik_id = ch.id AND chp.department_id = {department} AND chp.date BETWEEN '{start}' AND '{end}'
    #         LEFT JOIN chorvachilik_actual cha ON cha.chorvachilik_id = ch.id AND cha.department_id = {department} AND cha.date BETWEEN '{start}' AND '{end}'
    #         LEFT JOIN chorvachilik_type cht ON cht.chorvachilik_id = ch.id
    #         where cht.chorvachiliktypes_id = {self.type_name.pk}
    #         GROUP BY ch.id ORDER BY ch.id
    #         """
    #         cursor.execute(query)
    #         row = cursor.fetchall()
    #         return row
