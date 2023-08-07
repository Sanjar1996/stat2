from datetime import datetime
from django.db import connection
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter, column_index_from_string
from ..models import Chorvachilik, ChorvachilikTypes, ChorvachilikPlan, ChorvachilikActual
from ...accounts.models import Department


class CattleSheet:
    def __init__(self, start=None, end=None):
        self.start_date = start
        self.end_date = end

    def generate_mshm_excel_report(self):
        now = datetime.now()
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f"attachment; filename=livestock-{now.strftime('%d-%m-%Y')}.xlsx"
        wb = Workbook()
        sheet = wb.active

        align = Alignment(wrapText=True, horizontal='center', vertical='center')

        if self.start_date[:4] == self.end_date[:4]:
            year = f'{self.start_date[:4]}'
        else:
            year = f'{self.start_date[:4]}-{self.end_date[:4]}'

        sheet.merge_cells('A1:U1')
        sheet.column_dimensions['A'].width = 5
        sheet.column_dimensions['B'].width = 35
        sheet.row_dimensions[1].height = 45  # height size in the first row

        title = sheet['A1']
        title.value = f"Ўрмон хўжалиги давлат қўмитаси тизимидаги - давлат ўрмон хўжалигида" \
                      f" майда шоҳли молларни ривожлантириш бўйича " \
                      f"{self.start_date[:4]} йил {self.start_date[5:]} ҳолатига МАЬЛУМОТ"
        title.alignment = align
        title.font = Font(name='Times New Roman', bold=True, size=13)
        sheet.merge_cells('A2:A3')
        sheet.merge_cells('B2:B3')
        sheet.merge_cells('C2:D2')
        sheet.merge_cells('E2:G2')
        sheet.merge_cells('H2:J2')
        sheet.merge_cells('K2:K3')
        sheet.merge_cells('L2:N2')
        sheet.merge_cells('O2:O3')
        sheet.merge_cells('P2:Q2')
        sheet.merge_cells('R2:S2')
        sheet.merge_cells('T2:T3')
        sheet.merge_cells('U2:U3')
        sheet.row_dimensions[2].height = 40
        sheet.row_dimensions[3].height = 70

        sheet.cell(row=2, column=1, value='№')
        sheet.cell(row=2, column=2, value='Хужалик номи')
        sheet.cell(row=2, column=3, value='01.01.2020 йилда мавжуд')
        sheet.cell(row=3, column=3, value='Жами бош сони')
        sheet.cell(row=3, column=4, value='шундан Совлиқ-лар')
        sheet.cell(row=2, column=5, value='2020 йилда қўзи-улоқ олиш (дона)')
        sheet.cell(row=3, column=5, value='Топшириқ')
        sheet.cell(row=3, column=6, value='Амалда')
        sheet.cell(row=3, column=7, value='Фоиз')
        sheet.cell(row=2, column=8, value='Янгидан ташкил этиш')
        sheet.cell(row=3, column=8, value='Топшириқ')
        sheet.cell(row=3, column=9, value='Амалда')
        sheet.cell(row=3, column=10, value='Фоиз')
        sheet.cell(row=2, column=11, value='Гўштга сўйиш (бош)')
        sheet.cell(row=2, column=12, value='Гўшт ишлаб чиқариш (цн)')
        sheet.cell(row=3, column=12, value='Топшириқ')
        sheet.cell(row=3, column=13, value='Амалда')
        sheet.cell(row=3, column=14, value='Фоиз')
        sheet.cell(row=2, column=15, value='Жами чиқим (дона)')
        sheet.cell(row=2, column=16, value='Жун ишлаб чиқариш (кг)')
        sheet.cell(row=3, column=16, value='Топшириқ')
        sheet.cell(row=3, column=17, value='Амалда')
        sheet.cell(row=2, column=18, value='31.12.2020 йил ҳолатига')
        sheet.cell(row=3, column=18, value='Жами бош сони')
        sheet.cell(row=3, column=19, value='шундан Совлиқ-лар')
        sheet.cell(row=2, column=20, value='2018 йилга нисбаттан жами бош сонини  ўсиши, (%)')
        sheet.cell(row=2, column=21, value='МШМ дан олинган жами даромад (МЛН. Сум.)')

        departments = Department.objects.all().order_by('region_id')
        row_index = 4
        region = 0
        for index, department_value in enumerate(departments):
            region = department_value.region

            department_id = sheet.cell(column=1, row=row_index, value=department_value.id)
            department_id.font = Font(name='Times New Roman', size=11)
            department_id.alignment = align

            department_name = sheet.cell(column=2, row=row_index, value=department_value.name)
            department_name.alignment = align
            department_name.font = Font(name='Times New Roman', size=11)
            row_index += 1

            # agricultural_crops_list = self.agricultural_product_prepare_raw_sql(department=department_value.pk)
            if index == departments.count() - 1 or region.pk != departments[index + 1].region.pk:
                region_name = sheet.cell(column=2, row=row_index, value=region.name)
                region_name.alignment = align
                region_name.font = Font(name='Times New Roman', bold=True, size=13)
                row_index += 1

        def set_border(ws, cell_range):
            thin = Side(border_style="thin", color="000000")
            for row in ws[cell_range]:
                for cell in row:
                    cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                    cell.alignment = align
        set_border(sheet, f'A1:{get_column_letter(sheet.max_column)}{sheet.max_row}')

        # sheet.freeze_panes = 'A4'
        wb.save(response)
        return response

    def agricultural_product_prepare_raw_sql(self, department):
        start = self.start_date  # date format '2021-01-01'
        end = self.end_date
        with connection.cursor() as cursor:
            query = f""""""
            cursor.execute(query)
            row = cursor.fetchall()
            return row
